"""
Cell Manager Tool for hiel_excel_mcp.

Provides comprehensive cell manipulation operations including row/column insertion/deletion,
cell formatting, cell information retrieval, cell updates, and cell clearing.
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List

from ..core.base_tool import BaseTool, operation_route, OperationResponse, create_success_response, create_error_response
from ..core.workbook_context import workbook_context

# Import existing functionality
import sys
import os

# Add the src directory to the path to import existing modules
src_path = os.path.join(os.path.dirname(__file__), '..', '..', 'src')
if src_path not in sys.path:
    sys.path.insert(0, src_path)

    insert_row, insert_cols, delete_rows, delete_cols
)

logger = logging.getLogger(__name__)


class CellManager(BaseTool):
    """
    Comprehensive cell manipulation tool.
    
    Handles cell-level operations including row/column insertion/deletion,
    cell formatting, information retrieval, updates, and clearing.
    """
    
    def get_tool_name(self) -> str:
        return "cell_manager"
    
    def get_tool_description(self) -> str:
        return "Comprehensive cell manipulation and formatting operations tool"
    
    @operation_route(
        name="insert_rows",
        description="Insert one or more rows at the specified position",
        required_params=["filepath", "sheet_name", "start_row"],
        optional_params=["count"]
    )
    def insert_rows(self, filepath: str, sheet_name: str, start_row: int, 
                   count: int = 1, **kwargs) -> OperationResponse:
        """
        Insert one or more rows at the specified position.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_row: Row number where to start inserting (1-based)
            count: Number of rows to insert (default: 1)
            
        Returns:
            OperationResponse with insertion results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Insert rows using existing functionality
            result = insert_row(validated_path, sheet_name, start_row, count)
            
            return create_success_response(
                operation="insert_rows",
                message=f"Inserted {count} row(s) starting at row {start_row} in worksheet '{sheet_name}'",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "start_row": start_row,
                    "count": count,
                    "message": result.get("message", "")
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to insert rows in {sheet_name} of {filepath}: {e}")
            return create_error_response("insert_rows", e)
    
    @operation_route(
        name="insert_columns",
        description="Insert one or more columns at the specified position",
        required_params=["filepath", "sheet_name", "start_col"],
        optional_params=["count"]
    )
    def insert_columns(self, filepath: str, sheet_name: str, start_col: int, 
                      count: int = 1, **kwargs) -> OperationResponse:
        """
        Insert one or more columns at the specified position.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_col: Column number where to start inserting (1-based)
            count: Number of columns to insert (default: 1)
            
        Returns:
            OperationResponse with insertion results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Insert columns using existing functionality
            result = insert_cols(validated_path, sheet_name, start_col, count)
            
            return create_success_response(
                operation="insert_columns",
                message=f"Inserted {count} column(s) starting at column {start_col} in worksheet '{sheet_name}'",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "start_col": start_col,
                    "count": count,
                    "message": result.get("message", "")
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to insert columns in {sheet_name} of {filepath}: {e}")
            return create_error_response("insert_columns", e)
    
    @operation_route(
        name="delete_rows",
        description="Delete one or more rows starting at the specified position",
        required_params=["filepath", "sheet_name", "start_row"],
        optional_params=["count"]
    )
    def delete_rows(self, filepath: str, sheet_name: str, start_row: int, 
                   count: int = 1, **kwargs) -> OperationResponse:
        """
        Delete one or more rows starting at the specified position.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_row: Row number where to start deleting (1-based)
            count: Number of rows to delete (default: 1)
            
        Returns:
            OperationResponse with deletion results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Delete rows using existing functionality
            result = delete_rows(validated_path, sheet_name, start_row, count)
            
            return create_success_response(
                operation="delete_rows",
                message=f"Deleted {count} row(s) starting at row {start_row} in worksheet '{sheet_name}'",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "start_row": start_row,
                    "count": count,
                    "message": result.get("message", "")
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to delete rows in {sheet_name} of {filepath}: {e}")
            return create_error_response("delete_rows", e)
    
    @operation_route(
        name="delete_columns",
        description="Delete one or more columns starting at the specified position",
        required_params=["filepath", "sheet_name", "start_col"],
        optional_params=["count"]
    )
    def delete_columns(self, filepath: str, sheet_name: str, start_col: int, 
                      count: int = 1, **kwargs) -> OperationResponse:
        """
        Delete one or more columns starting at the specified position.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_col: Column number where to start deleting (1-based)
            count: Number of columns to delete (default: 1)
            
        Returns:
            OperationResponse with deletion results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Delete columns using existing functionality
            result = delete_cols(validated_path, sheet_name, start_col, count)
            
            return create_success_response(
                operation="delete_columns",
                message=f"Deleted {count} column(s) starting at column {start_col} in worksheet '{sheet_name}'",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "start_col": start_col,
                    "count": count,
                    "message": result.get("message", "")
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to delete columns in {sheet_name} of {filepath}: {e}")
            return create_error_response("delete_columns", e)
    
    @operation_route(
        name="format_range",
        description="Apply formatting to a range of cells",
        required_params=["filepath", "sheet_name", "start_cell"],
        optional_params=["end_cell", "bold", "italic", "underline", "font_size", "font_color", 
                        "bg_color", "border_style", "border_color", "number_format", 
                        "alignment", "wrap_text", "merge_cells"]
    )
    def format_range(self, filepath: str, sheet_name: str, start_cell: str, 
                    end_cell: Optional[str] = None, **kwargs) -> OperationResponse:
        """
        Apply formatting to a range of cells.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_cell: Starting cell address (e.g., "A1")
            end_cell: Ending cell address (optional)
            **kwargs: Formatting options (bold, italic, font_size, colors, etc.)
            
        Returns:
            OperationResponse with formatting results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Validate cell references
            if not validate_cell_reference(start_cell):
                raise ValueError(f"Invalid start cell reference: {start_cell}")
            
            if end_cell and not validate_cell_reference(end_cell):
                raise ValueError(f"Invalid end cell reference: {end_cell}")
            
            # Apply formatting using existing functionality
            result = format_range(
                validated_path, sheet_name, start_cell, end_cell, **kwargs
            )
            
            range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
            
            return create_success_response(
                operation="format_range",
                message=f"Applied formatting to range {range_str} in worksheet '{sheet_name}'",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "range": range_str,
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "formatting_applied": {k: v for k, v in kwargs.items() if v is not None},
                    "message": result.get("message", "")
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to format range in {sheet_name} of {filepath}: {e}")
            return create_error_response("format_range", e)
    
    @operation_route(
        name="get_cell_info",
        description="Get detailed information about a cell or range of cells",
        required_params=["filepath", "sheet_name", "start_cell"],
        optional_params=["end_cell", "include_validation"]
    )
    def get_cell_info(self, filepath: str, sheet_name: str, start_cell: str, 
                     end_cell: Optional[str] = None, include_validation: bool = True, 
                     **kwargs) -> OperationResponse:
        """
        Get detailed information about a cell or range of cells.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_cell: Starting cell address (e.g., "A1")
            end_cell: Ending cell address (optional)
            include_validation: Whether to include validation information
            
        Returns:
            OperationResponse with cell information
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Validate cell references
            if not validate_cell_reference(start_cell):
                raise ValueError(f"Invalid start cell reference: {start_cell}")
            
            if end_cell and not validate_cell_reference(end_cell):
                raise ValueError(f"Invalid end cell reference: {end_cell}")
            
            # Get cell information using existing functionality
            cell_info = read_excel_range_with_metadata(
                validated_path, sheet_name, start_cell, end_cell, include_validation
            )
            
            range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
            
            return create_success_response(
                operation="get_cell_info",
                message=f"Retrieved information for {len(cell_info.get('cells', []))} cells in range {range_str}",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "range": range_str,
                    "cell_info": cell_info
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to get cell info from {sheet_name} of {filepath}: {e}")
            return create_error_response("get_cell_info", e)
    
    @operation_route(
        name="update_cell",
        description="Update the value of a specific cell",
        required_params=["filepath", "sheet_name", "cell_address", "value"]
    )
    def update_cell(self, filepath: str, sheet_name: str, cell_address: str, 
                   value: Any, **kwargs) -> OperationResponse:
        """
        Update the value of a specific cell.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            cell_address: Cell address (e.g., "A1")
            value: New value for the cell
            
        Returns:
            OperationResponse with update results
        """
        try:
            from openpyxl import load_workbook
            
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Validate cell reference
            if not validate_cell_reference(cell_address):
                raise ValueError(f"Invalid cell reference: {cell_address}")
            
            # Load workbook and update cell
            wb = load_workbook(validated_path)
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # Get old value for comparison
            old_value = ws[cell_address].value
            
            # Update cell value
            ws[cell_address] = value
            
            # Save workbook
            wb.save(validated_path)
            wb.close()
            
            return create_success_response(
                operation="update_cell",
                message=f"Updated cell {cell_address} in worksheet '{sheet_name}'",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "cell_address": cell_address,
                    "old_value": old_value,
                    "new_value": value
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to update cell {cell_address} in {sheet_name} of {filepath}: {e}")
            return create_error_response("update_cell", e)
    
    @operation_route(
        name="clear_cells",
        description="Clear the contents of a cell or range of cells",
        required_params=["filepath", "sheet_name", "start_cell"],
        optional_params=["end_cell", "clear_formatting"]
    )
    def clear_cells(self, filepath: str, sheet_name: str, start_cell: str, 
                   end_cell: Optional[str] = None, clear_formatting: bool = False, 
                   **kwargs) -> OperationResponse:
        """
        Clear the contents of a cell or range of cells.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_cell: Starting cell address (e.g., "A1")
            end_cell: Ending cell address (optional)
            clear_formatting: Whether to also clear formatting (default: False)
            
        Returns:
            OperationResponse with clearing results
        """
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, Border, PatternFill
            
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Validate cell references
            if not validate_cell_reference(start_cell):
                raise ValueError(f"Invalid start cell reference: {start_cell}")
            
            if end_cell and not validate_cell_reference(end_cell):
                raise ValueError(f"Invalid end cell reference: {end_cell}")
            
            # Load workbook and clear cells
            wb = load_workbook(validated_path)
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # Parse range
            start_row, start_col, end_row, end_col = parse_cell_range(start_cell, end_cell)
            
            if end_row is None:
                end_row = start_row
            if end_col is None:
                end_col = start_col
            
            cells_cleared = 0
            
            # Clear cells in range
            for row in range(start_row, end_row + 1):
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row, column=col)
                    
                    # Clear value
                    if cell.value is not None:
                        cell.value = None
                        cells_cleared += 1
                    
                    # Clear formatting if requested
                    if clear_formatting:
                        cell.font = Font()
                        cell.border = Border()
                        cell.fill = PatternFill()
                        cell.number_format = "General"
                        cell.alignment = None
            
            # Save workbook
            wb.save(validated_path)
            wb.close()
            
            range_str = f"{start_cell}:{end_cell}" if end_cell else start_cell
            
            return create_success_response(
                operation="clear_cells",
                message=f"Cleared {cells_cleared} cells in range {range_str}",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "range": range_str,
                    "cells_cleared": cells_cleared,
                    "formatting_cleared": clear_formatting
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to clear cells in {sheet_name} of {filepath}: {e}")
            return create_error_response("clear_cells", e)


# Create global instance
cell_manager = CellManager()


def cell_manager_tool(operation: str, **kwargs) -> str:
    """
    MCP tool function for cell management operations.
    
    Args:
        operation: The operation to perform
        **kwargs: Operation-specific parameters
        
    Returns:
        JSON string with operation results
    """
    try:
        response = cell_manager.execute_operation(operation, **kwargs)
        return response.to_json()
    except Exception as e:
        logger.error(f"Unexpected error in cell_manager_tool: {e}", exc_info=True)
        error_response = create_error_response(operation, e)
        return error_response.to_json()