"""
Data Manager Tool for hiel_excel_mcp.

Provides comprehensive data operations including reading, writing, copying,
deleting, validating ranges, and data transformations.
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List

from ..core.base_tool import BaseTool, operation_route, OperationResponse, create_success_response, create_error_response
from ..core.workbook_context import workbook_context

# Import existing functionality
import sys
import os

# Add the src directory to the path to import existing modules
src_path = os.path.join(os.path.dirname(__file__), '..', '..', 'src')
if src_path not in sys.path:
    sys.path.insert(0, src_path)


logger = logging.getLogger(__name__)


class DataManager(BaseTool):
    """
    Comprehensive data management tool.
    
    Handles data operations including reading, writing, copying, deleting,
    range validation, and data transformations.
    """
    
    def get_tool_name(self) -> str:
        return "data_manager"
    
    def get_tool_description(self) -> str:
        return "Comprehensive data operations and transformation management tool"
    
    @operation_route(
        name="read",
        description="Read data from Excel range with optional metadata",
        required_params=["filepath", "sheet_name"],
        optional_params=["start_cell", "end_cell", "include_metadata", "include_validation"]
    )
    def read(self, filepath: str, sheet_name: str, start_cell: str = "A1", 
             end_cell: Optional[str] = None, include_metadata: bool = False,
             include_validation: bool = False, **kwargs) -> OperationResponse:
        """
        Read data from Excel range with optional metadata.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_cell: Starting cell address (default: A1)
            end_cell: Ending cell address (optional)
            include_metadata: Whether to include cell metadata
            include_validation: Whether to include validation information
            
        Returns:
            OperationResponse with read data
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            if include_metadata:
                # Use metadata-aware reading
                result = read_excel_range_with_metadata(
                    validated_path, sheet_name, start_cell, end_cell, include_validation
                )
                message = f"Read {len(result.get('cells', []))} cells with metadata from {sheet_name}"
            else:
                # Use simple data reading
                data = read_excel_range(validated_path, sheet_name, start_cell, end_cell)
                result = {
                    "data": data,
                    "sheet_name": sheet_name,
                    "range": f"{start_cell}:{end_cell}" if end_cell else start_cell,
                    "row_count": len(data),
                    "column_count": len(data[0]) if data else 0
                }
                message = f"Read {len(data)} rows from {sheet_name}"
            
            return create_success_response(
                operation="read",
                message=message,
                data=result,
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to read data from {filepath}: {e}")
            return create_error_response("read", e)
    
    @operation_route(
        name="write",
        description="Write data to Excel sheet",
        required_params=["filepath", "data"],
        optional_params=["sheet_name", "start_cell"]
    )
    def write(self, filepath: str, data: List[List], sheet_name: Optional[str] = None,
              start_cell: str = "A1", **kwargs) -> OperationResponse:
        """
        Write data to Excel sheet.
        
        Args:
            filepath: Path to the Excel file
            data: Data to write as list of lists
            sheet_name: Name of the worksheet (uses active if None)
            start_cell: Starting cell address (default: A1)
            
        Returns:
            OperationResponse with write results
        """
        try:
            from openpyxl import Workbook, load_workbook
            import os
            
            # Validate path (allow creation for write operations)
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=True)
            
            if not data:
                raise ValueError("No data provided to write")
            
            # Check if file exists, create if not
            if not os.path.exists(validated_path):
                # Create new workbook
                wb = Workbook()
                if sheet_name:
                    wb.active.title = sheet_name
                    active_sheet_name = sheet_name
                else:
                    active_sheet_name = wb.active.title
                wb.save(validated_path)
                wb.close()
            
            # Now use existing write_data function
            result = write_data(validated_path, sheet_name, data, start_cell)
            
            return create_success_response(
                operation="write",
                message=f"Data written successfully to {result.get('active_sheet', sheet_name)}",
                data={
                    "filepath": validated_path,
                    "sheet_name": result.get("active_sheet", sheet_name),
                    "start_cell": start_cell,
                    "rows_written": len(data),
                    "columns_written": len(data[0]) if data else 0,
                    "message": result.get("message", "")
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to write data to {filepath}: {e}")
            return create_error_response("write", e)
    
    @operation_route(
        name="copy_range",
        description="Copy data from one range to another",
        required_params=["filepath", "source_sheet", "source_range", "dest_sheet", "dest_start_cell"],
        optional_params=["dest_filepath"]
    )
    def copy_range(self, filepath: str, source_sheet: str, source_range: str,
                   dest_sheet: str, dest_start_cell: str, 
                   dest_filepath: Optional[str] = None, **kwargs) -> OperationResponse:
        """
        Copy data from one range to another.
        
        Args:
            filepath: Path to source Excel file
            source_sheet: Source worksheet name
            source_range: Source range (e.g., "A1:C10" or just "A1")
            dest_sheet: Destination worksheet name
            dest_start_cell: Starting cell for destination
            dest_filepath: Destination file path (uses source if None)
            
        Returns:
            OperationResponse with copy results
        """
        try:
            # Validate source path
            validated_source_path, source_warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Parse source range
            if ':' in source_range:
                start_cell, end_cell = source_range.split(':', 1)
            else:
                start_cell, end_cell = source_range, None
            
            # Read source data
            source_data = read_excel_range(validated_source_path, source_sheet, start_cell, end_cell)
            
            if not source_data:
                return create_success_response(
                    operation="copy_range",
                    message="No data found in source range",
                    data={
                        "source_filepath": validated_source_path,
                        "source_sheet": source_sheet,
                        "source_range": source_range,
                        "rows_copied": 0
                    },
                    warnings=["No data found in source range"]
                )
            
            # Determine destination file
            dest_file = dest_filepath or validated_source_path
            validated_dest_path, dest_warnings = PathValidator.validate_path(dest_file, allow_create=True)
            
            # Write to destination
            write_result = write_data(validated_dest_path, dest_sheet, source_data, dest_start_cell)
            
            # Combine warnings
            all_warnings = []
            if source_warnings:
                all_warnings.extend(source_warnings)
            if dest_warnings:
                all_warnings.extend(dest_warnings)
            
            return create_success_response(
                operation="copy_range",
                message=f"Copied {len(source_data)} rows from {source_sheet} to {dest_sheet}",
                data={
                    "source_filepath": validated_source_path,
                    "source_sheet": source_sheet,
                    "source_range": source_range,
                    "dest_filepath": validated_dest_path,
                    "dest_sheet": write_result.get("active_sheet", dest_sheet),
                    "dest_start_cell": dest_start_cell,
                    "rows_copied": len(source_data),
                    "columns_copied": len(source_data[0]) if source_data else 0
                },
                warnings=all_warnings if all_warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to copy range: {e}")
            return create_error_response("copy_range", e)
    
    @operation_route(
        name="delete_range",
        description="Delete data from a range (clear cell values)",
        required_params=["filepath", "sheet_name", "range_ref"]
    )
    def delete_range(self, filepath: str, sheet_name: str, range_ref: str, **kwargs) -> OperationResponse:
        """
        Delete data from a range by clearing cell values.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            range_ref: Range to clear (e.g., "A1:C10")
            
        Returns:
            OperationResponse with deletion results
        """
        try:
            from openpyxl import load_workbook
            
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Load workbook and clear range
            wb = load_workbook(validated_path)
            
            if sheet_name not in wb.sheetnames:
                raise ValueError(f"Sheet '{sheet_name}' not found")
            
            ws = wb[sheet_name]
            
            # Parse range and clear cells
            if ':' in range_ref:
                # Range of cells
                cell_range = ws[range_ref]
                cells_cleared = 0
                
                # Handle both 2D ranges and single row/column ranges
                if hasattr(cell_range, '__iter__'):
                    if hasattr(cell_range[0], '__iter__'):
                        # 2D range
                        for row in cell_range:
                            for cell in row:
                                if cell.value is not None:
                                    cell.value = None
                                    cells_cleared += 1
                    else:
                        # 1D range (single row or column)
                        for cell in cell_range:
                            if cell.value is not None:
                                cell.value = None
                                cells_cleared += 1
                else:
                    # Single cell
                    if cell_range.value is not None:
                        cell_range.value = None
                        cells_cleared = 1
            else:
                # Single cell
                cell = ws[range_ref]
                cells_cleared = 1 if cell.value is not None else 0
                cell.value = None
            
            # Save workbook
            wb.save(validated_path)
            wb.close()
            
            return create_success_response(
                operation="delete_range",
                message=f"Cleared {cells_cleared} cells in range {range_ref}",
                data={
                    "filepath": validated_path,
                    "sheet_name": sheet_name,
                    "range": range_ref,
                    "cells_cleared": cells_cleared
                },
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to delete range {range_ref}: {e}")
            return create_error_response("delete_range", e)
    
    @operation_route(
        name="validate_range",
        description="Validate if a range exists and get range information",
        required_params=["filepath", "sheet_name", "start_cell"],
        optional_params=["end_cell"]
    )
    def validate_range(self, filepath: str, sheet_name: str, start_cell: str,
                      end_cell: Optional[str] = None, **kwargs) -> OperationResponse:
        """
        Validate if a range exists and get range information.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            start_cell: Starting cell address
            end_cell: Ending cell address (optional)
            
        Returns:
            OperationResponse with validation results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Validate range using existing functionality
            validation_result = validate_range_in_sheet_operation(
                validated_path, sheet_name, start_cell, end_cell
            )
            
            return create_success_response(
                operation="validate_range",
                message=validation_result["message"],
                data=validation_result,
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to validate range: {e}")
            return create_error_response("validate_range", e)
    
    @operation_route(
        name="transform",
        description="Apply data transformations to a range",
        required_params=["filepath", "sheet_name", "range_ref", "transformations"],
        optional_params=["output_filepath"]
    )
    def transform(self, filepath: str, sheet_name: str, range_ref: str,
                  transformations: List[Dict[str, Any]], 
                  output_filepath: Optional[str] = None, **kwargs) -> OperationResponse:
        """
        Apply data transformations to a range.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of the worksheet
            range_ref: Range to transform (e.g., "A1:C10")
            transformations: List of transformation configurations
            output_filepath: Optional output file path
            
        Returns:
            OperationResponse with transformation results
        """
        try:
            # Validate path
            validated_path, warnings = PathValidator.validate_path(filepath, allow_create=False)
            
            # Apply transformations using existing functionality
            result = DataTransformer.transform_range(
                validated_path, sheet_name, range_ref, transformations, output_filepath
            )
            
            return create_success_response(
                operation="transform",
                message=result["message"],
                data=result,
                warnings=warnings if warnings else None
            )
            
        except Exception as e:
            logger.error(f"Failed to transform data: {e}")
            return create_error_response("transform", e)


# Create global instance
data_manager = DataManager()


def data_manager_tool(operation: str, **kwargs) -> str:
    """
    MCP tool function for data management operations.
    
    Args:
        operation: The operation to perform
        **kwargs: Operation-specific parameters
        
    Returns:
        JSON string with operation results
    """
    try:
        response = data_manager.execute_operation(operation, **kwargs)
        return response.to_json()
    except Exception as e:
        logger.error(f"Unexpected error in data_manager_tool: {e}", exc_info=True)
        error_response = create_error_response(operation, e)
        return error_response.to_json()