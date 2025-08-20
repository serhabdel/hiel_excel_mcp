"""
Workbook Manager Tool for hiel_excel_mcp.

Provides comprehensive workbook lifecycle management including creation,
metadata retrieval, safety validation, backup operations, and path management.
"""

import json
import logging
from pathlib import Path
from typing import Dict, Any, Optional, List
from openpyxl import Workbook, load_workbook
import os
import shutil
from datetime import datetime

from ..core.base_tool import BaseTool, operation_route, OperationResponse, create_success_response, create_error_response
from ..core.workbook_context import WorkbookContext
from ..core.utils import ExcelMCPUtils, ValidationError
from ..core.error_handler import handle_excel_errors
from ..core.config import config

logger = logging.getLogger(__name__)


class WorkbookManager(BaseTool):
    """
    Comprehensive workbook management tool.
    
    Handles workbook lifecycle operations including creation, metadata retrieval,
    safety validation, backup management, and path operations.
    """
    
    def get_tool_name(self) -> str:
        return "workbook_manager"
    
    def get_tool_description(self) -> str:
        return "Comprehensive workbook lifecycle and metadata management tool"
    
    @handle_excel_errors("create_workbook", "workbook_manager")
    def create_workbook(self, filepath: str, sheet_names: Optional[List[str]] = None) -> Dict[str, Any]:
        """
        Create a new Excel workbook with optional custom sheet names.
        
        Args:
            filepath: Path where to create the workbook
            sheet_names: Optional list of sheet names to create
            
        Returns:
            Dict with creation results
        """
        # Validate and get safe path
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=True)
        
        # Create workbook
        wb = Workbook()
        
        # Set up sheets
        if sheet_names:
            # Remove default sheet if we're creating custom ones
            if wb.worksheets:
                wb.remove(wb.active)
            
            # Create specified sheets
            for sheet_name in sheet_names:
                sanitized_name = ExcelMCPUtils.sanitize_filename(sheet_name)
                wb.create_sheet(title=sanitized_name)
        else:
            # Just rename the default sheet
            wb.active.title = "Sheet1"
        
        # Save workbook
        wb.save(validated_path)
        wb.close()
        
        return {
            "success": True,
            "message": f"Workbook created successfully: {validated_path}",
            "filepath": validated_path,
            "sheet_names": sheet_names or ["Sheet1"],
            "warnings": warnings
        }
    
    @handle_excel_errors("open_workbook", "workbook_manager")
    def open_workbook(self, filepath: str, read_only: bool = False) -> Dict[str, Any]:
        """
        Open an existing Excel workbook and get basic information.
        
        Args:
            filepath: Path to the Excel file
            read_only: Whether to open in read-only mode
            
        Returns:
            Dict with workbook information
        """
        # Validate path
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        with WorkbookContext(validated_path, read_only=read_only) as wb_context:
            wb = wb_context.workbook
            
            # Collect basic information
            sheet_info = []
            for sheet in wb.worksheets:
                sheet_info.append({
                    "name": sheet.title,
                    "visible": sheet.sheet_state == 'visible',
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column
                })
            
            return {
                "success": True,
                "message": f"Workbook opened successfully: {validated_path}",
                "filepath": validated_path,
                "read_only": read_only,
                "sheet_count": len(wb.worksheets),
                "sheets": sheet_info,
                "warnings": warnings
            }
    
    @handle_excel_errors("save_workbook", "workbook_manager")
    def save_workbook(self, filepath: str, new_filepath: Optional[str] = None) -> Dict[str, Any]:
        """
        Save workbook to current or new location.
        
        Args:
            filepath: Current filepath of workbook
            new_filepath: Optional new filepath to save to
            
        Returns:
            Dict with save results
        """
        # Validate current path
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        save_path = validated_path
        if new_filepath:
            save_path, new_warnings = ExcelMCPUtils.validate_filepath(new_filepath, allow_create=True)
            warnings.extend(new_warnings)
        
        with WorkbookContext(validated_path) as wb_context:
            wb_context.workbook.save(save_path)
            
            return {
                "success": True,
                "message": f"Workbook saved successfully to: {save_path}",
                "original_filepath": validated_path,
                "saved_filepath": save_path,
                "warnings": warnings
            }
    
    @handle_excel_errors("get_workbook_info", "workbook_manager")
    def get_workbook_info(self, filepath: str) -> Dict[str, Any]:
        """
        Get comprehensive information about a workbook.
        
        Args:
            filepath: Path to the Excel file
            
        Returns:
            Dict with detailed workbook information
        """
        # Validate path
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        with WorkbookContext(validated_path, read_only=True) as wb_context:
            wb = wb_context.workbook
            
            # Get file statistics
            file_stats = os.stat(validated_path)
            
            # Collect comprehensive information
            sheets_info = []
            total_cells = 0
            
            for sheet in wb.worksheets:
                # Calculate used range
                used_range = f"A1:{sheet.calculate_dimension()}" if sheet.max_row > 0 else "A1:A1"
                cell_count = sheet.max_row * sheet.max_column if sheet.max_row > 0 else 0
                total_cells += cell_count
                
                sheet_info = {
                    "name": sheet.title,
                    "state": sheet.sheet_state,
                    "used_range": used_range,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "cell_count": cell_count,
                    "has_filters": bool(sheet.auto_filter.ref) if sheet.auto_filter else False,
                    "protection": {
                        "protected": sheet.protection.sheet,
                        "password_protected": bool(sheet.protection.password)
                    }
                }
                sheets_info.append(sheet_info)
            
            # Workbook-level information
            workbook_info = {
                "success": True,
                "filepath": validated_path,
                "file_size": file_stats.st_size,
                "modified_time": datetime.fromtimestamp(file_stats.st_mtime).isoformat(),
                "sheet_count": len(wb.worksheets),
                "total_estimated_cells": total_cells,
                "active_sheet": wb.active.title if wb.active else None,
                "properties": {
                    "creator": getattr(wb.properties, 'creator', None),
                    "title": getattr(wb.properties, 'title', None),
                    "description": getattr(wb.properties, 'description', None),
                    "created": getattr(wb.properties, 'created', None),
                    "modified": getattr(wb.properties, 'modified', None)
                },
                "sheets": sheets_info,
                "warnings": warnings
            }
            
            return workbook_info
    
    @handle_excel_errors("add_worksheet", "workbook_manager")
    def add_worksheet(self, filepath: str, sheet_name: str, index: Optional[int] = None) -> Dict[str, Any]:
        """
        Add a new worksheet to the workbook.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name for the new worksheet
            index: Optional position to insert the sheet
            
        Returns:
            Dict with operation results
        """
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        # Sanitize sheet name
        sanitized_name = ExcelMCPUtils.sanitize_filename(sheet_name)
        
        with WorkbookContext(validated_path) as wb_context:
            wb = wb_context.workbook
            
            # Check if sheet name already exists
            if sanitized_name in [ws.title for ws in wb.worksheets]:
                raise ValidationError(f"Sheet '{sanitized_name}' already exists")
            
            # Create new worksheet
            if index is not None:
                ws = wb.create_sheet(title=sanitized_name, index=index)
            else:
                ws = wb.create_sheet(title=sanitized_name)
            
            return {
                "success": True,
                "message": f"Worksheet '{sanitized_name}' added successfully",
                "sheet_name": sanitized_name,
                "original_name": sheet_name,
                "position": wb.worksheets.index(ws),
                "total_sheets": len(wb.worksheets),
                "warnings": warnings
            }
    
    @handle_excel_errors("delete_worksheet", "workbook_manager")
    def delete_worksheet(self, filepath: str, sheet_name: str) -> Dict[str, Any]:
        """
        Delete a worksheet from the workbook.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of worksheet to delete
            
        Returns:
            Dict with operation results
        """
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        with WorkbookContext(validated_path) as wb_context:
            wb = wb_context.workbook
            
            # Find worksheet
            worksheet = None
            for ws in wb.worksheets:
                if ws.title == sheet_name:
                    worksheet = ws
                    break
            
            if not worksheet:
                raise ValidationError(f"Worksheet '{sheet_name}' not found")
            
            # Prevent deletion of last worksheet
            if len(wb.worksheets) <= 1:
                raise ValidationError("Cannot delete the last worksheet in the workbook")
            
            # Remove worksheet
            wb.remove(worksheet)
            
            return {
                "success": True,
                "message": f"Worksheet '{sheet_name}' deleted successfully",
                "deleted_sheet": sheet_name,
                "remaining_sheets": [ws.title for ws in wb.worksheets],
                "warnings": warnings
            }
    
    @handle_excel_errors("rename_worksheet", "workbook_manager")
    def rename_worksheet(self, filepath: str, old_name: str, new_name: str) -> Dict[str, Any]:
        """
        Rename a worksheet.
        
        Args:
            filepath: Path to the Excel file
            old_name: Current name of worksheet
            new_name: New name for worksheet
            
        Returns:
            Dict with operation results
        """
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        # Sanitize new name
        sanitized_name = ExcelMCPUtils.sanitize_filename(new_name)
        
        with WorkbookContext(validated_path) as wb_context:
            wb = wb_context.workbook
            
            # Find worksheet
            worksheet = None
            for ws in wb.worksheets:
                if ws.title == old_name:
                    worksheet = ws
                    break
            
            if not worksheet:
                raise ValidationError(f"Worksheet '{old_name}' not found")
            
            # Check if new name already exists
            if sanitized_name in [ws.title for ws in wb.worksheets if ws != worksheet]:
                raise ValidationError(f"Worksheet '{sanitized_name}' already exists")
            
            # Rename worksheet
            worksheet.title = sanitized_name
            
            return {
                "success": True,
                "message": f"Worksheet renamed from '{old_name}' to '{sanitized_name}'",
                "old_name": old_name,
                "new_name": sanitized_name,
                "warnings": warnings
            }
    
    @handle_excel_errors("copy_worksheet", "workbook_manager")
    def copy_worksheet(self, filepath: str, sheet_name: str, new_sheet_name: str) -> Dict[str, Any]:
        """
        Copy a worksheet within the workbook.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of worksheet to copy
            new_sheet_name: Name for the copied worksheet
            
        Returns:
            Dict with operation results
        """
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        # Sanitize new name
        sanitized_name = ExcelMCPUtils.sanitize_filename(new_sheet_name)
        
        with WorkbookContext(validated_path) as wb_context:
            wb = wb_context.workbook
            
            # Find source worksheet
            source_worksheet = None
            for ws in wb.worksheets:
                if ws.title == sheet_name:
                    source_worksheet = ws
                    break
            
            if not source_worksheet:
                raise ValidationError(f"Worksheet '{sheet_name}' not found")
            
            # Check if new name already exists
            if sanitized_name in [ws.title for ws in wb.worksheets]:
                raise ValidationError(f"Worksheet '{sanitized_name}' already exists")
            
            # Copy worksheet
            new_worksheet = wb.copy_worksheet(source_worksheet)
            new_worksheet.title = sanitized_name
            
            return {
                "success": True,
                "message": f"Worksheet '{sheet_name}' copied to '{sanitized_name}'",
                "source_sheet": sheet_name,
                "new_sheet": sanitized_name,
                "position": wb.worksheets.index(new_worksheet),
                "warnings": warnings
            }
    
    @handle_excel_errors("get_worksheet_info", "workbook_manager")
    def get_worksheet_info(self, filepath: str, sheet_name: str) -> Dict[str, Any]:
        """
        Get detailed information about a specific worksheet.
        
        Args:
            filepath: Path to the Excel file
            sheet_name: Name of worksheet
            
        Returns:
            Dict with worksheet information
        """
        validated_path, warnings = ExcelMCPUtils.validate_filepath(filepath, allow_create=False)
        
        with WorkbookContext(validated_path, read_only=True) as wb_context:
            wb = wb_context.workbook
            
            # Find worksheet
            worksheet = None
            for ws in wb.worksheets:
                if ws.title == sheet_name:
                    worksheet = ws
                    break
            
            if not worksheet:
                raise ValidationError(f"Worksheet '{sheet_name}' not found")
            
            # Collect detailed information
            worksheet_info = {
                "success": True,
                "sheet_name": sheet_name,
                "state": worksheet.sheet_state,
                "dimensions": {
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "used_range": worksheet.calculate_dimension() if worksheet.max_row > 0 else "A1:A1"
                },
                "protection": {
                    "protected": worksheet.protection.sheet,
                    "password_protected": bool(worksheet.protection.password),
                    "format_cells": worksheet.protection.formatCells,
                    "format_columns": worksheet.protection.formatColumns,
                    "format_rows": worksheet.protection.formatRows,
                    "insert_columns": worksheet.protection.insertColumns,
                    "insert_rows": worksheet.protection.insertRows,
                    "insert_hyperlinks": worksheet.protection.insertHyperlinks,
                    "delete_columns": worksheet.protection.deleteColumns,
                    "delete_rows": worksheet.protection.deleteRows,
                    "select_locked_cells": worksheet.protection.selectLockedCells,
                    "sort": worksheet.protection.sort,
                    "auto_filter": worksheet.protection.autoFilter,
                    "pivot_tables": worksheet.protection.pivotTables,
                    "select_unlocked_cells": worksheet.protection.selectUnlockedCells
                },
                "filters": {
                    "has_auto_filter": bool(worksheet.auto_filter.ref) if worksheet.auto_filter else False,
                    "filter_range": worksheet.auto_filter.ref if worksheet.auto_filter and worksheet.auto_filter.ref else None
                },
                "merged_cells": [str(range_) for range_ in worksheet.merged_cells.ranges],
                "position": wb.worksheets.index(worksheet),
                "warnings": warnings
            }
            
            return worksheet_info