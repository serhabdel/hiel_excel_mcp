"""
Import/Export Manager Tool for hiel_excel_mcp.

Provides comprehensive import and export operations for Excel files,
including CSV import/export, HTML/JSON export, batch operations, and preview functionality.
"""

import json
import csv
from pathlib import Path
from typing import Dict, Any, List, Optional, Union
import logging

from ..core.base_tool import BaseTool, operation_route, OperationResponse, create_success_response, create_error_response
from ..core.workbook_context import workbook_context

logger = logging.getLogger(__name__)


class ImportExportManager(BaseTool):
    """
    Tool for managing import and export operations.
    
    Provides operations for:
    - CSV import/export with preview
    - HTML/JSON export
    - Batch import operations
    - Data format conversion
    """
    
    def get_tool_name(self) -> str:
        """Get the name of this tool."""
        return "import_export_manager"
    
    def get_tool_description(self) -> str:
        """Get the description of this tool."""
        return "Manage import and export operations for Excel files including CSV, HTML, JSON formats"
    
    @operation_route(
        name="import_csv",
        description="Import CSV data into Excel worksheet",
        required_params=["csv_path", "excel_path", "sheet_name"],
        optional_params=["start_cell", "has_header", "delimiter", "encoding", "max_rows", "skip_rows"]
    )
    def import_csv(
        self,
        csv_path: str,
        excel_path: str,
        sheet_name: str,
        start_cell: str = "A1",
        has_header: bool = True,
        delimiter: Optional[str] = None,
        encoding: str = "utf-8",
        max_rows: Optional[int] = None,
        skip_rows: int = 0
    ) -> OperationResponse:
        """
        Import CSV data into Excel worksheet.
        
        Args:
            csv_path: Path to CSV file
            excel_path: Path to Excel file (will be created if doesn't exist)
            sheet_name: Name of worksheet to import into
            start_cell: Starting cell for import (default A1)
            has_header: Whether CSV has header row (default True)
            delimiter: CSV delimiter (auto-detected if None)
            encoding: File encoding (default utf-8)
            max_rows: Maximum number of rows to import (None for all)
            skip_rows: Number of rows to skip from beginning
            
        Returns:
            OperationResponse with import results
        """
        try:
            # Validate CSV file exists
            csv_file = Path(csv_path)
            if not csv_file.exists():
                return create_error_response("import_csv", FileNotFoundError(f"CSV file not found: {csv_path}"))
            
            # Read CSV data
            csv_data = self._read_csv_file(
                csv_path, delimiter, encoding, has_header, max_rows, skip_rows
            )
            
            if not csv_data:
                return create_error_response("import_csv", ValueError("No data found in CSV file"))
            
            # Import into Excel
            with workbook_context(excel_path) as wb:
                # Create sheet if it doesn't exist
                if sheet_name not in wb.sheetnames:
                    wb.create_sheet(sheet_name)
                
                ws = wb[sheet_name]
                
                # Parse start cell
                from openpyxl.utils import coordinate_from_string
                col, row = coordinate_from_string(start_cell)
                start_col = ws[start_cell].column
                start_row = ws[start_cell].row
                
                # Write data to worksheet
                for row_idx, csv_row in enumerate(csv_data):
                    for col_idx, value in enumerate(csv_row):
                        cell = ws.cell(
                            row=start_row + row_idx,
                            column=start_col + col_idx,
                            value=value
                        )
            
            return create_success_response(
                "import_csv",
                f"Successfully imported {len(csv_data)} rows from CSV to Excel",
                {
                    "csv_path": csv_path,
                    "excel_path": excel_path,
                    "sheet_name": sheet_name,
                    "start_cell": start_cell,
                    "rows_imported": len(csv_data),
                    "columns_imported": len(csv_data[0]) if csv_data else 0,
                    "has_header": has_header
                }
            )
            
        except Exception as e:
            logger.error(f"Error importing CSV: {e}")
            return create_error_response("import_csv", e)
    
    @operation_route(
        name="export_csv",
        description="Export Excel data to CSV file",
        required_params=["excel_path", "sheet_name", "csv_path"],
        optional_params=["start_cell", "end_cell", "include_header", "delimiter", "encoding"]
    )
    def export_csv(
        self,
        excel_path: str,
        sheet_name: str,
        csv_path: str,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
        include_header: bool = True,
        delimiter: str = ",",
        encoding: str = "utf-8"
    ) -> OperationResponse:
        """
        Export Excel data to CSV file.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of worksheet to export from
            csv_path: Path to output CSV file
            start_cell: Starting cell for export (default A1)
            end_cell: Ending cell for export (None for auto-detect)
            include_header: Whether to include header row (default True)
            delimiter: CSV delimiter (default comma)
            encoding: File encoding (default utf-8)
            
        Returns:
            OperationResponse with export results
        """
        try:
            # Validate Excel file exists
            excel_file = Path(excel_path)
            if not excel_file.exists():
                return create_error_response("export_csv", FileNotFoundError(f"Excel file not found: {excel_path}"))
            
            # Read Excel data
            with workbook_context(excel_path, read_only=True) as wb:
                if sheet_name not in wb.sheetnames:
                    return create_error_response("export_csv", ValueError(f"Sheet '{sheet_name}' not found"))
                
                ws = wb[sheet_name]
                
                # Determine range
                if end_cell:
                    cell_range = ws[f"{start_cell}:{end_cell}"]
                else:
                    # Auto-detect used range
                    max_row = ws.max_row
                    max_col = ws.max_column
                    from openpyxl.utils import coordinate_from_string, get_column_letter
                    start_col, start_row = coordinate_from_string(start_cell)
                    end_cell = f"{get_column_letter(max_col)}{max_row}"
                    cell_range = ws[f"{start_cell}:{end_cell}"]
                
                # Extract data
                excel_data = []
                if hasattr(cell_range, '__iter__') and not isinstance(cell_range, str):
                    if hasattr(cell_range[0], '__iter__'):
                        # Multiple rows
                        for row in cell_range:
                            row_data = [cell.value for cell in row]
                            excel_data.append(row_data)
                    else:
                        # Single row
                        row_data = [cell.value for cell in cell_range]
                        excel_data.append(row_data)
                else:
                    # Single cell
                    excel_data.append([cell_range.value])
            
            if not excel_data:
                return create_error_response("export_csv", ValueError("No data found in specified range"))
            
            # Write CSV file
            self._write_csv_file(csv_path, excel_data, delimiter, encoding)
            
            return create_success_response(
                "export_csv",
                f"Successfully exported {len(excel_data)} rows to CSV",
                {
                    "excel_path": excel_path,
                    "sheet_name": sheet_name,
                    "csv_path": csv_path,
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "rows_exported": len(excel_data),
                    "columns_exported": len(excel_data[0]) if excel_data else 0,
                    "delimiter": delimiter,
                    "encoding": encoding
                }
            )
            
        except Exception as e:
            logger.error(f"Error exporting CSV: {e}")
            return create_error_response("export_csv", e)
    
    @operation_route(
        name="preview_csv",
        description="Preview CSV file contents before import",
        required_params=["csv_path"],
        optional_params=["num_rows", "delimiter", "encoding", "has_header"]
    )
    def preview_csv(
        self,
        csv_path: str,
        num_rows: int = 5,
        delimiter: Optional[str] = None,
        encoding: str = "utf-8",
        has_header: bool = True
    ) -> OperationResponse:
        """
        Preview CSV file contents before import.
        
        Args:
            csv_path: Path to CSV file
            num_rows: Number of rows to preview (default 5)
            delimiter: CSV delimiter (auto-detected if None)
            encoding: File encoding (default utf-8)
            has_header: Whether CSV has header row (default True)
            
        Returns:
            OperationResponse with preview data
        """
        try:
            # Validate CSV file exists
            csv_file = Path(csv_path)
            if not csv_file.exists():
                return create_error_response("preview_csv", FileNotFoundError(f"CSV file not found: {csv_path}"))
            
            # Read preview data
            preview_data = self._read_csv_file(
                csv_path, delimiter, encoding, has_header, max_rows=num_rows
            )
            
            # Get file statistics
            file_stats = self._get_csv_file_stats(csv_path, encoding)
            
            # Detect delimiter if not specified
            detected_delimiter = delimiter
            if delimiter is None:
                detected_delimiter = self._detect_csv_delimiter(csv_path, encoding)
            
            return create_success_response(
                "preview_csv",
                f"Preview of {len(preview_data)} rows from CSV file",
                {
                    "csv_path": csv_path,
                    "preview_data": preview_data,
                    "rows_previewed": len(preview_data),
                    "columns_detected": len(preview_data[0]) if preview_data else 0,
                    "detected_delimiter": detected_delimiter,
                    "encoding": encoding,
                    "has_header": has_header,
                    "file_stats": file_stats
                }
            )
            
        except Exception as e:
            logger.error(f"Error previewing CSV: {e}")
            return create_error_response("preview_csv", e)
    
    @operation_route(
        name="export_html",
        description="Export Excel data to HTML format",
        required_params=["excel_path", "sheet_name", "html_path"],
        optional_params=["start_cell", "end_cell", "include_styles", "table_title"]
    )
    def export_html(
        self,
        excel_path: str,
        sheet_name: str,
        html_path: str,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
        include_styles: bool = True,
        table_title: Optional[str] = None
    ) -> OperationResponse:
        """
        Export Excel data to HTML format.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of worksheet to export from
            html_path: Path to output HTML file
            start_cell: Starting cell for export (default A1)
            end_cell: Ending cell for export (None for auto-detect)
            include_styles: Whether to include CSS styling (default True)
            table_title: Optional title for the HTML table
            
        Returns:
            OperationResponse with export results
        """
        try:
            # Validate Excel file exists
            excel_file = Path(excel_path)
            if not excel_file.exists():
                return create_error_response("export_html", FileNotFoundError(f"Excel file not found: {excel_path}"))
            
            # Read Excel data (same logic as CSV export)
            with workbook_context(excel_path, read_only=True) as wb:
                if sheet_name not in wb.sheetnames:
                    return create_error_response("export_html", ValueError(f"Sheet '{sheet_name}' not found"))
                
                ws = wb[sheet_name]
                
                # Determine range
                if end_cell:
                    cell_range = ws[f"{start_cell}:{end_cell}"]
                else:
                    max_row = ws.max_row
                    max_col = ws.max_column
                    from openpyxl.utils import coordinate_from_string, get_column_letter
                    start_col, start_row = coordinate_from_string(start_cell)
                    end_cell = f"{get_column_letter(max_col)}{max_row}"
                    cell_range = ws[f"{start_cell}:{end_cell}"]
                
                # Extract data
                excel_data = []
                if hasattr(cell_range, '__iter__') and not isinstance(cell_range, str):
                    if hasattr(cell_range[0], '__iter__'):
                        for row in cell_range:
                            row_data = [cell.value for cell in row]
                            excel_data.append(row_data)
                    else:
                        row_data = [cell.value for cell in cell_range]
                        excel_data.append(row_data)
                else:
                    excel_data.append([cell_range.value])
            
            if not excel_data:
                return create_error_response("export_html", ValueError("No data found in specified range"))
            
            # Generate HTML
            html_content = self._generate_html_table(
                excel_data, include_styles, table_title or f"{sheet_name} Export"
            )
            
            # Write HTML file
            html_file = Path(html_path)
            html_file.parent.mkdir(parents=True, exist_ok=True)
            html_file.write_text(html_content, encoding='utf-8')
            
            return create_success_response(
                "export_html",
                f"Successfully exported {len(excel_data)} rows to HTML",
                {
                    "excel_path": excel_path,
                    "sheet_name": sheet_name,
                    "html_path": html_path,
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "rows_exported": len(excel_data),
                    "columns_exported": len(excel_data[0]) if excel_data else 0,
                    "include_styles": include_styles,
                    "table_title": table_title
                }
            )
            
        except Exception as e:
            logger.error(f"Error exporting HTML: {e}")
            return create_error_response("export_html", e)
    
    @operation_route(
        name="export_json",
        description="Export Excel data to JSON format",
        required_params=["excel_path", "sheet_name", "json_path"],
        optional_params=["start_cell", "end_cell", "include_headers", "format_style"]
    )
    def export_json(
        self,
        excel_path: str,
        sheet_name: str,
        json_path: str,
        start_cell: str = "A1",
        end_cell: Optional[str] = None,
        include_headers: bool = True,
        format_style: str = "records"
    ) -> OperationResponse:
        """
        Export Excel data to JSON format.
        
        Args:
            excel_path: Path to Excel file
            sheet_name: Name of worksheet to export from
            json_path: Path to output JSON file
            start_cell: Starting cell for export (default A1)
            end_cell: Ending cell for export (None for auto-detect)
            include_headers: Whether first row contains headers (default True)
            format_style: JSON format style - 'records', 'values', or 'index' (default 'records')
            
        Returns:
            OperationResponse with export results
        """
        try:
            # Validate format style
            if format_style not in ['records', 'values', 'index']:
                return create_error_response("export_json", ValueError(f"Invalid format_style: {format_style}"))
            
            # Validate Excel file exists
            excel_file = Path(excel_path)
            if not excel_file.exists():
                return create_error_response("export_json", FileNotFoundError(f"Excel file not found: {excel_path}"))
            
            # Read Excel data (same logic as other exports)
            with workbook_context(excel_path, read_only=True) as wb:
                if sheet_name not in wb.sheetnames:
                    return create_error_response("export_json", ValueError(f"Sheet '{sheet_name}' not found"))
                
                ws = wb[sheet_name]
                
                # Determine range
                if end_cell:
                    cell_range = ws[f"{start_cell}:{end_cell}"]
                else:
                    max_row = ws.max_row
                    max_col = ws.max_column
                    from openpyxl.utils import coordinate_from_string, get_column_letter
                    start_col, start_row = coordinate_from_string(start_cell)
                    end_cell = f"{get_column_letter(max_col)}{max_row}"
                    cell_range = ws[f"{start_cell}:{end_cell}"]
                
                # Extract data
                excel_data = []
                if hasattr(cell_range, '__iter__') and not isinstance(cell_range, str):
                    if hasattr(cell_range[0], '__iter__'):
                        for row in cell_range:
                            row_data = [cell.value for cell in row]
                            excel_data.append(row_data)
                    else:
                        row_data = [cell.value for cell in cell_range]
                        excel_data.append(row_data)
                else:
                    excel_data.append([cell_range.value])
            
            if not excel_data:
                return create_error_response("export_json", ValueError("No data found in specified range"))
            
            # Format data based on style
            json_data = self._format_json_data(excel_data, format_style, include_headers)
            
            # Write JSON file
            json_file = Path(json_path)
            json_file.parent.mkdir(parents=True, exist_ok=True)
            with json_file.open('w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, default=str, ensure_ascii=False)
            
            return create_success_response(
                "export_json",
                f"Successfully exported {len(excel_data)} rows to JSON",
                {
                    "excel_path": excel_path,
                    "sheet_name": sheet_name,
                    "json_path": json_path,
                    "start_cell": start_cell,
                    "end_cell": end_cell,
                    "rows_exported": len(excel_data),
                    "columns_exported": len(excel_data[0]) if excel_data else 0,
                    "format_style": format_style,
                    "include_headers": include_headers
                }
            )
            
        except Exception as e:
            logger.error(f"Error exporting JSON: {e}")
            return create_error_response("export_json", e)
    
    @operation_route(
        name="batch_import",
        description="Import multiple CSV files into Excel workbook",
        required_params=["csv_files", "excel_path"],
        optional_params=["sheet_prefix", "has_headers", "delimiter", "encoding", "max_rows_per_file"]
    )
    def batch_import(
        self,
        csv_files: List[str],
        excel_path: str,
        sheet_prefix: str = "Sheet",
        has_headers: bool = True,
        delimiter: Optional[str] = None,
        encoding: str = "utf-8",
        max_rows_per_file: Optional[int] = None
    ) -> OperationResponse:
        """
        Import multiple CSV files into Excel workbook.
        
        Args:
            csv_files: List of CSV file paths to import
            excel_path: Path to Excel file (will be created if doesn't exist)
            sheet_prefix: Prefix for sheet names (default "Sheet")
            has_headers: Whether CSV files have header rows (default True)
            delimiter: CSV delimiter (auto-detected if None)
            encoding: File encoding (default utf-8)
            max_rows_per_file: Maximum rows to import per file (None for all)
            
        Returns:
            OperationResponse with batch import results
        """
        try:
            if not csv_files:
                return create_error_response("batch_import", ValueError("No CSV files provided"))
            
            import_results = []
            total_rows = 0
            
            with workbook_context(excel_path) as wb:
                # Remove default sheet if it exists and is empty
                if "Sheet" in wb.sheetnames and len(wb.sheetnames) == 1:
                    wb.remove(wb["Sheet"])
                
                for i, csv_path in enumerate(csv_files):
                    try:
                        # Validate CSV file
                        csv_file = Path(csv_path)
                        if not csv_file.exists():
                            import_results.append({
                                "csv_path": csv_path,
                                "success": False,
                                "error": f"File not found: {csv_path}"
                            })
                            continue
                        
                        # Generate sheet name
                        sheet_name = f"{sheet_prefix}_{i+1}_{csv_file.stem}"
                        
                        # Read CSV data
                        csv_data = self._read_csv_file(
                            csv_path, delimiter, encoding, has_headers, max_rows_per_file
                        )
                        
                        if not csv_data:
                            import_results.append({
                                "csv_path": csv_path,
                                "sheet_name": sheet_name,
                                "success": False,
                                "error": "No data found in CSV file"
                            })
                            continue
                        
                        # Create sheet and import data
                        ws = wb.create_sheet(sheet_name)
                        
                        for row_idx, csv_row in enumerate(csv_data):
                            for col_idx, value in enumerate(csv_row):
                                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
                        
                        rows_imported = len(csv_data)
                        total_rows += rows_imported
                        
                        import_results.append({
                            "csv_path": csv_path,
                            "sheet_name": sheet_name,
                            "success": True,
                            "rows_imported": rows_imported,
                            "columns_imported": len(csv_data[0]) if csv_data else 0
                        })
                        
                    except Exception as e:
                        import_results.append({
                            "csv_path": csv_path,
                            "success": False,
                            "error": str(e)
                        })
            
            successful_imports = sum(1 for result in import_results if result["success"])
            
            return create_success_response(
                "batch_import",
                f"Batch import completed: {successful_imports}/{len(csv_files)} files imported successfully",
                {
                    "excel_path": excel_path,
                    "total_files": len(csv_files),
                    "successful_imports": successful_imports,
                    "failed_imports": len(csv_files) - successful_imports,
                    "total_rows_imported": total_rows,
                    "import_results": import_results
                }
            )
            
        except Exception as e:
            logger.error(f"Error in batch import: {e}")
            return create_error_response("batch_import", e)
    
    # Helper methods
    
    def _read_csv_file(
        self,
        csv_path: str,
        delimiter: Optional[str],
        encoding: str,
        has_header: bool,
        max_rows: Optional[int] = None,
        skip_rows: int = 0
    ) -> List[List[Any]]:
        """Read CSV file and return data as list of lists."""
        # Detect delimiter if not provided
        if delimiter is None:
            delimiter = self._detect_csv_delimiter(csv_path, encoding)
        
        data = []
        with open(csv_path, 'r', encoding=encoding, newline='') as f:
            reader = csv.reader(f, delimiter=delimiter)
            
            # Skip initial rows
            for _ in range(skip_rows):
                try:
                    next(reader)
                except StopIteration:
                    break
            
            row_count = 0
            for row in reader:
                # Clean cell values
                cleaned_row = [self._clean_cell_value(cell) for cell in row]
                data.append(cleaned_row)
                
                row_count += 1
                if max_rows and row_count >= max_rows:
                    break
        
        return data
    
    def _write_csv_file(
        self,
        csv_path: str,
        data: List[List[Any]],
        delimiter: str,
        encoding: str
    ):
        """Write data to CSV file."""
        csv_file = Path(csv_path)
        csv_file.parent.mkdir(parents=True, exist_ok=True)
        
        with open(csv_path, 'w', encoding=encoding, newline='') as f:
            writer = csv.writer(f, delimiter=delimiter)
            for row in data:
                # Convert values to strings
                string_row = [str(cell) if cell is not None else '' for cell in row]
                writer.writerow(string_row)
    
    def _detect_csv_delimiter(self, csv_path: str, encoding: str) -> str:
        """Detect CSV delimiter from file sample."""
        with open(csv_path, 'r', encoding=encoding) as f:
            sample = f.read(8192)  # Read first 8KB
        
        try:
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample, delimiters=',;\t|:').delimiter
            return delimiter
        except csv.Error:
            return ','  # Default to comma
    
    def _clean_cell_value(self, value: Any) -> Any:
        """Clean and normalize cell value."""
        if value is None:
            return ''
        
        if isinstance(value, str):
            value = value.strip()
            if value.lower() in ['null', 'none', 'n/a', 'na', '#n/a']:
                return ''
        
        return value
    
    def _get_csv_file_stats(self, csv_path: str, encoding: str) -> Dict[str, Any]:
        """Get statistics about CSV file."""
        try:
            csv_file = Path(csv_path)
            file_size = csv_file.stat().st_size
            
            with open(csv_path, 'r', encoding=encoding) as f:
                line_count = sum(1 for _ in f)
            
            return {
                "file_size_bytes": file_size,
                "file_size_mb": round(file_size / (1024 * 1024), 2),
                "total_lines": line_count,
                "estimated_rows": line_count - 1 if line_count > 0 else 0
            }
        except Exception:
            return {
                "file_size_bytes": 0,
                "file_size_mb": 0,
                "total_lines": 0,
                "estimated_rows": 0
            }
    
    def _generate_html_table(
        self,
        data: List[List[Any]],
        include_styles: bool,
        title: str = "Excel Export"
    ) -> str:
        """Generate HTML table from data."""
        html_parts = ['<!DOCTYPE html>', '<html>', '<head>']
        html_parts.append(f'<title>{title}</title>')
        html_parts.append('<meta charset="utf-8">')
        
        if include_styles:
            html_parts.append('''
<style>
    body { font-family: Arial, sans-serif; margin: 20px; }
    table { border-collapse: collapse; width: 100%; }
    th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
    th { background-color: #f2f2f2; font-weight: bold; }
    tr:nth-child(even) { background-color: #f9f9f9; }
    .number { text-align: right; }
</style>
            ''')
        
        html_parts.extend(['</head>', '<body>'])
        html_parts.append(f'<h1>{title}</h1>')
        html_parts.append('<table>')
        
        # Add table data
        for row_idx, row in enumerate(data):
            if row_idx == 0:
                # First row as header
                html_parts.append('<thead><tr>')
                for cell in row:
                    html_parts.append(f'<th>{self._escape_html(str(cell or ""))}</th>')
                html_parts.append('</tr></thead><tbody>')
            else:
                html_parts.append('<tr>')
                for cell in row:
                    cell_value = str(cell or "")
                    cell_class = 'number' if self._is_number(cell) else ''
                    class_attr = f' class="{cell_class}"' if cell_class else ''
                    html_parts.append(f'<td{class_attr}>{self._escape_html(cell_value)}</td>')
                html_parts.append('</tr>')
        
        if len(data) > 1:
            html_parts.append('</tbody>')
        
        html_parts.extend(['</table>', '</body>', '</html>'])
        return '\n'.join(html_parts)
    
    def _escape_html(self, text: str) -> str:
        """Escape HTML special characters."""
        return (text.replace('&', '&amp;')
                   .replace('<', '&lt;')
                   .replace('>', '&gt;')
                   .replace('"', '&quot;')
                   .replace("'", '&#x27;'))
    
    def _is_number(self, value: Any) -> bool:
        """Check if value is a number."""
        return isinstance(value, (int, float)) and not isinstance(value, bool)
    
    def _format_json_data(
        self,
        data: List[List[Any]],
        format_style: str,
        include_headers: bool
    ) -> Union[List[Dict], List[List], Dict]:
        """Format data for JSON export based on style."""
        if format_style == 'records' and include_headers and data:
            headers = data[0]
            records = []
            for row in data[1:]:
                record = {}
                for i, header in enumerate(headers):
                    value = row[i] if i < len(row) else None
                    record[str(header)] = value
                records.append(record)
            return records
        elif format_style == 'index':
            return {i: row for i, row in enumerate(data)}
        else:  # 'values'
            return data