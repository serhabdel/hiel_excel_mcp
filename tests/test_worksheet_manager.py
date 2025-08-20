"""
Tests for WorksheetManager tool.

Tests all worksheet management operations including creation, copying, deletion,
renaming, cell merging operations, and validation information retrieval.
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from hiel_excel_mcp.tools.worksheet_manager import worksheet_manager
from hiel_excel_mcp.core.base_tool import OperationStatus


class TestWorksheetManager:
    """Test suite for WorksheetManager operations."""
    
    @pytest.fixture
    def temp_workbook(self):
        """Create a temporary workbook for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    @pytest.fixture
    def temp_workbook_with_sheets(self):
        """Create a temporary workbook with multiple sheets for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            # Rename default sheet
            wb.active.title = "Sheet1"
            # Add some data to Sheet1 to make validation work
            ws = wb.active
            ws['A1'] = "Header1"
            ws['B1'] = "Header2"
            ws['C1'] = "Header3"
            ws['A2'] = "Data1"
            ws['B2'] = "Data2"
            ws['C2'] = "Data3"
            ws['A3'] = "Data4"
            ws['B3'] = "Data5"
            ws['C3'] = "Data6"
            # Add additional sheets
            wb.create_sheet("Sheet2")
            wb.create_sheet("TestSheet")
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_create_worksheet_success(self, temp_workbook):
        """Test successful worksheet creation."""
        response = worksheet_manager.execute_operation(
            "create",
            filepath=temp_workbook,
            sheet_name="NewSheet"
        )
        
        assert response.success is True
        assert response.operation == "create"
        assert response.status == OperationStatus.SUCCESS
        assert "NewSheet" in response.message
        assert response.data["sheet_name"] == "NewSheet"
        assert response.data["filepath"] == temp_workbook
    
    def test_create_worksheet_duplicate_name(self, temp_workbook_with_sheets):
        """Test creating worksheet with duplicate name fails."""
        response = worksheet_manager.execute_operation(
            "create",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1"  # Already exists
        )
        
        assert response.success is False
        assert response.operation == "create"
        assert response.status == OperationStatus.ERROR
        assert "already exists" in response.message.lower()
    
    def test_create_worksheet_invalid_file(self):
        """Test creating worksheet in non-existent file fails."""
        response = worksheet_manager.execute_operation(
            "create",
            filepath="/nonexistent/file.xlsx",
            sheet_name="NewSheet"
        )
        
        assert response.success is False
        assert response.operation == "create"
        assert response.status == OperationStatus.ERROR
    
    def test_copy_worksheet_success(self, temp_workbook_with_sheets):
        """Test successful worksheet copying."""
        response = worksheet_manager.execute_operation(
            "copy",
            filepath=temp_workbook_with_sheets,
            source_sheet="Sheet1",
            target_sheet="CopiedSheet"
        )
        
        assert response.success is True
        assert response.operation == "copy"
        assert response.status == OperationStatus.SUCCESS
        assert "copied" in response.message.lower()
        assert response.data["source_sheet"] == "Sheet1"
        assert response.data["target_sheet"] == "CopiedSheet"
    
    def test_copy_worksheet_nonexistent_source(self, temp_workbook_with_sheets):
        """Test copying non-existent worksheet fails."""
        response = worksheet_manager.execute_operation(
            "copy",
            filepath=temp_workbook_with_sheets,
            source_sheet="NonExistentSheet",
            target_sheet="CopiedSheet"
        )
        
        assert response.success is False
        assert response.operation == "copy"
        assert response.status == OperationStatus.ERROR
        assert "not found" in response.message.lower()
    
    def test_copy_worksheet_duplicate_target(self, temp_workbook_with_sheets):
        """Test copying to existing sheet name fails."""
        response = worksheet_manager.execute_operation(
            "copy",
            filepath=temp_workbook_with_sheets,
            source_sheet="Sheet1",
            target_sheet="Sheet2"  # Already exists
        )
        
        assert response.success is False
        assert response.operation == "copy"
        assert response.status == OperationStatus.ERROR
        assert "already exists" in response.message.lower()
    
    def test_delete_worksheet_success(self, temp_workbook_with_sheets):
        """Test successful worksheet deletion."""
        response = worksheet_manager.execute_operation(
            "delete",
            filepath=temp_workbook_with_sheets,
            sheet_name="TestSheet"
        )
        
        assert response.success is True
        assert response.operation == "delete"
        assert response.status == OperationStatus.SUCCESS
        assert "deleted" in response.message.lower()
        assert response.data["sheet_name"] == "TestSheet"
    
    def test_delete_worksheet_nonexistent(self, temp_workbook_with_sheets):
        """Test deleting non-existent worksheet fails."""
        response = worksheet_manager.execute_operation(
            "delete",
            filepath=temp_workbook_with_sheets,
            sheet_name="NonExistentSheet"
        )
        
        assert response.success is False
        assert response.operation == "delete"
        assert response.status == OperationStatus.ERROR
        assert "not found" in response.message.lower()
    
    def test_delete_last_worksheet(self, temp_workbook):
        """Test deleting the only worksheet fails."""
        response = worksheet_manager.execute_operation(
            "delete",
            filepath=temp_workbook,
            sheet_name="Sheet"  # Default sheet name
        )
        
        assert response.success is False
        assert response.operation == "delete"
        assert response.status == OperationStatus.ERROR
        assert "only sheet" in response.message.lower()
    
    def test_rename_worksheet_success(self, temp_workbook_with_sheets):
        """Test successful worksheet renaming."""
        response = worksheet_manager.execute_operation(
            "rename",
            filepath=temp_workbook_with_sheets,
            old_name="TestSheet",
            new_name="RenamedSheet"
        )
        
        assert response.success is True
        assert response.operation == "rename"
        assert response.status == OperationStatus.SUCCESS
        assert "renamed" in response.message.lower()
        assert response.data["old_name"] == "TestSheet"
        assert response.data["new_name"] == "RenamedSheet"
    
    def test_rename_worksheet_nonexistent(self, temp_workbook_with_sheets):
        """Test renaming non-existent worksheet fails."""
        response = worksheet_manager.execute_operation(
            "rename",
            filepath=temp_workbook_with_sheets,
            old_name="NonExistentSheet",
            new_name="RenamedSheet"
        )
        
        assert response.success is False
        assert response.operation == "rename"
        assert response.status == OperationStatus.ERROR
        assert "not found" in response.message.lower()
    
    def test_rename_worksheet_duplicate_name(self, temp_workbook_with_sheets):
        """Test renaming to existing sheet name fails."""
        response = worksheet_manager.execute_operation(
            "rename",
            filepath=temp_workbook_with_sheets,
            old_name="TestSheet",
            new_name="Sheet1"  # Already exists
        )
        
        assert response.success is False
        assert response.operation == "rename"
        assert response.status == OperationStatus.ERROR
        assert "already exists" in response.message.lower()
    
    def test_get_merged_cells_empty(self, temp_workbook_with_sheets):
        """Test getting merged cells from worksheet with no merged cells."""
        response = worksheet_manager.execute_operation(
            "get_merged_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1"
        )
        
        assert response.success is True
        assert response.operation == "get_merged_cells"
        assert response.status == OperationStatus.SUCCESS
        assert response.data["merged_ranges"] == []
        assert response.data["count"] == 0
    
    def test_get_merged_cells_nonexistent_sheet(self, temp_workbook_with_sheets):
        """Test getting merged cells from non-existent worksheet fails."""
        response = worksheet_manager.execute_operation(
            "get_merged_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="NonExistentSheet"
        )
        
        assert response.success is False
        assert response.operation == "get_merged_cells"
        assert response.status == OperationStatus.ERROR
        assert "not found" in response.message.lower()
    
    def test_merge_cells_success(self, temp_workbook_with_sheets):
        """Test successful cell merging."""
        response = worksheet_manager.execute_operation(
            "merge_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="C3"
        )
        
        assert response.success is True
        assert response.operation == "merge_cells"
        assert response.status == OperationStatus.SUCCESS
        assert "merged" in response.message.lower()
        assert response.data["start_cell"] == "A1"
        assert response.data["end_cell"] == "C3"
        assert response.data["range"] == "A1:C3"
    
    def test_merge_cells_invalid_range(self, temp_workbook_with_sheets):
        """Test merging cells with invalid range fails."""
        response = worksheet_manager.execute_operation(
            "merge_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="invalid"
        )
        
        assert response.success is False
        assert response.operation == "merge_cells"
        assert response.status == OperationStatus.ERROR
    
    def test_unmerge_cells_success(self, temp_workbook_with_sheets):
        """Test successful cell unmerging."""
        # First merge some cells
        merge_response = worksheet_manager.execute_operation(
            "merge_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="B2"
        )
        assert merge_response.success is True
        
        # Then unmerge them
        response = worksheet_manager.execute_operation(
            "unmerge_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="B2"
        )
        
        assert response.success is True
        assert response.operation == "unmerge_cells"
        assert response.status == OperationStatus.SUCCESS
        assert "unmerged" in response.message.lower()
        assert response.data["start_cell"] == "A1"
        assert response.data["end_cell"] == "B2"
    
    def test_unmerge_cells_not_merged(self, temp_workbook_with_sheets):
        """Test unmerging cells that aren't merged fails."""
        response = worksheet_manager.execute_operation(
            "unmerge_cells",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="B2"
        )
        
        assert response.success is False
        assert response.operation == "unmerge_cells"
        assert response.status == OperationStatus.ERROR
        assert "not merged" in response.message.lower()
    
    def test_get_validation_info_success(self, temp_workbook_with_sheets):
        """Test successful validation info retrieval."""
        response = worksheet_manager.execute_operation(
            "get_validation_info",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1",
            end_cell="C3"
        )
        
        assert response.success is True
        assert response.operation == "get_validation_info"
        assert response.status == OperationStatus.SUCCESS
        assert "validation_info" in response.data
        assert response.data["start_cell"] == "A1"
        assert response.data["end_cell"] == "C3"
    
    def test_get_validation_info_single_cell(self, temp_workbook_with_sheets):
        """Test validation info for single cell."""
        response = worksheet_manager.execute_operation(
            "get_validation_info",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="A1"
        )
        
        assert response.success is True
        assert response.operation == "get_validation_info"
        assert response.status == OperationStatus.SUCCESS
        assert "validation_info" in response.data
        assert response.data["start_cell"] == "A1"
        assert response.data["end_cell"] is None
    
    def test_get_validation_info_invalid_range(self, temp_workbook_with_sheets):
        """Test validation info with invalid range fails."""
        response = worksheet_manager.execute_operation(
            "get_validation_info",
            filepath=temp_workbook_with_sheets,
            sheet_name="Sheet1",
            start_cell="invalid",
            end_cell="A1"
        )
        
        assert response.success is False
        assert response.operation == "get_validation_info"
        assert response.status == OperationStatus.ERROR
    
    def test_invalid_operation(self, temp_workbook):
        """Test invalid operation fails gracefully."""
        response = worksheet_manager.execute_operation(
            "invalid_operation",
            filepath=temp_workbook
        )
        
        assert response.success is False
        assert response.operation == "invalid_operation"
        assert response.status == OperationStatus.ERROR
        assert "not supported" in response.message.lower()
    
    def test_missing_required_parameters(self, temp_workbook):
        """Test missing required parameters fails gracefully."""
        response = worksheet_manager.execute_operation(
            "create",
            filepath=temp_workbook
            # Missing sheet_name parameter
        )
        
        assert response.success is False
        assert response.operation == "create"
        assert response.status == OperationStatus.ERROR
        assert "missing required parameters" in response.message.lower()
    
    def test_get_available_operations(self):
        """Test getting list of available operations."""
        operations = worksheet_manager.get_available_operations()
        
        expected_operations = [
            "create", "copy", "delete", "rename", 
            "get_merged_cells", "merge_cells", "unmerge_cells", 
            "get_validation_info"
        ]
        
        for op in expected_operations:
            assert op in operations
    
    def test_get_tool_info(self):
        """Test getting comprehensive tool information."""
        info = worksheet_manager.get_tool_info()
        
        assert info["name"] == "worksheet_manager"
        assert "description" in info
        assert "operations" in info
        
        # Check that all expected operations are documented
        expected_operations = [
            "create", "copy", "delete", "rename", 
            "get_merged_cells", "merge_cells", "unmerge_cells", 
            "get_validation_info"
        ]
        
        for op in expected_operations:
            assert op in info["operations"]
            assert "description" in info["operations"][op]
            assert "required_params" in info["operations"][op]
            assert "optional_params" in info["operations"][op]


class TestWorksheetManagerIntegration:
    """Integration tests for WorksheetManager with complex workflows."""
    
    @pytest.fixture
    def temp_workbook_complex(self):
        """Create a complex workbook for integration testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            wb.active.title = "MainSheet"
            
            # Add some data and formatting
            ws = wb.active
            ws['A1'] = "Header 1"
            ws['B1'] = "Header 2"
            ws['A2'] = "Data 1"
            ws['B2'] = "Data 2"
            
            # Create additional sheets
            wb.create_sheet("DataSheet")
            wb.create_sheet("AnalysisSheet")
            
            wb.save(tmp.name)
            wb.close()
            yield tmp.name
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_complete_worksheet_workflow(self, temp_workbook_complex):
        """Test a complete workflow of worksheet operations."""
        # 1. Create a new worksheet
        create_response = worksheet_manager.execute_operation(
            "create",
            filepath=temp_workbook_complex,
            sheet_name="NewWorksheet"
        )
        assert create_response.success is True
        
        # 2. Copy an existing worksheet
        copy_response = worksheet_manager.execute_operation(
            "copy",
            filepath=temp_workbook_complex,
            source_sheet="MainSheet",
            target_sheet="MainSheetCopy"
        )
        assert copy_response.success is True
        
        # 3. Rename a worksheet
        rename_response = worksheet_manager.execute_operation(
            "rename",
            filepath=temp_workbook_complex,
            old_name="NewWorksheet",
            new_name="ProcessedData"
        )
        assert rename_response.success is True
        
        # 4. Merge some cells
        merge_response = worksheet_manager.execute_operation(
            "merge_cells",
            filepath=temp_workbook_complex,
            sheet_name="ProcessedData",
            start_cell="A1",
            end_cell="B1"
        )
        assert merge_response.success is True
        
        # 5. Get merged cells info
        merged_info_response = worksheet_manager.execute_operation(
            "get_merged_cells",
            filepath=temp_workbook_complex,
            sheet_name="ProcessedData"
        )
        assert merged_info_response.success is True
        assert len(merged_info_response.data["merged_ranges"]) == 1
        
        # 6. Get validation info
        validation_response = worksheet_manager.execute_operation(
            "get_validation_info",
            filepath=temp_workbook_complex,
            sheet_name="MainSheet",
            start_cell="A1",
            end_cell="B2"
        )
        assert validation_response.success is True
        
        # 7. Delete a worksheet
        delete_response = worksheet_manager.execute_operation(
            "delete",
            filepath=temp_workbook_complex,
            sheet_name="AnalysisSheet"
        )
        assert delete_response.success is True
    
    def test_merge_unmerge_workflow(self, temp_workbook_complex):
        """Test merge and unmerge workflow."""
        # Merge cells
        merge_response = worksheet_manager.execute_operation(
            "merge_cells",
            filepath=temp_workbook_complex,
            sheet_name="MainSheet",
            start_cell="A1",
            end_cell="C1"
        )
        assert merge_response.success is True
        
        # Verify merge
        merged_info = worksheet_manager.execute_operation(
            "get_merged_cells",
            filepath=temp_workbook_complex,
            sheet_name="MainSheet"
        )
        assert merged_info.success is True
        assert len(merged_info.data["merged_ranges"]) == 1
        
        # Unmerge cells
        unmerge_response = worksheet_manager.execute_operation(
            "unmerge_cells",
            filepath=temp_workbook_complex,
            sheet_name="MainSheet",
            start_cell="A1",
            end_cell="C1"
        )
        assert unmerge_response.success is True
        
        # Verify unmerge
        merged_info_after = worksheet_manager.execute_operation(
            "get_merged_cells",
            filepath=temp_workbook_complex,
            sheet_name="MainSheet"
        )
        assert merged_info_after.success is True
        assert len(merged_info_after.data["merged_ranges"]) == 0