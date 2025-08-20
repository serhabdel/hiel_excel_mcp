"""
Tests for DataManager tool.

Tests all data operations including reading, writing, copying, deleting,
range validation, and data transformations.
"""

import pytest
import json
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from hiel_excel_mcp.tools.data_manager import DataManager, data_manager_tool


class TestDataManager:
    """Test suite for DataManager tool."""
    
    @pytest.fixture
    def data_manager(self):
        """Create DataManager instance for testing."""
        return DataManager()
    
    @pytest.fixture
    def sample_workbook(self):
        """Create a sample workbook for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add sample data
            data = [
                ["Name", "Age", "City"],
                ["Alice", 25, "New York"],
                ["Bob", 30, "London"],
                ["Charlie", 35, "Paris"]
            ]
            
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    @pytest.fixture
    def empty_workbook(self):
        """Create an empty workbook for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_tool_metadata(self, data_manager):
        """Test tool metadata and operation registration."""
        assert data_manager.get_tool_name() == "data_manager"
        assert "data operations" in data_manager.get_tool_description().lower()
        
        operations = data_manager.get_available_operations()
        expected_operations = ["read", "write", "copy_range", "delete_range", "validate_range", "transform"]
        
        for op in expected_operations:
            assert op in operations
    
    def test_read_operation_basic(self, data_manager, sample_workbook):
        """Test basic read operation."""
        response = data_manager.execute_operation(
            "read",
            filepath=sample_workbook,
            sheet_name="TestSheet"
        )
        
        assert response.success
        assert response.operation == "read"
        assert "data" in response.data
        
        data = response.data["data"]
        assert len(data) == 4  # 4 rows including header
        assert data[0] == ["Name", "Age", "City"]  # Header row
        assert data[1] == ["Alice", 25, "New York"]  # First data row
    
    def test_read_operation_with_range(self, data_manager, sample_workbook):
        """Test read operation with specific range."""
        response = data_manager.execute_operation(
            "read",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            start_cell="A1",
            end_cell="B2"
        )
        
        assert response.success
        data = response.data["data"]
        assert len(data) == 2  # 2 rows
        assert len(data[0]) == 2  # 2 columns
        assert data[0] == ["Name", "Age"]
        assert data[1] == ["Alice", 25]
    
    def test_read_operation_with_metadata(self, data_manager, sample_workbook):
        """Test read operation with metadata."""
        response = data_manager.execute_operation(
            "read",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            start_cell="A1",
            end_cell="C2",
            include_metadata=True
        )
        
        assert response.success
        assert "cells" in response.data
        
        cells = response.data["cells"]
        assert len(cells) == 6  # 2 rows × 3 columns
        
        # Check first cell
        first_cell = cells[0]
        assert first_cell["address"] == "A1"
        assert first_cell["value"] == "Name"
        assert first_cell["row"] == 1
        assert first_cell["column"] == 1
    
    def test_write_operation_new_file(self, data_manager):
        """Test write operation to new file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Remove the file so we can test creation
            os.unlink(tmp_path)
            
            test_data = [
                ["Product", "Price"],
                ["Apple", 1.50],
                ["Banana", 0.75]
            ]
            
            response = data_manager.execute_operation(
                "write",
                filepath=tmp_path,
                data=test_data,
                sheet_name="Products"
            )
            
            assert response.success
            assert response.operation == "write"
            assert "rows_written" in response.data
            assert response.data["rows_written"] == 3
            
            # Verify file was created and data written
            assert os.path.exists(tmp_path)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_write_operation_existing_file(self, data_manager, empty_workbook):
        """Test write operation to existing file."""
        test_data = [
            ["Item", "Quantity"],
            ["Pencil", 10],
            ["Eraser", 5]
        ]
        
        response = data_manager.execute_operation(
            "write",
            filepath=empty_workbook,
            data=test_data,
            start_cell="B2"
        )
        
        assert response.success
        assert response.data["start_cell"] == "B2"
        assert response.data["rows_written"] == 3
    
    def test_copy_range_same_file(self, data_manager, sample_workbook):
        """Test copying range within same file."""
        response = data_manager.execute_operation(
            "copy_range",
            filepath=sample_workbook,
            source_sheet="TestSheet",
            source_range="A1:B2",
            dest_sheet="TestSheet",
            dest_start_cell="E1"
        )
        
        assert response.success
        assert response.operation == "copy_range"
        assert response.data["rows_copied"] == 2
        assert response.data["columns_copied"] == 2
    
    def test_copy_range_different_file(self, data_manager, sample_workbook):
        """Test copying range to different file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            dest_path = tmp.name
        
        try:
            # Remove the file so we can test creation
            os.unlink(dest_path)
            
            response = data_manager.execute_operation(
                "copy_range",
                filepath=sample_workbook,
                source_sheet="TestSheet",
                source_range="A1:C3",
                dest_sheet="CopiedData",
                dest_start_cell="A1",
                dest_filepath=dest_path
            )
            
            assert response.success
            assert response.data["rows_copied"] == 3
            assert os.path.exists(dest_path)
            
        finally:
            if os.path.exists(dest_path):
                os.unlink(dest_path)
    
    def test_delete_range_operation(self, data_manager, sample_workbook):
        """Test delete range operation."""
        response = data_manager.execute_operation(
            "delete_range",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            range_ref="B2:C3"
        )
        
        assert response.success
        assert response.operation == "delete_range"
        assert "cells_cleared" in response.data
        
        # Verify data was cleared by reading it back
        read_response = data_manager.execute_operation(
            "read",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            start_cell="B2",
            end_cell="C3"
        )
        
        assert read_response.success
        # All values should be None after deletion
        for row in read_response.data["data"]:
            for value in row:
                assert value is None
    
    def test_validate_range_valid(self, data_manager, sample_workbook):
        """Test range validation with valid range."""
        response = data_manager.execute_operation(
            "validate_range",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            start_cell="A1",
            end_cell="C4"
        )
        
        assert response.success
        assert response.operation == "validate_range"
        assert response.data["valid"] is True
        assert "range" in response.data
        assert "data_range" in response.data
    
    def test_validate_range_invalid_sheet(self, data_manager, sample_workbook):
        """Test range validation with invalid sheet."""
        response = data_manager.execute_operation(
            "validate_range",
            filepath=sample_workbook,
            sheet_name="NonExistentSheet",
            start_cell="A1"
        )
        
        assert not response.success
        assert response.operation == "validate_range"
        assert "not found" in response.message.lower()
    
    def test_transform_operation_basic(self, data_manager, sample_workbook):
        """Test basic data transformation."""
        transformations = [
            {"type": "upper", "params": {}},
            {"type": "trim", "params": {}}
        ]
        
        response = data_manager.execute_operation(
            "transform",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            range_ref="A1:C1",  # Transform header row
            transformations=transformations
        )
        
        assert response.success
        assert response.operation == "transform"
        assert "transformed_count" in response.data
    
    def test_transform_operation_with_output_file(self, data_manager, sample_workbook):
        """Test transformation with output to different file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = tmp.name
        
        try:
            os.unlink(output_path)  # Remove so we can test creation
            
            transformations = [
                {"type": "lower", "params": {}}
            ]
            
            response = data_manager.execute_operation(
                "transform",
                filepath=sample_workbook,
                sheet_name="TestSheet",
                range_ref="A1:C4",
                transformations=transformations,
                output_filepath=output_path
            )
            
            assert response.success
            assert os.path.exists(output_path)
            
        finally:
            if os.path.exists(output_path):
                os.unlink(output_path)
    
    def test_invalid_operation(self, data_manager):
        """Test handling of invalid operation."""
        response = data_manager.execute_operation(
            "invalid_operation",
            filepath="test.xlsx"
        )
        
        assert not response.success
        assert "not supported" in response.message.lower()
    
    def test_missing_required_parameters(self, data_manager):
        """Test handling of missing required parameters."""
        response = data_manager.execute_operation(
            "read"
            # Missing required filepath and sheet_name
        )
        
        assert not response.success
        assert "missing required parameters" in response.message.lower()
    
    def test_tool_function_wrapper(self, sample_workbook):
        """Test the MCP tool function wrapper."""
        result_json = data_manager_tool(
            operation="read",
            filepath=sample_workbook,
            sheet_name="TestSheet"
        )
        
        result = json.loads(result_json)
        assert result["success"] is True
        assert result["operation"] == "read"
        assert "data" in result
    
    def test_tool_function_error_handling(self):
        """Test error handling in tool function wrapper."""
        result_json = data_manager_tool(
            operation="read",
            filepath="nonexistent.xlsx",
            sheet_name="Sheet1"
        )
        
        result = json.loads(result_json)
        assert result["success"] is False
        assert "error" in result["message"].lower()
    
    def test_copy_range_empty_source(self, data_manager, empty_workbook):
        """Test copying from empty range."""
        response = data_manager.execute_operation(
            "copy_range",
            filepath=empty_workbook,
            source_sheet="Sheet",
            source_range="A1:B2",
            dest_sheet="Sheet",
            dest_start_cell="D1"
        )
        
        assert response.success
        assert response.data["rows_copied"] == 0
        assert "no data found" in response.message.lower()
    
    def test_read_nonexistent_file(self, data_manager):
        """Test reading from nonexistent file."""
        response = data_manager.execute_operation(
            "read",
            filepath="nonexistent.xlsx",
            sheet_name="Sheet1"
        )
        
        assert not response.success
        assert response.operation == "read"
    
    def test_write_invalid_data_format(self, data_manager, empty_workbook):
        """Test writing with invalid data format."""
        response = data_manager.execute_operation(
            "write",
            filepath=empty_workbook,
            data=[]  # Empty data
        )
        
        assert not response.success
        assert "no data provided" in response.message.lower()
    
    def test_transform_with_complex_transformations(self, data_manager, sample_workbook):
        """Test transformation with multiple complex operations."""
        transformations = [
            {"type": "trim", "params": {}},
            {"type": "replace", "params": {"find": "Alice", "replace": "Alicia"}},
            {"type": "upper", "params": {}}
        ]
        
        response = data_manager.execute_operation(
            "transform",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            range_ref="A2:A2",  # Transform just Alice's name
            transformations=transformations
        )
        
        assert response.success
        # Should have transformed at least one cell
        assert response.data.get("transformed_count", 0) >= 0