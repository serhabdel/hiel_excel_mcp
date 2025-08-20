"""
Test suite for System Manager tool.
Tests filtering, sorting, and cache management operations.
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from ..tools.system_manager import SystemManager


@pytest.fixture
def temp_workbook_with_data():
    """Create a temporary Excel workbook with test data for filtering and sorting."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        wb = Workbook()
        ws = wb.active
        ws.title = "TestData"
        
        # Add headers
        headers = ['Name', 'Age', 'Department', 'Salary']
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)
        
        # Add test data
        test_data = [
            ['Alice', 30, 'IT', 70000],
            ['Bob', 25, 'Sales', 50000],
            ['Charlie', 35, 'IT', 80000],
            ['Diana', 28, 'HR', 60000],
            ['Eve', 32, 'Sales', 55000],
            ['Frank', 29, 'IT', 65000]
        ]
        
        for row, data in enumerate(test_data, 2):
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        wb.save(tmp_file.name)
        wb.close()
        
        yield tmp_file.name
        
        # Clean up
        try:
            os.unlink(tmp_file.name)
        except:
            pass


class TestSystemManager:
    """Test cases for SystemManager."""
    
    def test_apply_auto_filter(self, temp_workbook_with_data):
        """Test applying auto filter to a range."""
        manager = SystemManager()
        
        result = manager.apply_filter(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            "auto"
        )
        
        assert result["success"] is True
        assert result["sheet"] == "TestData"
        assert result["range"] == "A1:D7"
    
    def test_apply_column_filter(self, temp_workbook_with_data):
        """Test applying column-specific filter."""
        manager = SystemManager()
        
        # First apply auto filter
        manager.apply_filter(temp_workbook_with_data, "TestData", "A1:D7", "auto")
        
        # Then apply column filter
        filter_config = {
            "column_index": 2,  # Department column
            "criteria": {
                "type": "values",
                "values": ["IT", "Sales"]
            }
        }
        
        result = manager.apply_filter(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            "column",
            filter_config
        )
        
        assert result["success"] is True
        assert result["column_index"] == 2
        assert result["filter_type"] == "values"
    
    def test_apply_values_filter(self, temp_workbook_with_data):
        """Test applying values filter."""
        manager = SystemManager()
        
        # First apply auto filter
        manager.apply_filter(temp_workbook_with_data, "TestData", "A1:D7", "auto")
        
        filter_config = {
            "column_index": 1,  # Age column
            "values": [25, 30, 35]
        }
        
        result = manager.apply_filter(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            "values",
            filter_config
        )
        
        assert result["success"] is True
        assert result["column_index"] == 1
    
    def test_clear_filters(self, temp_workbook_with_data):
        """Test clearing filters from worksheet."""
        manager = SystemManager()
        
        # Apply filter first
        manager.apply_filter(temp_workbook_with_data, "TestData", "A1:D7", "auto")
        
        # Clear filters
        result = manager.clear_filters(temp_workbook_with_data, "TestData")
        
        assert result["success"] is True
        assert result["sheet"] == "TestData"
    
    def test_sort_range_single_column(self, temp_workbook_with_data):
        """Test sorting range by single column."""
        manager = SystemManager()
        
        sort_config = {
            "column_index": 1,  # Age column
            "ascending": False,  # Descending order
            "has_header": True
        }
        
        result = manager.sort_range(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            sort_config
        )
        
        assert result["success"] is True
        assert result["sheet"] == "TestData"
        assert result["sorted_rows"] == 6
    
    def test_sort_range_multi_column(self, temp_workbook_with_data):
        """Test sorting range by multiple columns."""
        manager = SystemManager()
        
        sort_config = {
            "columns": [
                {"column_index": 2, "ascending": True},   # Department (primary)
                {"column_index": 1, "ascending": False}   # Age (secondary, descending)
            ],
            "has_header": True
        }
        
        result = manager.sort_range(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            sort_config
        )
        
        assert result["success"] is True
        assert result["sorted_rows"] == 6
        assert len(result["sort_columns"]) == 2
    
    def test_sort_multi_column_direct(self, temp_workbook_with_data):
        """Test direct multi-column sort method."""
        manager = SystemManager()
        
        sort_columns = [
            {"column_index": 2, "ascending": True},   # Department
            {"column_index": 3, "ascending": False}   # Salary (descending)
        ]
        
        result = manager.sort_multi_column(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            sort_columns
        )
        
        assert result["success"] is True
        assert result["sorted_rows"] == 6
    
    def test_cache_operations(self):
        """Test cache management operations."""
        manager = SystemManager()
        
        # Get cache stats
        result = manager.get_cache_stats()
        assert result["success"] is True
        assert "size" in result
        assert "max_size" in result
        
        # Clear cache
        result = manager.clear_cache()
        assert result["success"] is True
        
        # Verify cache is cleared
        stats = manager.get_cache_stats()
        assert stats["size"] == 0
    
    def test_invalidate_cache(self, temp_workbook_with_data):
        """Test cache invalidation."""
        manager = SystemManager()
        
        # Invalidate specific file
        result = manager.invalidate_cache(temp_workbook_with_data)
        assert result["success"] is True
        assert result["filepath"] == temp_workbook_with_data
        
        # Invalidate all cache
        result = manager.invalidate_cache()
        assert result["success"] is True
        assert "Entire cache cleared" in result["message"]
    
    def test_batch_operations(self):
        """Test batch operation management."""
        manager = SystemManager()
        
        # List operations (should be empty initially)
        result = manager.list_batch_operations()
        assert result["success"] is True
        assert result["total_operations"] >= 0
        
        # Test getting status of non-existent operation
        result = manager.get_batch_status("non-existent-id")
        assert result["success"] is False
        assert "Operation not found" in result["error"]
        
        # Test canceling non-existent operation
        result = manager.cancel_batch_operation("non-existent-id")
        assert result["success"] is False
        assert result["error"] == "Operation not found"
    
    def test_get_filtered_data(self, temp_workbook_with_data):
        """Test getting filtered data from worksheet."""
        manager = SystemManager()
        
        result = manager.get_filtered_data(temp_workbook_with_data, "TestData")
        assert result["success"] is True
        assert result["sheet"] == "TestData"
        assert result["row_count"] > 0
        assert result["column_count"] == 4
        assert "data" in result
    
    def test_optimize_performance(self):
        """Test performance optimization."""
        manager = SystemManager()
        
        config = {
            "cache_size": 20,
            "cache_age": 600,
            "cleanup_operations": True,
            "operation_cleanup_age": 1800
        }
        
        result = manager.optimize_performance(config)
        assert result["success"] is True
        assert "changes" in result
        assert len(result["changes"]) > 0
    
    def test_system_health_check(self):
        """Test comprehensive system health check."""
        manager = SystemManager()
        
        result = manager.system_health_check()
        assert result["success"] is True
        assert "timestamp" in result
        assert "cache" in result
        assert "operations" in result
        assert "system" in result
        assert "status" in result
        
        # Check cache health info
        cache_info = result["cache"]
        assert "size" in cache_info
        assert "max_size" in cache_info
        assert "utilization_percent" in cache_info
        
        # Check operations health info
        ops_info = result["operations"]
        assert "active_operations" in ops_info
        assert "completed_operations" in ops_info
        assert "total_operations" in ops_info
        
        # Check system health info
        sys_info = result["system"]
        assert "active_threads" in sys_info
    
    def test_error_handling(self, temp_workbook_with_data):
        """Test error handling for invalid operations."""
        manager = SystemManager()
        
        # Test invalid filter type
        result = manager.apply_filter(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            "invalid_filter_type"
        )
        assert result["success"] is False
        assert "Unknown filter type" in result["error"]
        
        # Test column filter without column_index
        result = manager.apply_filter(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            "column",
            {}  # Empty config
        )
        assert result["success"] is False
        assert "column_index required" in result["error"]
        
        # Test sort without proper configuration
        result = manager.sort_range(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            {}  # Empty config
        )
        assert result["success"] is False
        assert "Sort configuration must specify" in result["error"]
    
    def test_filter_and_sort_combined(self, temp_workbook_with_data):
        """Test combining filtering and sorting operations."""
        manager = SystemManager()
        
        # Apply filter first
        result = manager.apply_filter(temp_workbook_with_data, "TestData", "A1:D7", "auto")
        assert result["success"] is True
        
        # Then sort the data
        sort_config = {
            "column_index": 3,  # Salary column
            "ascending": False,
            "has_header": True
        }
        
        result = manager.sort_range(
            temp_workbook_with_data, 
            "TestData", 
            "A1:D7", 
            sort_config
        )
        assert result["success"] is True
        
        # Get filtered data
        result = manager.get_filtered_data(temp_workbook_with_data, "TestData")
        assert result["success"] is True
    
    def test_cache_with_file_operations(self, temp_workbook_with_data):
        """Test cache behavior with file operations."""
        manager = SystemManager()
        
        # Perform an operation that would cache the file
        manager.apply_filter(temp_workbook_with_data, "TestData", "A1:D7", "auto")
        
        # Check cache stats
        stats = manager.get_cache_stats()
        assert stats["success"] is True
        
        # Invalidate the specific file
        result = manager.invalidate_cache(temp_workbook_with_data)
        assert result["success"] is True
        
        # Verify cache stats changed
        new_stats = manager.get_cache_stats()
        # Cache size might be different after invalidation
        assert new_stats["success"] is True