"""
Tests for workbook context management and caching system.
"""

import pytest
import tempfile
import time
import threading
from pathlib import Path
from unittest.mock import patch, MagicMock

from openpyxl import Workbook

from hiel_excel_mcp.core.workbook_context import (
    WorkbookContext,
    WorkbookCache,
    workbook_context,
    get_cache_stats,
    invalidate_cache,
    clear_cache,
    configure_cache
)


class TestWorkbookContext:
    """Test WorkbookContext class."""
    
    def test_context_manager_new_file(self):
        """Test context manager with new file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Remove the file so we test creation
            Path(tmp_path).unlink()
            
            with WorkbookContext(tmp_path) as wb:
                assert wb is not None
                ws = wb.active
                ws['A1'] = 'Test'
            
            # File should be saved
            assert Path(tmp_path).exists()
            
            # Verify content
            with WorkbookContext(tmp_path, read_only=True) as wb:
                assert wb.active['A1'].value == 'Test'
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_context_manager_existing_file(self):
        """Test context manager with existing file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Create initial file
            wb = Workbook()
            wb.active['A1'] = 'Initial'
            wb.save(tmp_path)
            wb.close()
            
            # Test loading existing file
            with WorkbookContext(tmp_path) as wb:
                assert wb.active['A1'].value == 'Initial'
                wb.active['B1'] = 'Modified'
            
            # Verify changes were saved
            with WorkbookContext(tmp_path, read_only=True) as wb:
                assert wb.active['A1'].value == 'Initial'
                assert wb.active['B1'].value == 'Modified'
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_read_only_mode(self):
        """Test read-only mode prevents saving."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Create initial file
            wb = Workbook()
            wb.active['A1'] = 'Initial'
            wb.save(tmp_path)
            wb.close()
            
            # Test read-only context
            context = WorkbookContext(tmp_path, read_only=True)
            
            with pytest.raises(ValueError, match="Cannot save read-only workbook"):
                context.save()
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_mark_dirty(self):
        """Test marking workbook as dirty."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            context = WorkbookContext(tmp_path)
            assert not context.is_dirty
            
            context.mark_dirty()
            assert context.is_dirty
            
            # After save, should not be dirty
            context.get_workbook()  # Load workbook
            context.save()
            assert not context.is_dirty
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_expiration(self):
        """Test context expiration."""
        context = WorkbookContext('test.xlsx')
        
        # Should not be expired initially
        assert not context.is_expired(max_age_seconds=1)
        
        # Mock time to simulate expiration
        with patch('time.time', return_value=time.time() + 2):
            assert context.is_expired(max_age_seconds=1)
    
    def test_access_tracking(self):
        """Test access count tracking."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            context = WorkbookContext(tmp_path)
            assert context.access_count == 0
            
            context.get_workbook()
            assert context.access_count == 1
            
            context.get_workbook()
            assert context.access_count == 2
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_thread_safety(self):
        """Test thread safety of context operations."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            context = WorkbookContext(tmp_path)
            results = []
            errors = []
            
            def worker():
                try:
                    wb = context.get_workbook()
                    results.append(wb is not None)
                except Exception as e:
                    errors.append(e)
            
            # Start multiple threads
            threads = [threading.Thread(target=worker) for _ in range(5)]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join()
            
            # All should succeed
            assert len(errors) == 0
            assert all(results)
            assert context.access_count == 5
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()


class TestWorkbookCache:
    """Test WorkbookCache class."""
    
    def test_cache_hit_miss(self):
        """Test cache hit and miss behavior."""
        cache = WorkbookCache(max_size=2)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # First access should be a miss
            context1 = cache.get_context(tmp_path)
            stats = cache.get_stats()
            assert stats['misses'] == 1
            assert stats['hits'] == 0
            
            # Second access should be a hit
            context2 = cache.get_context(tmp_path)
            assert context1 is context2
            stats = cache.get_stats()
            assert stats['hits'] == 1
            assert stats['misses'] == 1
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_cache_eviction(self):
        """Test LRU eviction when cache is full."""
        cache = WorkbookCache(max_size=2)
        
        files = []
        try:
            # Create test files
            for i in range(3):
                tmp = tempfile.NamedTemporaryFile(suffix=f'_{i}.xlsx', delete=False)
                files.append(tmp.name)
                tmp.close()
            
            # Fill cache
            context1 = cache.get_context(files[0])
            context2 = cache.get_context(files[1])
            
            # Access first context to make it more recent
            time.sleep(0.01)  # Ensure different timestamps
            context1.get_workbook()
            
            # Add third context, should evict second (LRU)
            context3 = cache.get_context(files[2])
            
            stats = cache.get_stats()
            assert stats['evictions'] == 1
            assert stats['size'] == 2
            
            # Accessing second file should be a miss now
            context2_new = cache.get_context(files[1])
            assert context2_new is not context2
        
        finally:
            for file_path in files:
                if Path(file_path).exists():
                    Path(file_path).unlink()
    
    def test_cache_expiration(self):
        """Test automatic expiration of cached contexts."""
        cache = WorkbookCache(max_age_seconds=0.1)  # Very short expiration
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Get context
            context1 = cache.get_context(tmp_path)
            
            # Wait for expiration
            time.sleep(0.2)
            
            # Should get new context due to expiration
            context2 = cache.get_context(tmp_path)
            assert context1 is not context2
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_cache_invalidation(self):
        """Test cache invalidation for specific files."""
        cache = WorkbookCache()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Get context
            context1 = cache.get_context(tmp_path)
            assert cache.get_stats()['size'] == 1
            
            # Invalidate cache
            cache.invalidate(tmp_path)
            assert cache.get_stats()['size'] == 0
            
            # Should get new context
            context2 = cache.get_context(tmp_path)
            assert context1 is not context2
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_cache_clear(self):
        """Test clearing entire cache."""
        cache = WorkbookCache()
        
        files = []
        try:
            # Create multiple cached contexts
            for i in range(3):
                tmp = tempfile.NamedTemporaryFile(suffix=f'_{i}.xlsx', delete=False)
                files.append(tmp.name)
                tmp.close()
                cache.get_context(tmp.name)
            
            assert cache.get_stats()['size'] == 3
            
            # Clear cache
            cache.clear()
            assert cache.get_stats()['size'] == 0
        
        finally:
            for file_path in files:
                if Path(file_path).exists():
                    Path(file_path).unlink()
    
    def test_different_modes_different_cache(self):
        """Test that different modes create different cache entries."""
        cache = WorkbookCache()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Create workbook first
            wb = Workbook()
            wb.save(tmp_path)
            wb.close()
            
            # Get contexts with different modes
            context1 = cache.get_context(tmp_path, read_only=False)
            context2 = cache.get_context(tmp_path, read_only=True)
            context3 = cache.get_context(tmp_path, data_only=True)
            
            # Should be different contexts
            assert context1 is not context2
            assert context1 is not context3
            assert context2 is not context3
            
            # Should have 3 cache entries
            assert cache.get_stats()['size'] == 3
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_thread_safety(self):
        """Test thread safety of cache operations."""
        cache = WorkbookCache()
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            contexts = []
            errors = []
            
            def worker():
                try:
                    context = cache.get_context(tmp_path)
                    contexts.append(context)
                except Exception as e:
                    errors.append(e)
            
            # Start multiple threads
            threads = [threading.Thread(target=worker) for _ in range(10)]
            for thread in threads:
                thread.start()
            for thread in threads:
                thread.join()
            
            # All should succeed and get same context
            assert len(errors) == 0
            assert len(contexts) == 10
            assert all(ctx is contexts[0] for ctx in contexts)
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()


class TestGlobalFunctions:
    """Test global convenience functions."""
    
    def test_workbook_context_function(self):
        """Test workbook_context convenience function."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Remove file to test creation
            Path(tmp_path).unlink()
            
            with workbook_context(tmp_path) as wb:
                assert wb is not None
                wb.active['A1'] = 'Test'
            
            # File should exist and contain data
            assert Path(tmp_path).exists()
            
            with workbook_context(tmp_path, read_only=True) as wb:
                assert wb.active['A1'].value == 'Test'
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_global_cache_functions(self):
        """Test global cache management functions."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Use workbook_context to populate cache
            with workbook_context(tmp_path) as wb:
                wb.active['A1'] = 'Test'
            
            # Check stats
            stats = get_cache_stats()
            assert stats['size'] >= 1
            
            # Invalidate specific file
            invalidate_cache(tmp_path)
            stats = get_cache_stats()
            assert stats['size'] == 0
            
            # Populate cache again
            with workbook_context(tmp_path) as wb:
                pass
            
            # Clear entire cache
            clear_cache()
            stats = get_cache_stats()
            assert stats['size'] == 0
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_configure_cache(self):
        """Test cache configuration."""
        # Configure with custom settings
        configure_cache(max_size=5, max_age_seconds=60)
        
        # Create some contexts to test new settings
        files = []
        try:
            for i in range(3):
                tmp = tempfile.NamedTemporaryFile(suffix=f'_{i}.xlsx', delete=False)
                files.append(tmp.name)
                tmp.close()
                
                with workbook_context(tmp.name) as wb:
                    wb.active['A1'] = f'Test{i}'
            
            stats = get_cache_stats()
            assert stats['max_size'] == 5
            assert stats['size'] == 3
        
        finally:
            for file_path in files:
                if Path(file_path).exists():
                    Path(file_path).unlink()
            clear_cache()


class TestPerformanceOptimizations:
    """Test performance optimization features."""
    
    def test_workbook_reuse(self):
        """Test that workbook instances are reused for performance."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Create initial file
            wb = Workbook()
            wb.active['A1'] = 'Initial'
            wb.save(tmp_path)
            wb.close()
            
            # Multiple operations should reuse same workbook
            with workbook_context(tmp_path) as wb1:
                wb1.active['B1'] = 'First'
            
            with workbook_context(tmp_path) as wb2:
                wb2.active['C1'] = 'Second'
                # Should see previous change
                assert wb2.active['B1'].value == 'First'
            
            # Verify all changes were saved
            with workbook_context(tmp_path, read_only=True) as wb:
                assert wb.active['A1'].value == 'Initial'
                assert wb.active['B1'].value == 'First'
                assert wb.active['C1'].value == 'Second'
        
        finally:
            if Path(tmp_path).exists():
                Path(tmp_path).unlink()
    
    def test_memory_management(self):
        """Test memory management with cache limits."""
        # Use small cache to test eviction
        configure_cache(max_size=2, max_age_seconds=300)
        
        files = []
        try:
            # Create more files than cache can hold
            for i in range(4):
                tmp = tempfile.NamedTemporaryFile(suffix=f'_{i}.xlsx', delete=False)
                files.append(tmp.name)
                tmp.close()
                
                with workbook_context(tmp.name) as wb:
                    wb.active['A1'] = f'Test{i}'
            
            # Cache should only hold 2 items
            stats = get_cache_stats()
            assert stats['size'] == 2
            assert stats['evictions'] >= 2
        
        finally:
            for file_path in files:
                if Path(file_path).exists():
                    Path(file_path).unlink()
            clear_cache()


if __name__ == '__main__':
    pytest.main([__file__])