"""
Tests for Cell Manager Tool.

Tests all cell manipulation operations including row/column insertion/deletion,
cell formatting, information retrieval, updates, and clearing.
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook

from hiel_excel_mcp.tools.cell_manager import cell_manager, cell_manager_tool
from hiel_excel_mcp.core.base_tool import OperationStatus


class TestCellManager:
    """Test suite for CellManager operations."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add some test data
            ws['A1'] = "Header1"
            ws['B1'] = "Header2"
            ws['C1'] = "Header3"
            ws['A2'] = "Data1"
            ws['B2'] = "Data2"
            ws['C2'] = "Data3"
            ws['A3'] = 100
            ws['B3'] = 200
            ws['C3'] = 300
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_tool_metadata(self):
        """Test tool metadata and operation registration."""
        assert cell_manager.get_tool_name() == "cell_manager"
        assert "cell manipulation" in cell_manager.get_tool_description().lower()
        
        operations = cell_manager.get_available_operations()
        expected_operations = [
            "insert_rows", "insert_columns", "delete_rows", "delete_columns",
            "format_range", "get_cell_info", "update_cell", "clear_cells"
        ]
        
        for op in expected_operations:
            assert op in operations
    
    def test_insert_rows(self, temp_excel_file):
        """Test row insertion operation."""
        response = cell_manager.execute_operation(
            "insert_rows",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_row=2,
            count=2
        )
        
        assert response.success is True
        assert response.operation == "insert_rows"
        assert "2 row(s)" in response.message
        
        # Verify rows were inserted
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        # Original data should be shifted down
        assert ws['A1'].value == "Header1"  # Headers unchanged
        assert ws['A2'].value is None  # New empty row
        assert ws['A3'].value is None  # New empty row
        assert ws['A4'].value == "Data1"  # Original data shifted
        
        wb.close()
    
    def test_insert_columns(self, temp_excel_file):
        """Test column insertion operation."""
        response = cell_manager.execute_operation(
            "insert_columns",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_col=2,
            count=1
        )
        
        assert response.success is True
        assert response.operation == "insert_columns"
        assert "1 column(s)" in response.message
        
        # Verify column was inserted
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        # Original data should be shifted right
        assert ws['A1'].value == "Header1"  # First column unchanged
        assert ws['B1'].value is None  # New empty column
        assert ws['C1'].value == "Header2"  # Original data shifted
        
        wb.close()
    
    def test_delete_rows(self, temp_excel_file):
        """Test row deletion operation."""
        response = cell_manager.execute_operation(
            "delete_rows",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_row=2,
            count=1
        )
        
        assert response.success is True
        assert response.operation == "delete_rows"
        assert "1 row(s)" in response.message
        
        # Verify row was deleted
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        # Data should be shifted up
        assert ws['A1'].value == "Header1"  # Headers unchanged
        assert ws['A2'].value == 100  # Row 3 data moved to row 2
        
        wb.close()
    
    def test_delete_columns(self, temp_excel_file):
        """Test column deletion operation."""
        response = cell_manager.execute_operation(
            "delete_columns",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_col=2,
            count=1
        )
        
        assert response.success is True
        assert response.operation == "delete_columns"
        assert "1 column(s)" in response.message
        
        # Verify column was deleted
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        # Data should be shifted left
        assert ws['A1'].value == "Header1"  # First column unchanged
        assert ws['B1'].value == "Header3"  # Column C moved to B
        
        wb.close()
    
    def test_format_range_basic(self, temp_excel_file):
        """Test basic cell formatting operation."""
        response = cell_manager.execute_operation(
            "format_range",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1",
            end_cell="C1",
            bold=True,
            font_size=14,
            font_color="FF0000"
        )
        
        assert response.success is True
        assert response.operation == "format_range"
        assert "A1:C1" in response.message
        
        # Verify formatting was applied
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        cell = ws['A1']
        assert cell.font.bold is True
        assert cell.font.size == 14
        
        wb.close()
    
    def test_format_range_single_cell(self, temp_excel_file):
        """Test formatting a single cell."""
        response = cell_manager.execute_operation(
            "format_range",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="B2",
            italic=True,
            bg_color="FFFF00"
        )
        
        assert response.success is True
        assert response.operation == "format_range"
        assert "B2" in response.message
        
        # Verify formatting was applied
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        cell = ws['B2']
        assert cell.font.italic is True
        
        wb.close()
    
    def test_get_cell_info_single_cell(self, temp_excel_file):
        """Test getting information for a single cell."""
        response = cell_manager.execute_operation(
            "get_cell_info",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1"
        )
        
        assert response.success is True
        assert response.operation == "get_cell_info"
        assert response.data is not None
        
        cell_info = response.data["cell_info"]
        assert cell_info["sheet_name"] == "TestSheet"
        assert len(cell_info["cells"]) == 1
        assert cell_info["cells"][0]["address"] == "A1"
        assert cell_info["cells"][0]["value"] == "Header1"
    
    def test_get_cell_info_range(self, temp_excel_file):
        """Test getting information for a range of cells."""
        response = cell_manager.execute_operation(
            "get_cell_info",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1",
            end_cell="C2"
        )
        
        assert response.success is True
        assert response.operation == "get_cell_info"
        assert response.data is not None
        
        cell_info = response.data["cell_info"]
        assert cell_info["sheet_name"] == "TestSheet"
        assert len(cell_info["cells"]) == 6  # 3x2 range
        
        # Check specific cells
        cells_by_address = {cell["address"]: cell for cell in cell_info["cells"]}
        assert cells_by_address["A1"]["value"] == "Header1"
        assert cells_by_address["B1"]["value"] == "Header2"
        assert cells_by_address["A2"]["value"] == "Data1"
    
    def test_update_cell(self, temp_excel_file):
        """Test updating a cell value."""
        response = cell_manager.execute_operation(
            "update_cell",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            cell_address="B2",
            value="UpdatedData"
        )
        
        assert response.success is True
        assert response.operation == "update_cell"
        assert response.data["old_value"] == "Data2"
        assert response.data["new_value"] == "UpdatedData"
        
        # Verify cell was updated
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        assert ws['B2'].value == "UpdatedData"
        wb.close()
    
    def test_update_cell_numeric(self, temp_excel_file):
        """Test updating a cell with numeric value."""
        response = cell_manager.execute_operation(
            "update_cell",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            cell_address="A3",
            value=999
        )
        
        assert response.success is True
        assert response.operation == "update_cell"
        assert response.data["old_value"] == 100
        assert response.data["new_value"] == 999
        
        # Verify cell was updated
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        assert ws['A3'].value == 999
        wb.close()
    
    def test_clear_cells_single(self, temp_excel_file):
        """Test clearing a single cell."""
        response = cell_manager.execute_operation(
            "clear_cells",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="B2"
        )
        
        assert response.success is True
        assert response.operation == "clear_cells"
        assert response.data["cells_cleared"] == 1
        
        # Verify cell was cleared
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        assert ws['B2'].value is None
        wb.close()
    
    def test_clear_cells_range(self, temp_excel_file):
        """Test clearing a range of cells."""
        response = cell_manager.execute_operation(
            "clear_cells",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A2",
            end_cell="C3"
        )
        
        assert response.success is True
        assert response.operation == "clear_cells"
        assert response.data["cells_cleared"] == 6  # All cells had values
        
        # Verify cells were cleared
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        for row in range(2, 4):
            for col in ['A', 'B', 'C']:
                assert ws[f'{col}{row}'].value is None
        
        wb.close()
    
    def test_clear_cells_with_formatting(self, temp_excel_file):
        """Test clearing cells with formatting."""
        # First apply some formatting
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        from openpyxl.styles import Font
        ws['A1'].font = Font(bold=True, size=16)
        wb.save(temp_excel_file)
        wb.close()
        
        # Clear with formatting
        response = cell_manager.execute_operation(
            "clear_cells",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1",
            clear_formatting=True
        )
        
        assert response.success is True
        assert response.data["formatting_cleared"] is True
        
        # Verify formatting was cleared
        wb = load_workbook(temp_excel_file)
        ws = wb["TestSheet"]
        
        cell = ws['A1']
        assert cell.value is None
        assert cell.font.bold is False or cell.font.bold is None
        
        wb.close()
    
    def test_invalid_operation(self, temp_excel_file):
        """Test handling of invalid operation."""
        response = cell_manager.execute_operation(
            "invalid_operation",
            filepath=temp_excel_file,
            sheet_name="TestSheet"
        )
        
        assert response.success is False
        assert response.status == OperationStatus.ERROR
        assert "not supported" in response.message
    
    def test_invalid_cell_reference(self, temp_excel_file):
        """Test handling of invalid cell references."""
        response = cell_manager.execute_operation(
            "update_cell",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            cell_address="INVALID",
            value="test"
        )
        
        assert response.success is False
        assert response.status == OperationStatus.ERROR
        assert "Invalid cell reference" in response.message
    
    def test_nonexistent_sheet(self, temp_excel_file):
        """Test handling of nonexistent sheet."""
        response = cell_manager.execute_operation(
            "update_cell",
            filepath=temp_excel_file,
            sheet_name="NonexistentSheet",
            cell_address="A1",
            value="test"
        )
        
        assert response.success is False
        assert response.status == OperationStatus.ERROR
        assert "not found" in response.message
    
    def test_missing_required_parameters(self, temp_excel_file):
        """Test handling of missing required parameters."""
        response = cell_manager.execute_operation(
            "update_cell",
            filepath=temp_excel_file,
            sheet_name="TestSheet"
            # Missing cell_address and value
        )
        
        assert response.success is False
        assert response.status == OperationStatus.ERROR
        assert "Missing required parameters" in response.message
    
    def test_mcp_tool_function(self, temp_excel_file):
        """Test the MCP tool function wrapper."""
        result_json = cell_manager_tool(
            operation="get_cell_info",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1"
        )
        
        import json
        result = json.loads(result_json)
        
        assert result["success"] is True
        assert result["operation"] == "get_cell_info"
        assert "cell_info" in result["data"]
    
    def test_tool_info(self):
        """Test getting comprehensive tool information."""
        tool_info = cell_manager.get_tool_info()
        
        assert tool_info["name"] == "cell_manager"
        assert "operations" in tool_info
        
        # Check that all expected operations are documented
        operations = tool_info["operations"]
        expected_operations = [
            "insert_rows", "insert_columns", "delete_rows", "delete_columns",
            "format_range", "get_cell_info", "update_cell", "clear_cells"
        ]
        
        for op in expected_operations:
            assert op in operations
            assert "description" in operations[op]
            assert "required_params" in operations[op]
            assert "optional_params" in operations[op]


class TestCellManagerIntegration:
    """Integration tests for CellManager with complex scenarios."""
    
    @pytest.fixture
    def complex_excel_file(self):
        """Create a more complex Excel file for integration testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            
            # Create multiple sheets
            ws1 = wb.active
            ws1.title = "Data"
            
            ws2 = wb.create_sheet("Summary")
            
            # Add data to first sheet
            headers = ["Name", "Age", "Salary", "Department"]
            for col, header in enumerate(headers, 1):
                ws1.cell(row=1, column=col, value=header)
            
            data = [
                ["Alice", 30, 50000, "Engineering"],
                ["Bob", 25, 45000, "Marketing"],
                ["Charlie", 35, 60000, "Engineering"],
                ["Diana", 28, 48000, "Sales"]
            ]
            
            for row, record in enumerate(data, 2):
                for col, value in enumerate(record, 1):
                    ws1.cell(row=row, column=col, value=value)
            
            # Add summary to second sheet
            ws2['A1'] = "Summary Report"
            ws2['A2'] = "Total Employees"
            ws2['B2'] = len(data)
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_multi_operation_workflow(self, complex_excel_file):
        """Test a workflow involving multiple operations."""
        # 1. Insert a new row for a new employee
        response1 = cell_manager.execute_operation(
            "insert_rows",
            filepath=complex_excel_file,
            sheet_name="Data",
            start_row=6,
            count=1
        )
        assert response1.success is True
        
        # 2. Add the new employee data
        new_employee = ["Eve", 32, 55000, "Engineering"]
        for col, value in enumerate(new_employee, 1):
            response = cell_manager.execute_operation(
                "update_cell",
                filepath=complex_excel_file,
                sheet_name="Data",
                cell_address=f"{chr(64+col)}6",  # A6, B6, C6, D6
                value=value
            )
            assert response.success is True
        
        # 3. Format the header row
        response3 = cell_manager.execute_operation(
            "format_range",
            filepath=complex_excel_file,
            sheet_name="Data",
            start_cell="A1",
            end_cell="D1",
            bold=True,
            bg_color="CCCCCC"
        )
        assert response3.success is True
        
        # 4. Get info about the updated data
        response4 = cell_manager.execute_operation(
            "get_cell_info",
            filepath=complex_excel_file,
            sheet_name="Data",
            start_cell="A1",
            end_cell="D6"
        )
        assert response4.success is True
        assert len(response4.data["cell_info"]["cells"]) == 24  # 4x6 grid
        
        # Verify the final state
        wb = load_workbook(complex_excel_file)
        ws = wb["Data"]
        
        # Check new employee was added
        assert ws['A6'].value == "Eve"
        assert ws['B6'].value == 32
        assert ws['C6'].value == 55000
        assert ws['D6'].value == "Engineering"
        
        # Check formatting was applied
        assert ws['A1'].font.bold is True
        
        wb.close()
    
    def test_error_recovery(self, complex_excel_file):
        """Test error handling and recovery in complex scenarios."""
        # Try to insert rows with invalid parameters
        response1 = cell_manager.execute_operation(
            "insert_rows",
            filepath=complex_excel_file,
            sheet_name="Data",
            start_row=-1  # Invalid row number
        )
        assert response1.success is False
        
        # File should still be intact - try a valid operation
        response2 = cell_manager.execute_operation(
            "get_cell_info",
            filepath=complex_excel_file,
            sheet_name="Data",
            start_cell="A1"
        )
        assert response2.success is True
        
        # Try to format with invalid cell reference
        response3 = cell_manager.execute_operation(
            "format_range",
            filepath=complex_excel_file,
            sheet_name="Data",
            start_cell="INVALID_CELL",
            bold=True
        )
        assert response3.success is False
        
        # File should still be intact
        response4 = cell_manager.execute_operation(
            "update_cell",
            filepath=complex_excel_file,
            sheet_name="Data",
            cell_address="A1",
            value="Updated Header"
        )
        assert response4.success is True