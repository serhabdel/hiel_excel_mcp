"""
Tests for Import/Export Manager Tool.

Tests all import/export operations including CSV, HTML, JSON formats,
batch operations, and preview functionality.
"""

import pytest
import json
import csv
from pathlib import Path
from tempfile import TemporaryDirectory
from typing import List, Dict, Any

from ..tools.import_export_manager import ImportExportManager
from ..core.base_tool import OperationResponse


class TestImportExportManager:
    """Test suite for ImportExportManager tool."""
    
    @pytest.fixture
    def manager(self):
        """Create ImportExportManager instance."""
        return ImportExportManager()
    
    @pytest.fixture
    def temp_dir(self):
        """Create temporary directory for test files."""
        with TemporaryDirectory() as temp_dir:
            yield Path(temp_dir)
    
    @pytest.fixture
    def sample_csv_data(self):
        """Sample CSV data for testing."""
        return [
            ["Name", "Age", "City"],
            ["Alice", "25", "New York"],
            ["Bob", "30", "Los Angeles"],
            ["Charlie", "35", "Chicago"]
        ]
    
    @pytest.fixture
    def sample_csv_file(self, temp_dir, sample_csv_data):
        """Create sample CSV file."""
        csv_path = temp_dir / "sample.csv"
        with open(csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            writer.writerows(sample_csv_data)
        return csv_path
    
    @pytest.fixture
    def sample_excel_file(self, temp_dir, sample_csv_data):
        """Create sample Excel file with data."""
        from openpyxl import Workbook
        
        excel_path = temp_dir / "sample.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add sample data
        for row_idx, row_data in enumerate(sample_csv_data):
            for col_idx, value in enumerate(row_data):
                ws.cell(row=row_idx + 1, column=col_idx + 1, value=value)
        
        wb.save(excel_path)
        wb.close()
        return excel_path
    
    def test_tool_metadata(self, manager):
        """Test tool metadata and operations."""
        assert manager.get_tool_name() == "import_export_manager"
        assert "import and export operations" in manager.get_tool_description()
        
        operations = manager.get_available_operations()
        expected_operations = [
            "import_csv", "export_csv", "preview_csv",
            "export_html", "export_json", "batch_import"
        ]
        
        for op in expected_operations:
            assert op in operations
    
    def test_import_csv_success(self, manager, temp_dir, sample_csv_file):
        """Test successful CSV import."""
        excel_path = temp_dir / "output.xlsx"
        
        response = manager.execute_operation(
            "import_csv",
            csv_path=str(sample_csv_file),
            excel_path=str(excel_path),
            sheet_name="ImportedData"
        )
        
        assert response.success
        assert response.operation == "import_csv"
        assert "Successfully imported" in response.message
        assert response.data["rows_imported"] == 4
        assert response.data["columns_imported"] == 3
        
        # Verify Excel file was created
        assert excel_path.exists()
        
        # Verify data was imported correctly
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        assert "ImportedData" in wb.sheetnames
        ws = wb["ImportedData"]
        
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"
        assert ws.cell(2, 2).value == "25"
        wb.close()
    
    def test_import_csv_with_options(self, manager, temp_dir):
        """Test CSV import with various options."""
        # Create CSV with semicolon delimiter
        csv_data = [
            "Name;Age;City",
            "Alice;25;New York",
            "Bob;30;Los Angeles"
        ]
        csv_path = temp_dir / "semicolon.csv"
        csv_path.write_text('\n'.join(csv_data), encoding='utf-8')
        
        excel_path = temp_dir / "output.xlsx"
        
        response = manager.execute_operation(
            "import_csv",
            csv_path=str(csv_path),
            excel_path=str(excel_path),
            sheet_name="TestSheet",
            delimiter=";",
            start_cell="B2",
            max_rows=2,
            skip_rows=0
        )
        
        assert response.success
        assert response.data["rows_imported"] == 2
        
        # Verify data placement
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        ws = wb["TestSheet"]
        
        assert ws.cell(2, 2).value == "Name"  # B2
        assert ws.cell(3, 2).value == "Alice"  # B3
        wb.close()
    
    def test_import_csv_file_not_found(self, manager, temp_dir):
        """Test CSV import with non-existent file."""
        response = manager.execute_operation(
            "import_csv",
            csv_path=str(temp_dir / "nonexistent.csv"),
            excel_path=str(temp_dir / "output.xlsx"),
            sheet_name="TestSheet"
        )
        
        assert not response.success
        assert "not found" in response.message.lower()
    
    def test_export_csv_success(self, manager, temp_dir, sample_excel_file):
        """Test successful CSV export."""
        csv_path = temp_dir / "exported.csv"
        
        response = manager.execute_operation(
            "export_csv",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            csv_path=str(csv_path)
        )
        
        assert response.success
        assert response.operation == "export_csv"
        assert "Successfully exported" in response.message
        assert response.data["rows_exported"] == 4
        assert response.data["columns_exported"] == 3
        
        # Verify CSV file was created and contains correct data
        assert csv_path.exists()
        
        with open(csv_path, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            rows = list(reader)
        
        assert len(rows) == 4
        assert rows[0] == ["Name", "Age", "City"]
        assert rows[1] == ["Alice", "25", "New York"]
    
    def test_export_csv_with_range(self, manager, temp_dir, sample_excel_file):
        """Test CSV export with specific range."""
        csv_path = temp_dir / "exported_range.csv"
        
        response = manager.execute_operation(
            "export_csv",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            csv_path=str(csv_path),
            start_cell="A1",
            end_cell="B3",
            delimiter=";"
        )
        
        assert response.success
        assert response.data["rows_exported"] == 3
        assert response.data["columns_exported"] == 2
        
        # Verify exported data
        with open(csv_path, 'r', encoding='utf-8') as f:
            content = f.read()
        
        assert "Name;Age" in content
        assert "Alice;25" in content
        assert "Charlie" not in content  # Should not be included due to range
    
    def test_export_csv_sheet_not_found(self, manager, temp_dir, sample_excel_file):
        """Test CSV export with non-existent sheet."""
        csv_path = temp_dir / "exported.csv"
        
        response = manager.execute_operation(
            "export_csv",
            excel_path=str(sample_excel_file),
            sheet_name="NonExistentSheet",
            csv_path=str(csv_path)
        )
        
        assert not response.success
        assert "not found" in response.message.lower()
    
    def test_preview_csv_success(self, manager, sample_csv_file):
        """Test successful CSV preview."""
        response = manager.execute_operation(
            "preview_csv",
            csv_path=str(sample_csv_file),
            num_rows=3
        )
        
        assert response.success
        assert response.operation == "preview_csv"
        assert "Preview of" in response.message
        assert response.data["rows_previewed"] == 3
        assert response.data["columns_detected"] == 3
        
        preview_data = response.data["preview_data"]
        assert len(preview_data) == 3
        assert preview_data[0] == ["Name", "Age", "City"]
        assert preview_data[1] == ["Alice", "25", "New York"]
    
    def test_preview_csv_with_delimiter(self, manager, temp_dir):
        """Test CSV preview with custom delimiter."""
        # Create CSV with tab delimiter
        csv_data = "Name\tAge\tCity\nAlice\t25\tNew York\nBob\t30\tLos Angeles"
        csv_path = temp_dir / "tab_delimited.csv"
        csv_path.write_text(csv_data, encoding='utf-8')
        
        response = manager.execute_operation(
            "preview_csv",
            csv_path=str(csv_path),
            delimiter="\t",
            num_rows=2
        )
        
        assert response.success
        assert response.data["rows_previewed"] == 2
        assert response.data["detected_delimiter"] == "\t"
        
        preview_data = response.data["preview_data"]
        assert preview_data[0] == ["Name", "Age", "City"]
        assert preview_data[1] == ["Alice", "25", "New York"]
    
    def test_export_html_success(self, manager, temp_dir, sample_excel_file):
        """Test successful HTML export."""
        html_path = temp_dir / "exported.html"
        
        response = manager.execute_operation(
            "export_html",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            html_path=str(html_path),
            table_title="Test Export"
        )
        
        assert response.success
        assert response.operation == "export_html"
        assert "Successfully exported" in response.message
        assert response.data["rows_exported"] == 4
        
        # Verify HTML file was created
        assert html_path.exists()
        
        html_content = html_path.read_text(encoding='utf-8')
        assert "<!DOCTYPE html>" in html_content
        assert "<table>" in html_content
        assert "Test Export" in html_content
        assert "Alice" in html_content
        assert "New York" in html_content
    
    def test_export_html_without_styles(self, manager, temp_dir, sample_excel_file):
        """Test HTML export without CSS styles."""
        html_path = temp_dir / "exported_no_styles.html"
        
        response = manager.execute_operation(
            "export_html",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            html_path=str(html_path),
            include_styles=False
        )
        
        assert response.success
        
        html_content = html_path.read_text(encoding='utf-8')
        assert "<style>" not in html_content
        assert "<table>" in html_content
    
    def test_export_json_records_format(self, manager, temp_dir, sample_excel_file):
        """Test JSON export in records format."""
        json_path = temp_dir / "exported_records.json"
        
        response = manager.execute_operation(
            "export_json",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            json_path=str(json_path),
            format_style="records",
            include_headers=True
        )
        
        assert response.success
        assert response.operation == "export_json"
        assert response.data["format_style"] == "records"
        
        # Verify JSON file content
        assert json_path.exists()
        
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        assert isinstance(json_data, list)
        assert len(json_data) == 3  # 4 rows - 1 header = 3 records
        assert json_data[0]["Name"] == "Alice"
        assert json_data[0]["Age"] == "25"
        assert json_data[0]["City"] == "New York"
    
    def test_export_json_values_format(self, manager, temp_dir, sample_excel_file):
        """Test JSON export in values format."""
        json_path = temp_dir / "exported_values.json"
        
        response = manager.execute_operation(
            "export_json",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            json_path=str(json_path),
            format_style="values"
        )
        
        assert response.success
        
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        assert isinstance(json_data, list)
        assert len(json_data) == 4
        assert json_data[0] == ["Name", "Age", "City"]
        assert json_data[1] == ["Alice", "25", "New York"]
    
    def test_export_json_index_format(self, manager, temp_dir, sample_excel_file):
        """Test JSON export in index format."""
        json_path = temp_dir / "exported_index.json"
        
        response = manager.execute_operation(
            "export_json",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            json_path=str(json_path),
            format_style="index"
        )
        
        assert response.success
        
        with open(json_path, 'r', encoding='utf-8') as f:
            json_data = json.load(f)
        
        assert isinstance(json_data, dict)
        assert "0" in json_data
        assert json_data["0"] == ["Name", "Age", "City"]
        assert json_data["1"] == ["Alice", "25", "New York"]
    
    def test_export_json_invalid_format(self, manager, temp_dir, sample_excel_file):
        """Test JSON export with invalid format style."""
        json_path = temp_dir / "exported.json"
        
        response = manager.execute_operation(
            "export_json",
            excel_path=str(sample_excel_file),
            sheet_name="TestSheet",
            json_path=str(json_path),
            format_style="invalid_format"
        )
        
        assert not response.success
        assert "Invalid format_style" in response.message
    
    def test_batch_import_success(self, manager, temp_dir):
        """Test successful batch import of multiple CSV files."""
        # Create multiple CSV files
        csv_files = []
        
        for i in range(3):
            csv_data = [
                ["Name", "Value"],
                [f"Item{i}_1", f"{i*10 + 1}"],
                [f"Item{i}_2", f"{i*10 + 2}"]
            ]
            
            csv_path = temp_dir / f"batch_{i}.csv"
            with open(csv_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerows(csv_data)
            
            csv_files.append(str(csv_path))
        
        excel_path = temp_dir / "batch_output.xlsx"
        
        response = manager.execute_operation(
            "batch_import",
            csv_files=csv_files,
            excel_path=str(excel_path),
            sheet_prefix="Data"
        )
        
        assert response.success
        assert response.operation == "batch_import"
        assert response.data["total_files"] == 3
        assert response.data["successful_imports"] == 3
        assert response.data["failed_imports"] == 0
        assert response.data["total_rows_imported"] == 9  # 3 files * 3 rows each
        
        # Verify Excel file and sheets
        assert excel_path.exists()
        
        from openpyxl import load_workbook
        wb = load_workbook(excel_path)
        
        # Should have 3 sheets
        assert len(wb.sheetnames) == 3
        assert "Data_1_batch_0" in wb.sheetnames
        assert "Data_2_batch_1" in wb.sheetnames
        assert "Data_3_batch_2" in wb.sheetnames
        
        # Verify data in first sheet
        ws = wb["Data_1_batch_0"]
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Item0_1"
        assert ws.cell(2, 2).value == "1"
        
        wb.close()
    
    def test_batch_import_with_missing_files(self, manager, temp_dir, sample_csv_file):
        """Test batch import with some missing files."""
        csv_files = [
            str(sample_csv_file),
            str(temp_dir / "nonexistent.csv")
        ]
        
        excel_path = temp_dir / "batch_output.xlsx"
        
        response = manager.execute_operation(
            "batch_import",
            csv_files=csv_files,
            excel_path=str(excel_path)
        )
        
        assert response.success  # Overall operation succeeds
        assert response.data["successful_imports"] == 1
        assert response.data["failed_imports"] == 1
        
        # Check import results
        import_results = response.data["import_results"]
        assert len(import_results) == 2
        assert import_results[0]["success"] == True
        assert import_results[1]["success"] == False
        assert "not found" in import_results[1]["error"].lower()
    
    def test_batch_import_empty_list(self, manager, temp_dir):
        """Test batch import with empty file list."""
        excel_path = temp_dir / "batch_output.xlsx"
        
        response = manager.execute_operation(
            "batch_import",
            csv_files=[],
            excel_path=str(excel_path)
        )
        
        assert not response.success
        assert "No CSV files provided" in response.message
    
    def test_helper_methods(self, manager, temp_dir):
        """Test helper methods functionality."""
        # Test CSV delimiter detection
        csv_data = "Name;Age;City\nAlice;25;New York"
        csv_path = temp_dir / "test_delimiter.csv"
        csv_path.write_text(csv_data, encoding='utf-8')
        
        delimiter = manager._detect_csv_delimiter(str(csv_path), 'utf-8')
        assert delimiter == ';'
        
        # Test HTML escaping
        escaped = manager._escape_html('<script>alert("test")</script>')
        assert '&lt;script&gt;' in escaped
        assert '&quot;test&quot;' in escaped
        
        # Test number detection
        assert manager._is_number(42) == True
        assert manager._is_number(3.14) == True
        assert manager._is_number("42") == False
        assert manager._is_number(True) == False  # bool is not considered number
        
        # Test cell value cleaning
        assert manager._clean_cell_value(None) == ''
        assert manager._clean_cell_value('  test  ') == 'test'
        assert manager._clean_cell_value('null') == ''
        assert manager._clean_cell_value('N/A') == ''
    
    def test_parameter_validation(self, manager):
        """Test parameter validation for operations."""
        # Test missing required parameters
        response = manager.execute_operation("import_csv")
        assert not response.success
        assert "Missing required parameters" in response.message
        
        # Test invalid operation
        response = manager.execute_operation("invalid_operation")
        assert not response.success
        assert "not supported" in response.message
    
    def test_file_stats_generation(self, manager, sample_csv_file):
        """Test CSV file statistics generation."""
        stats = manager._get_csv_file_stats(str(sample_csv_file), 'utf-8')
        
        assert "file_size_bytes" in stats
        assert "file_size_mb" in stats
        assert "total_lines" in stats
        assert "estimated_rows" in stats
        assert stats["total_lines"] == 4  # Header + 3 data rows
        assert stats["estimated_rows"] == 3  # Excluding header
    
    def test_json_data_formatting(self, manager):
        """Test JSON data formatting for different styles."""
        sample_data = [
            ["Name", "Age", "City"],
            ["Alice", "25", "New York"],
            ["Bob", "30", "Los Angeles"]
        ]
        
        # Test records format
        records = manager._format_json_data(sample_data, "records", True)
        assert isinstance(records, list)
        assert len(records) == 2
        assert records[0]["Name"] == "Alice"
        assert records[0]["Age"] == "25"
        
        # Test values format
        values = manager._format_json_data(sample_data, "values", True)
        assert isinstance(values, list)
        assert values == sample_data
        
        # Test index format
        index = manager._format_json_data(sample_data, "index", True)
        assert isinstance(index, dict)
        assert index[0] == ["Name", "Age", "City"]
        assert index[1] == ["Alice", "25", "New York"]


if __name__ == "__main__":
    pytest.main([__file__])