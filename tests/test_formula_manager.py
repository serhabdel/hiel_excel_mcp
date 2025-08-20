"""
Tests for FormulaManager tool.

Tests all formula operations including applying formulas, validating syntax,
and batch formula operations.
"""

import pytest
import json
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from hiel_excel_mcp.tools.formula_manager import FormulaManager, formula_manager_tool


class TestFormulaManager:
    """Test suite for FormulaManager tool."""
    
    @pytest.fixture
    def formula_manager(self):
        """Create FormulaManager instance for testing."""
        return FormulaManager()
    
    @pytest.fixture
    def sample_workbook(self):
        """Create a sample workbook for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add sample data for formulas to reference
            data = [
                ["Value1", "Value2", "Result"],
                [10, 20, None],
                [15, 25, None],
                [30, 40, None]
            ]
            
            for row_idx, row_data in enumerate(data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    @pytest.fixture
    def empty_workbook(self):
        """Create an empty workbook for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_tool_metadata(self, formula_manager):
        """Test tool metadata and operation registration."""
        assert formula_manager.get_tool_name() == "formula_manager"
        assert "formula operations" in formula_manager.get_tool_description().lower()
        
        operations = formula_manager.get_available_operations()
        expected_operations = ["apply_formula", "validate_formula", "batch_apply_formulas"]
        
        for op in expected_operations:
            assert op in operations
    
    def test_apply_formula_basic(self, formula_manager, sample_workbook):
        """Test basic formula application."""
        response = formula_manager.execute_operation(
            "apply_formula",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="C2",
            formula="=A2+B2"
        )
        
        assert response.success
        assert response.operation == "apply_formula"
        assert response.data["cell"] == "C2"
        assert response.data["formula"] == "=A2+B2"
        assert response.data["applied_successfully"] is True
    
    def test_apply_formula_without_equals(self, formula_manager, sample_workbook):
        """Test formula application without leading equals sign."""
        response = formula_manager.execute_operation(
            "apply_formula",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="C3",
            formula="A3*B3"
        )
        
        assert response.success
        assert response.data["formula"] == "=A3*B3"  # Should add the equals sign
    
    def test_apply_formula_complex(self, formula_manager, sample_workbook):
        """Test complex formula application."""
        response = formula_manager.execute_operation(
            "apply_formula",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="C4",
            formula="=SUM(A2:A4)+AVERAGE(B2:B4)"
        )
        
        assert response.success
        assert response.data["formula"] == "=SUM(A2:A4)+AVERAGE(B2:B4)"
    
    def test_apply_formula_invalid_cell(self, formula_manager, sample_workbook):
        """Test formula application with invalid cell reference."""
        response = formula_manager.execute_operation(
            "apply_formula",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="INVALID",
            formula="=A1+B1"
        )
        
        assert not response.success
        assert "invalid cell reference" in response.message.lower()
    
    def test_apply_formula_invalid_sheet(self, formula_manager, sample_workbook):
        """Test formula application with invalid sheet name."""
        response = formula_manager.execute_operation(
            "apply_formula",
            filepath=sample_workbook,
            sheet_name="NonExistentSheet",
            cell="A1",
            formula="=1+1"
        )
        
        assert not response.success
        assert "not found" in response.message.lower()
    
    def test_apply_formula_unsafe_formula(self, formula_manager, sample_workbook):
        """Test formula application with unsafe formula."""
        response = formula_manager.execute_operation(
            "apply_formula",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="A1",
            formula="=INDIRECT(A1)"
        )
        
        assert not response.success
        assert "unsafe function" in response.message.lower()
    
    def test_validate_formula_valid(self, formula_manager):
        """Test validation of valid formula."""
        response = formula_manager.execute_operation(
            "validate_formula",
            formula="=SUM(A1:A10)"
        )
        
        assert response.success
        assert response.operation == "validate_formula"
        assert response.data["is_valid"] is True
        assert response.data["syntax_check"] == "passed"
    
    def test_validate_formula_invalid_syntax(self, formula_manager):
        """Test validation of formula with invalid syntax."""
        response = formula_manager.execute_operation(
            "validate_formula",
            formula="=SUM(A1:A10"  # Missing closing parenthesis
        )
        
        assert not response.success
        assert response.data["is_valid"] is False
        assert response.data["syntax_check"] == "failed"
        assert "parenthesis" in response.data["validation_message"].lower()
    
    def test_validate_formula_unsafe(self, formula_manager):
        """Test validation of unsafe formula."""
        response = formula_manager.execute_operation(
            "validate_formula",
            formula="=HYPERLINK(A1)"
        )
        
        assert not response.success
        assert response.data["is_valid"] is False
        assert "unsafe function" in response.data["validation_message"].lower()
    
    def test_validate_formula_without_equals(self, formula_manager):
        """Test validation of formula without equals sign."""
        response = formula_manager.execute_operation(
            "validate_formula",
            formula="SUM(A1:A10)"
        )
        
        assert not response.success
        assert response.data["is_valid"] is False
        assert "must start with" in response.data["validation_message"].lower()
    
    def test_validate_formula_with_context(self, formula_manager, sample_workbook):
        """Test formula validation with file context."""
        response = formula_manager.execute_operation(
            "validate_formula",
            formula="=A1+B1",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="C1"
        )
        
        assert response.success
        assert response.data["is_valid"] is True
        assert "context_validation" in response.data
        assert response.data["filepath"] == sample_workbook
    
    def test_validate_formula_context_invalid_file(self, formula_manager):
        """Test formula validation with invalid file context."""
        response = formula_manager.execute_operation(
            "validate_formula",
            formula="=A1+B1",
            filepath="nonexistent.xlsx",
            sheet_name="Sheet1",
            cell="C1"
        )
        
        assert response.success  # Basic validation should still pass
        assert response.data["is_valid"] is True
        assert "context_validation_error" in response.data
    
    def test_batch_apply_formulas_basic(self, formula_manager, sample_workbook):
        """Test basic batch formula application."""
        formulas = [
            {"cell": "C2", "formula": "=A2+B2"},
            {"cell": "C3", "formula": "=A3+B3"},
            {"cell": "C4", "formula": "=A4+B4"}
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=formulas
        )
        
        assert response.success
        assert response.operation == "batch_apply_formulas"
        assert response.data["total_formulas"] == 3
        assert response.data["successful_applications"] == 3
        assert response.data["failed_applications"] == 0
        
        # Check individual results
        results = response.data["results"]
        assert len(results) == 3
        for result in results:
            assert result["status"] == "success"
    
    def test_batch_apply_formulas_mixed_results(self, formula_manager, sample_workbook):
        """Test batch formula application with mixed success/failure."""
        formulas = [
            {"cell": "C2", "formula": "=A2+B2"},  # Valid
            {"cell": "INVALID", "formula": "=A3+B3"},  # Invalid cell
            {"cell": "C4", "formula": "=INDIRECT(A1)"},  # Unsafe formula
            {"cell": "C5", "formula": "=A5+B5"}  # Valid
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=formulas
        )
        
        # Should have partial success
        assert response.success  # Some formulas succeeded
        assert response.data["total_formulas"] == 4
        assert response.data["successful_applications"] == 2
        assert response.data["failed_applications"] == 2
        assert len(response.errors) == 2
        
        # Check individual results
        results = response.data["results"]
        assert len(results) == 4
        assert results[0]["status"] == "success"
        assert results[1]["status"] == "error"
        assert results[2]["status"] == "error"
        assert results[3]["status"] == "success"
    
    def test_batch_apply_formulas_all_failures(self, formula_manager, sample_workbook):
        """Test batch formula application with all failures."""
        formulas = [
            {"cell": "INVALID1", "formula": "=A1+B1"},
            {"cell": "INVALID2", "formula": "=A2+B2"}
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=formulas
        )
        
        assert not response.success
        assert response.data["successful_applications"] == 0
        assert response.data["failed_applications"] == 2
        assert len(response.errors) == 2
    
    def test_batch_apply_formulas_empty_list(self, formula_manager, sample_workbook):
        """Test batch formula application with empty formula list."""
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=[]
        )
        
        assert not response.success
        assert "no formulas provided" in response.message.lower()
    
    def test_batch_apply_formulas_invalid_format(self, formula_manager, sample_workbook):
        """Test batch formula application with invalid formula format."""
        formulas = [
            {"cell": "A1"},  # Missing formula key
            "invalid_format"  # Not a dictionary
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=formulas
        )
        
        assert not response.success
        assert "must contain" in response.message.lower()
    
    def test_batch_apply_formulas_invalid_sheet(self, formula_manager, sample_workbook):
        """Test batch formula application with invalid sheet."""
        formulas = [
            {"cell": "A1", "formula": "=1+1"}
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="NonExistentSheet",
            formulas=formulas
        )
        
        assert not response.success
        assert "not found" in response.message.lower()
    
    def test_batch_apply_formulas_without_equals(self, formula_manager, sample_workbook):
        """Test batch formula application with formulas missing equals signs."""
        formulas = [
            {"cell": "C2", "formula": "A2+B2"},  # No equals
            {"cell": "C3", "formula": "=A3*B3"}  # With equals
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=formulas
        )
        
        assert response.success
        assert response.data["successful_applications"] == 2
        
        # Check that equals was added to first formula
        results = response.data["results"]
        assert "=A2+B2" in str(results[0])
    
    def test_invalid_operation(self, formula_manager):
        """Test handling of invalid operation."""
        response = formula_manager.execute_operation(
            "invalid_operation",
            filepath="test.xlsx"
        )
        
        assert not response.success
        assert "not supported" in response.message.lower()
    
    def test_missing_required_parameters(self, formula_manager):
        """Test handling of missing required parameters."""
        response = formula_manager.execute_operation(
            "apply_formula"
            # Missing required parameters
        )
        
        assert not response.success
        assert "missing required parameters" in response.message.lower()
    
    def test_tool_function_wrapper(self, formula_manager, sample_workbook):
        """Test the MCP tool function wrapper."""
        result_json = formula_manager_tool(
            operation="apply_formula",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            cell="C2",
            formula="=A2+B2"
        )
        
        result = json.loads(result_json)
        assert result["success"] is True
        assert result["operation"] == "apply_formula"
        assert result["data"]["applied_successfully"] is True
    
    def test_tool_function_error_handling(self):
        """Test error handling in tool function wrapper."""
        result_json = formula_manager_tool(
            operation="apply_formula",
            filepath="nonexistent.xlsx",
            sheet_name="Sheet1",
            cell="A1",
            formula="=1+1"
        )
        
        result = json.loads(result_json)
        assert result["success"] is False
        assert "error" in result["message"].lower()
    
    def test_apply_formula_new_file(self, formula_manager):
        """Test applying formula to new file."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            tmp_path = tmp.name
        
        try:
            # Remove the file so we can test creation
            os.unlink(tmp_path)
            
            response = formula_manager.execute_operation(
                "apply_formula",
                filepath=tmp_path,
                sheet_name="NewSheet",
                cell="A1",
                formula="=1+1"
            )
            
            assert response.success
            assert os.path.exists(tmp_path)
            
        finally:
            if os.path.exists(tmp_path):
                os.unlink(tmp_path)
    
    def test_validate_formula_edge_cases(self, formula_manager):
        """Test formula validation with edge cases."""
        test_cases = [
            ("=1", True),  # Simple number
            ("=A1", True),  # Simple cell reference
            ("=SUM()", True),  # Function with no arguments
            ("=IF(A1>0,\"Yes\",\"No\")", True),  # Function with string literals
            ("=A1:B10", True),  # Range reference
            ("=(A1+B1)*2", True),  # Parentheses
            ("=", False),  # Just equals
            ("==A1", False),  # Double equals
            ("=A1+", False),  # Incomplete expression
        ]
        
        for formula, expected_valid in test_cases:
            response = formula_manager.execute_operation(
                "validate_formula",
                formula=formula
            )
            
            assert response.data["is_valid"] == expected_valid, f"Formula '{formula}' validation failed"
    
    def test_batch_apply_complex_formulas(self, formula_manager, sample_workbook):
        """Test batch application of complex formulas."""
        formulas = [
            {"cell": "D2", "formula": "=IF(A2>B2,A2,B2)"},  # IF function
            {"cell": "D3", "formula": "=ROUND(A3/B3,2)"},  # ROUND function
            {"cell": "D4", "formula": "=CONCATENATE(\"Sum: \",A4+B4)"},  # CONCATENATE
            {"cell": "D5", "formula": "=SUM(A2:A4)*AVERAGE(B2:B4)"}  # Multiple functions
        ]
        
        response = formula_manager.execute_operation(
            "batch_apply_formulas",
            filepath=sample_workbook,
            sheet_name="TestSheet",
            formulas=formulas
        )
        
        assert response.success
        assert response.data["successful_applications"] == 4
        assert response.data["failed_applications"] == 0
    
    def test_formula_validation_performance(self, formula_manager):
        """Test formula validation performance with many formulas."""
        # Test with a reasonable number of formulas
        for i in range(50):
            response = formula_manager.execute_operation(
                "validate_formula",
                formula=f"=A{i}+B{i}"
            )
            assert response.success
            assert response.data["is_valid"] is True