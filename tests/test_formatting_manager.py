"""
Tests for FormattingManager tool.

Tests all formatting operations including basic cell formatting,
conditional formatting rules, and advanced styling capabilities.
"""

import pytest
import json
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from hiel_excel_mcp.tools.formatting_manager import FormattingManager, formatting_manager_tool


class TestFormattingManager:
    """Test suite for FormattingManager tool."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            
            # Add some test data
            ws['A1'] = "Header 1"
            ws['B1'] = "Header 2"
            ws['C1'] = "Header 3"
            ws['A2'] = 10
            ws['B2'] = 20
            ws['C2'] = 30
            ws['A3'] = 15
            ws['B3'] = 25
            ws['C3'] = 35
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    @pytest.fixture
    def formatting_manager(self):
        """Create FormattingManager instance."""
        return FormattingManager()
    
    def test_tool_metadata(self, formatting_manager):
        """Test tool metadata and operation registration."""
        assert formatting_manager.get_tool_name() == "formatting_manager"
        assert "formatting" in formatting_manager.get_tool_description().lower()
        
        operations = formatting_manager.get_available_operations()
        expected_operations = [
            "apply_formatting",
            "add_conditional_formatting", 
            "remove_conditional_formatting",
            "list_conditional_formatting",
            "create_highlight_rule"
        ]
        
        for op in expected_operations:
            assert op in operations
    
    def test_apply_formatting_basic(self, formatting_manager, temp_excel_file):
        """Test basic cell formatting application."""
        response = formatting_manager.execute_operation(
            "apply_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1",
            bold=True,
            italic=True,
            font_size=14,
            font_color="FF0000",
            bg_color="FFFF00"
        )
        
        assert response.success
        assert response.operation == "apply_formatting"
        assert "A1" in response.message
        assert response.data["range"] == "A1"
        assert response.data["formatting_applied"]["bold"] is True
        assert response.data["formatting_applied"]["italic"] is True
        assert response.data["formatting_applied"]["font_size"] == 14
    
    def test_apply_formatting_range(self, formatting_manager, temp_excel_file):
        """Test formatting application to a range."""
        response = formatting_manager.execute_operation(
            "apply_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1",
            end_cell="C1",
            bold=True,
            border_style="thin",
            border_color="000000",
            alignment="center"
        )
        
        assert response.success
        assert response.operation == "apply_formatting"
        assert "A1:C1" in response.data["range"]
        assert response.data["formatting_applied"]["bold"] is True
        assert response.data["formatting_applied"]["border_style"] == "thin"
        assert response.data["formatting_applied"]["alignment"] == "center"
    
    def test_apply_formatting_with_merge(self, formatting_manager, temp_excel_file):
        """Test formatting with cell merging."""
        response = formatting_manager.execute_operation(
            "apply_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A4",
            end_cell="C4",
            merge_cells=True,
            alignment="center",
            bg_color="E0E0E0"
        )
        
        assert response.success
        assert response.data["formatting_applied"]["merge_cells"] is True
        assert response.data["formatting_applied"]["alignment"] == "center"
    
    def test_add_conditional_formatting_cell_is(self, formatting_manager, temp_excel_file):
        """Test adding cell_is conditional formatting rule."""
        rule_config = {
            "type": "cell_is",
            "operator": "greaterThan",
            "formula": ["20"],
            "format": {
                "fill": {"color": "FF0000"}
            }
        }
        
        response = formatting_manager.execute_operation(
            "add_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="A2:C3",
            rule_config=rule_config
        )
        
        assert response.success
        assert response.operation == "add_conditional_formatting"
        assert response.data["range"] == "A2:C3"
        assert response.data["rule_type"] == "cell_is"
        assert response.data["rule_config"]["operator"] == "greaterThan"
    
    def test_add_conditional_formatting_formula(self, formatting_manager, temp_excel_file):
        """Test adding formula-based conditional formatting rule."""
        rule_config = {
            "type": "formula",
            "formula": "MOD(ROW(),2)=0",
            "format": {
                "fill": {"color": "F0F0F0"}
            }
        }
        
        response = formatting_manager.execute_operation(
            "add_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="A1:C10",
            rule_config=rule_config
        )
        
        assert response.success
        assert response.data["rule_type"] == "formula"
        assert response.data["rule_config"]["formula"] == "MOD(ROW(),2)=0"
    
    def test_add_conditional_formatting_color_scale(self, formatting_manager, temp_excel_file):
        """Test adding color scale conditional formatting rule."""
        rule_config = {
            "type": "color_scale",
            "start_type": "min",
            "start_color": "FF0000",
            "end_type": "max", 
            "end_color": "00FF00"
        }
        
        response = formatting_manager.execute_operation(
            "add_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="A2:C3",
            rule_config=rule_config
        )
        
        assert response.success
        assert response.data["rule_type"] == "color_scale"
    
    def test_list_conditional_formatting(self, formatting_manager, temp_excel_file):
        """Test listing conditional formatting rules."""
        # First add a rule
        rule_config = {
            "type": "cell_is",
            "operator": "equal",
            "formula": ["25"],
            "format": {"fill": {"color": "FFFF00"}}
        }
        
        formatting_manager.execute_operation(
            "add_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="B2:B3",
            rule_config=rule_config
        )
        
        # Now list the rules
        response = formatting_manager.execute_operation(
            "list_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet"
        )
        
        assert response.success
        assert response.operation == "list_conditional_formatting"
        assert response.data["total_rules"] >= 1
        assert "conditional_formatting" in response.data
    
    def test_remove_conditional_formatting_range(self, formatting_manager, temp_excel_file):
        """Test removing conditional formatting from specific range."""
        # First add a rule
        rule_config = {
            "type": "cell_is",
            "operator": "greaterThan",
            "formula": ["15"],
            "format": {"fill": {"color": "FF0000"}}
        }
        
        formatting_manager.execute_operation(
            "add_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="A2:A3",
            rule_config=rule_config
        )
        
        # Remove from specific range
        response = formatting_manager.execute_operation(
            "remove_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="A2:A3"
        )
        
        assert response.success
        assert response.operation == "remove_conditional_formatting"
        assert response.data["range"] == "A2:A3"
    
    def test_remove_conditional_formatting_entire_sheet(self, formatting_manager, temp_excel_file):
        """Test removing all conditional formatting from sheet."""
        # First add multiple rules
        rule_configs = [
            {
                "type": "cell_is",
                "operator": "greaterThan",
                "formula": ["20"],
                "format": {"fill": {"color": "FF0000"}}
            },
            {
                "type": "cell_is", 
                "operator": "lessThan",
                "formula": ["15"],
                "format": {"fill": {"color": "0000FF"}}
            }
        ]
        
        for i, rule_config in enumerate(rule_configs):
            formatting_manager.execute_operation(
                "add_conditional_formatting",
                filepath=temp_excel_file,
                sheet_name="TestSheet",
                range_ref=f"A{i+2}:C{i+2}",
                rule_config=rule_config
            )
        
        # Remove all formatting
        response = formatting_manager.execute_operation(
            "remove_conditional_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet"
        )
        
        assert response.success
        assert response.data["range"] == "entire sheet"
        assert response.data["removed_count"] >= 0
    
    def test_create_highlight_rule(self, formatting_manager, temp_excel_file):
        """Test creating simple highlight rule."""
        response = formatting_manager.execute_operation(
            "create_highlight_rule",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="A2:C3",
            operator="greaterThan",
            value=20,
            highlight_color="FFFF00"
        )
        
        assert response.success
        assert response.operation == "create_highlight_rule"
        assert response.data["range"] == "A2:C3"
        assert response.data["operator"] == "greaterThan"
        assert response.data["value"] == 20
        assert response.data["highlight_color"] == "FFFF00"
        assert response.data["rule_type"] == "cell_is"
    
    def test_create_highlight_rule_default_color(self, formatting_manager, temp_excel_file):
        """Test creating highlight rule with default color."""
        response = formatting_manager.execute_operation(
            "create_highlight_rule",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            range_ref="B2:B3",
            operator="equal",
            value="25"
        )
        
        assert response.success
        assert response.data["highlight_color"] == "FFFF00"  # Default yellow
    
    def test_invalid_operation(self, formatting_manager):
        """Test handling of invalid operation."""
        response = formatting_manager.execute_operation(
            "invalid_operation",
            filepath="test.xlsx"
        )
        
        assert not response.success
        assert "not supported" in response.message
        assert "invalid_operation" in response.errors[0]
    
    def test_missing_required_parameters(self, formatting_manager):
        """Test handling of missing required parameters."""
        response = formatting_manager.execute_operation(
            "apply_formatting",
            filepath="test.xlsx"
            # Missing sheet_name and start_cell
        )
        
        assert not response.success
        assert "Missing required parameters" in response.message
        assert "sheet_name" in response.errors[0]
        assert "start_cell" in response.errors[0]
    
    def test_invalid_file_path(self, formatting_manager):
        """Test handling of invalid file path."""
        response = formatting_manager.execute_operation(
            "apply_formatting",
            filepath="/nonexistent/path/file.xlsx",
            sheet_name="Sheet1",
            start_cell="A1"
        )
        
        assert not response.success
        assert "apply_formatting" in response.operation


class TestFormattingManagerTool:
    """Test suite for formatting_manager_tool function."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file for testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestSheet"
            ws['A1'] = "Test"
            ws['A2'] = 100
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_tool_function_success(self, temp_excel_file):
        """Test successful tool function call."""
        result_json = formatting_manager_tool(
            operation="apply_formatting",
            filepath=temp_excel_file,
            sheet_name="TestSheet",
            start_cell="A1",
            bold=True
        )
        
        result = json.loads(result_json)
        assert result["success"] is True
        assert result["operation"] == "apply_formatting"
        assert "data" in result
    
    def test_tool_function_error(self):
        """Test tool function error handling."""
        result_json = formatting_manager_tool(
            operation="apply_formatting",
            filepath="/invalid/path.xlsx",
            sheet_name="Sheet1",
            start_cell="A1"
        )
        
        result = json.loads(result_json)
        assert result["success"] is False
        assert "errors" in result
    
    def test_tool_function_invalid_operation(self):
        """Test tool function with invalid operation."""
        result_json = formatting_manager_tool(
            operation="nonexistent_operation",
            filepath="test.xlsx"
        )
        
        result = json.loads(result_json)
        assert result["success"] is False
        assert "not supported" in result["message"]


if __name__ == "__main__":
    pytest.main([__file__])