"""
Test suite for Batch Manager tool.
Tests batch operations and template processing functionality.
"""

import pytest
import tempfile
import os
import json
from pathlib import Path
from openpyxl import Workbook

from ..tools.batch_manager import BatchManager


@pytest.fixture
def temp_workbooks():
    """Create multiple temporary Excel workbooks for testing."""
    workbooks = []
    for i in range(3):
        with tempfile.NamedTemporaryFile(suffix=f'_test_{i}.xlsx', delete=False) as tmp_file:
            wb = Workbook()
            ws = wb.active
            ws.title = f"TestSheet{i}"
            
            # Add some test data
            ws['A1'] = f"Test Data {i}"
            ws['B1'] = f"More Data {i}"
            ws['A2'] = 100 + i
            ws['B2'] = 200 + i
            
            wb.save(tmp_file.name)
            wb.close()
            workbooks.append(tmp_file.name)
    
    yield workbooks
    
    # Clean up
    for workbook in workbooks:
        try:
            os.unlink(workbook)
        except:
            pass


@pytest.fixture
def temp_template():
    """Create a temporary template file for testing."""
    with tempfile.NamedTemporaryFile(suffix='_template.xlsx', delete=False) as tmp_file:
        wb = Workbook()
        ws = wb.active
        ws.title = "Template"
        
        # Add template placeholders
        ws['A1'] = "{{title}}"
        ws['A2'] = "Name: {{name}}"
        ws['A3'] = "Date: {{date}}"
        ws['A4'] = "{{#if show_table}}Table Data:{{/if}}"
        
        wb.save(tmp_file.name)
        wb.close()
        
        yield tmp_file.name
        
        # Clean up
        try:
            os.unlink(tmp_file.name)
        except:
            pass


class TestBatchManager:
    """Test cases for BatchManager."""
    
    def test_batch_create_workbooks(self):
        """Test batch workbook creation."""
        manager = BatchManager()
        
        # Create temporary file paths
        temp_paths = []
        for i in range(3):
            temp_file = tempfile.NamedTemporaryFile(suffix=f'_batch_{i}.xlsx', delete=False)
            temp_paths.append(temp_file.name)
            temp_file.close()
        
        try:
            result = manager.batch_create_workbooks(temp_paths)
            assert result["success"] is True
            assert "operation_id" in result
            assert result["total_files"] == 3
            
        finally:
            # Clean up
            for path in temp_paths:
                try:
                    os.unlink(path)
                except:
                    pass
    
    def test_fill_template(self, temp_template):
        """Test single template filling."""
        manager = BatchManager()
        
        with tempfile.NamedTemporaryFile(suffix='_output.xlsx', delete=False) as output_file:
            output_path = output_file.name
        
        try:
            data = {
                "title": "Test Report",
                "name": "John Doe",
                "date": "2024-01-15",
                "show_table": True
            }
            
            result = manager.fill_template(temp_template, output_path, data)
            assert result["success"] is True
            assert result["template_path"] == temp_template
            assert result["output_path"] == output_path
            assert result["total_filled_cells"] >= 1
            
        finally:
            try:
                os.unlink(output_path)
            except:
                pass
    
    def test_fill_table_template(self, temp_template):
        """Test table template filling."""
        manager = BatchManager()
        
        with tempfile.NamedTemporaryFile(suffix='_table_output.xlsx', delete=False) as output_file:
            output_path = output_file.name
        
        try:
            table_data = [
                {"name": "Alice", "age": 30, "city": "New York"},
                {"name": "Bob", "age": 25, "city": "Los Angeles"},
                {"name": "Charlie", "age": 35, "city": "Chicago"}
            ]
            
            result = manager.fill_table_template(temp_template, output_path, table_data)
            assert result["success"] is True
            assert result["rows_filled"] == 3
            assert result["columns_filled"] == 3
            
        finally:
            try:
                os.unlink(output_path)
            except:
                pass
    
    def test_generate_report_template(self):
        """Test report template generation."""
        manager = BatchManager()
        
        with tempfile.NamedTemporaryFile(suffix='_report_template.xlsx', delete=False) as output_file:
            output_path = output_file.name
        
        try:
            report_config = {
                "title": "Monthly Report",
                "metadata": ["author", "date", "department"],
                "table": {
                    "headers": ["Item", "Quantity", "Price", "Total"],
                    "placeholder_rows": 5
                }
            }
            
            result = manager.generate_report_template(output_path, report_config)
            assert result["success"] is True
            assert result["template_path"] == output_path
            assert "sections" in result
            
        finally:
            try:
                os.unlink(output_path)
            except:
                pass
    
    def test_batch_fill_templates(self, temp_template):
        """Test batch template filling."""
        manager = BatchManager()
        
        output_paths = []
        for i in range(2):
            temp_file = tempfile.NamedTemporaryFile(suffix=f'_batch_template_{i}.xlsx', delete=False)
            output_paths.append(temp_file.name)
            temp_file.close()
        
        try:
            template_configs = [
                {
                    "template_path": temp_template,
                    "output_path": output_paths[0],
                    "data": {"title": "Report 1", "name": "Alice", "date": "2024-01-01"}
                },
                {
                    "template_path": temp_template,
                    "output_path": output_paths[1],
                    "data": {"title": "Report 2", "name": "Bob", "date": "2024-01-02"}
                }
            ]
            
            result = manager.batch_fill_templates(template_configs)
            assert result["success"] is True
            assert result["total_templates"] == 2
            assert result["successful"] == 2
            assert result["failed"] == 0
            assert len(result["results"]) == 2
            
        finally:
            for path in output_paths:
                try:
                    os.unlink(path)
                except:
                    pass
    
    def test_batch_generate_reports(self):
        """Test batch report generation."""
        manager = BatchManager()
        
        output_paths = []
        for i in range(2):
            temp_file = tempfile.NamedTemporaryFile(suffix=f'_batch_report_{i}.xlsx', delete=False)
            output_paths.append(temp_file.name)
            temp_file.close()
        
        try:
            report_configs = [
                {
                    "output_path": output_paths[0],
                    "config": {
                        "title": "Sales Report",
                        "metadata": ["period", "region"],
                        "table": {"headers": ["Product", "Sales", "Revenue"]}
                    }
                },
                {
                    "output_path": output_paths[1],
                    "config": {
                        "title": "Inventory Report",
                        "metadata": ["date", "location"],
                        "table": {"headers": ["Item", "Stock", "Value"]}
                    }
                }
            ]
            
            result = manager.batch_generate_reports(report_configs)
            assert result["success"] is True
            assert result["total_reports"] == 2
            assert result["successful"] == 2
            assert result["failed"] == 0
            
        finally:
            for path in output_paths:
                try:
                    os.unlink(path)
                except:
                    pass
    
    def test_batch_apply_formulas(self, temp_workbooks):
        """Test batch formula application."""
        manager = BatchManager()
        
        formula_configs = [
            {
                "filepath": temp_workbooks[0],
                "sheet_name": "TestSheet0",
                "cell": "C1",
                "formula": "=A2+B2"
            },
            {
                "filepath": temp_workbooks[1],
                "sheet_name": "TestSheet1",
                "cell": "C1",
                "formula": "=A2*B2"
            }
        ]
        
        result = manager.batch_apply_formulas(formula_configs)
        assert result["success"] is True
        assert "operation_id" in result
        assert result["total_operations"] == 2
    
    def test_get_batch_status(self):
        """Test getting batch operation status."""
        manager = BatchManager()
        
        # Test with invalid operation ID
        result = manager.get_batch_status("invalid-id")
        # Should return some status (even if operation doesn't exist)
        assert "success" in result
    
    def test_list_batch_operations(self):
        """Test listing batch operations."""
        manager = BatchManager()
        
        result = manager.list_batch_operations()
        assert result["success"] is True
        assert "total_operations" in result
        assert "operations" in result
    
    def test_cancel_batch_operation(self):
        """Test canceling batch operation."""
        manager = BatchManager()
        
        # Test with invalid operation ID
        result = manager.cancel_batch_operation("invalid-id")
        assert result["success"] is False  # Should fail for non-existent operation
        assert result["operation_id"] == "invalid-id"
    
    def test_batch_process_data_import(self):
        """Test batch data processing for import operations."""
        manager = BatchManager()
        
        # Create temporary CSV files
        csv_files = []
        excel_files = []
        
        for i in range(2):
            # Create CSV file
            csv_file = tempfile.NamedTemporaryFile(mode='w', suffix=f'_test_{i}.csv', delete=False)
            csv_file.write("Name,Age,City\n")
            csv_file.write(f"Person{i},2{i},City{i}\n")
            csv_file.close()
            csv_files.append(csv_file.name)
            
            # Create Excel output path
            excel_file = tempfile.NamedTemporaryFile(suffix=f'_import_{i}.xlsx', delete=False)
            excel_files.append(excel_file.name)
            excel_file.close()
        
        try:
            file_configs = [
                {
                    "input_path": csv_files[0],
                    "output_path": excel_files[0],
                    "sheet_name": "Sheet1",
                    "has_header": True
                },
                {
                    "input_path": csv_files[1],
                    "output_path": excel_files[1],
                    "sheet_name": "Sheet1",
                    "has_header": True
                }
            ]
            
            result = manager.batch_process_data("import", file_configs)
            assert result["success"] is True
            assert "operation_id" in result
            assert result["operation_type"] == "import"
            
        finally:
            # Clean up
            for file_path in csv_files + excel_files:
                try:
                    os.unlink(file_path)
                except:
                    pass
    
    def test_batch_process_data_export(self, temp_workbooks):
        """Test batch data processing for export operations."""
        manager = BatchManager()
        
        file_configs = [
            {
                "input_path": temp_workbooks[0],
                "format": "csv",
                "output_path": temp_workbooks[0].replace('.xlsx', '.csv')
            },
            {
                "input_path": temp_workbooks[1],
                "format": "csv",
                "output_path": temp_workbooks[1].replace('.xlsx', '.csv')
            }
        ]
        
        result = manager.batch_process_data("export", file_configs)
        assert result["success"] is True
        assert "operation_id" in result
        assert result["export_format"] == "csv"
    
    def test_error_handling(self):
        """Test error handling for invalid operations."""
        manager = BatchManager()
        
        # Test invalid operation type
        result = manager.batch_process_data("invalid_operation", [])
        assert result["success"] is False
        assert "Unknown operation type" in result["error"]
        
        # Test invalid template path
        result = manager.fill_template("/nonexistent/template.xlsx", "/tmp/output.xlsx", {})
        assert result["success"] is False
        assert "error" in result
    
    def test_batch_export_formats(self, temp_workbooks):
        """Test batch export with different formats."""
        manager = BatchManager()
        
        # Test CSV export
        result = manager.batch_export(temp_workbooks[:2], "csv")
        assert result["success"] is True
        assert result["export_format"] == "csv"
        
        # Test with custom configurations
        export_configs = [
            {"output_path": temp_workbooks[0].replace('.xlsx', '_custom.csv')},
            {"output_path": temp_workbooks[1].replace('.xlsx', '_custom.csv')}
        ]
        
        result = manager.batch_export(temp_workbooks[:2], "csv", export_configs)
        assert result["success"] is True
        assert result["export_format"] == "csv"