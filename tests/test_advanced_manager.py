"""
Test suite for Advanced Manager tool.
Tests named ranges, hyperlinks, and comments operations.
"""

import pytest
import tempfile
import os
from pathlib import Path
from openpyxl import Workbook

from ..tools.advanced_manager import AdvancedManager


@pytest.fixture
def temp_workbook():
    """Create a temporary Excel workbook for testing."""
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp_file:
        wb = Workbook()
        ws = wb.active
        ws.title = "TestSheet"
        
        # Add some test data
        ws['A1'] = "Test Data"
        ws['B1'] = "More Data"
        ws['A2'] = 100
        ws['B2'] = 200
        ws['C2'] = 300
        
        wb.save(tmp_file.name)
        wb.close()
        
        yield tmp_file.name
        
        # Clean up
        try:
            os.unlink(tmp_file.name)
        except:
            pass


class TestAdvancedManager:
    """Test cases for AdvancedManager."""
    
    def test_named_range_operations(self, temp_workbook):
        """Test named range creation, listing, and deletion."""
        manager = AdvancedManager()
        
        # Create named range
        result = manager.create_named_range(
            temp_workbook, "TestRange", "A1:B2", "TestSheet"
        )
        assert result["success"] is True
        assert result["name"] == "TestRange"
        
        # List named ranges
        result = manager.list_named_ranges(temp_workbook)
        assert result["success"] is True
        assert result["total_ranges"] >= 1
        assert any(r["name"] == "TestRange" for r in result["named_ranges"])
        
        # Get named range value
        result = manager.get_named_range_value(temp_workbook, "TestRange")
        assert result["success"] is True
        assert result["name"] == "TestRange"
        assert result["rows"] == 2
        assert result["columns"] == 2
        
        # Delete named range
        result = manager.delete_named_range(temp_workbook, "TestRange")
        assert result["success"] is True
        assert result["name"] == "TestRange"
        
        # Verify deletion
        result = manager.list_named_ranges(temp_workbook)
        assert result["success"] is True
        assert not any(r["name"] == "TestRange" for r in result["named_ranges"])
    
    def test_hyperlink_operations(self, temp_workbook):
        """Test hyperlink creation, listing, and removal."""
        manager = AdvancedManager()
        
        # Add hyperlink
        result = manager.add_hyperlink(
            temp_workbook, "TestSheet", "C1", "https://example.com", "Example Link"
        )
        assert result["success"] is True
        assert result["cell"] == "C1"
        assert result["display_text"] == "Example Link"
        
        # List hyperlinks
        result = manager.list_hyperlinks(temp_workbook, "TestSheet")
        assert result["success"] is True
        assert result["total_hyperlinks"] >= 1
        assert any(h["cell"] == "C1" for h in result["hyperlinks"])
        
        # Remove hyperlink
        result = manager.remove_hyperlink(temp_workbook, "TestSheet", "C1", True)
        assert result["success"] is True
        assert result["cell"] == "C1"
        assert result["text_kept"] is True
    
    def test_comment_operations(self, temp_workbook):
        """Test comment management operations."""
        manager = AdvancedManager()
        
        # Add comment
        result = manager.manage_comments(
            temp_workbook, "add", "TestSheet", "D1", 
            "This is a test comment", "Test Author"
        )
        assert result["success"] is True
        assert result["cell"] == "D1"
        assert result["text"] == "This is a test comment"
        assert result["author"] == "Test Author"
        
        # Get comment
        result = manager.manage_comments(
            temp_workbook, "get", "TestSheet", "D1"
        )
        assert result["success"] is True
        assert result["cell"] == "D1"
        assert result["comment_text"] == "This is a test comment"
        
        # Edit comment
        result = manager.manage_comments(
            temp_workbook, "edit", "TestSheet", "D1", 
            "Updated comment text"
        )
        assert result["success"] is True
        assert result["cell"] == "D1"
        assert result["new_text"] == "Updated comment text"
        
        # Delete comment
        result = manager.manage_comments(
            temp_workbook, "delete", "TestSheet", "D1"
        )
        assert result["success"] is True
        assert result["cell"] == "D1"
        assert result["deleted_text"] == "Updated comment text"
    
    def test_search_operations(self, temp_workbook):
        """Test search across advanced features."""
        manager = AdvancedManager()
        
        # Set up test data
        manager.create_named_range(temp_workbook, "SearchRange", "A1:A2", "TestSheet")
        manager.add_hyperlink(temp_workbook, "TestSheet", "B1", "https://search.com", "Search Link")
        manager.manage_comments(temp_workbook, "add", "TestSheet", "C1", "Search in comment")
        
        # Search named ranges
        result = manager.search_advanced_features(
            temp_workbook, "named_ranges", "Search"
        )
        assert result["success"] is True
        assert result["total_matches"] >= 1
        
        # Search hyperlinks
        result = manager.search_advanced_features(
            temp_workbook, "hyperlinks", "search"
        )
        assert result["success"] is True
        assert result["total_matches"] >= 1
        
        # Search comments
        result = manager.search_advanced_features(
            temp_workbook, "comments", "Search"
        )
        assert result["success"] is True
        assert result["total_matches"] >= 1
    
    def test_advanced_summary(self, temp_workbook):
        """Test comprehensive summary of advanced features."""
        manager = AdvancedManager()
        
        # Set up test data
        manager.create_named_range(temp_workbook, "SummaryRange", "A1:B1", "TestSheet")
        manager.add_hyperlink(temp_workbook, "TestSheet", "C1", "https://example.com")
        manager.manage_comments(temp_workbook, "add", "TestSheet", "D1", "Summary comment")
        
        # Get summary
        result = manager.get_advanced_summary(temp_workbook)
        assert result["success"] is True
        assert "named_ranges" in result
        assert "hyperlinks" in result
        assert "comments" in result
        assert "totals" in result
        
        assert result["totals"]["named_ranges"] >= 1
        assert result["totals"]["hyperlinks"] >= 1
        assert result["totals"]["comments"] >= 1
        assert result["totals"]["total_advanced_features"] >= 3
    
    def test_error_handling(self, temp_workbook):
        """Test error handling for invalid operations."""
        manager = AdvancedManager()
        
        # Test invalid comment action
        with pytest.raises(ValueError, match="Invalid action"):
            manager.manage_comments(
                temp_workbook, "invalid", "TestSheet", "A1"
            )
        
        # Test invalid search type
        with pytest.raises(ValueError, match="Invalid search_type"):
            manager.search_advanced_features(
                temp_workbook, "invalid", "search"
            )
        
        # Test missing text for comment add
        with pytest.raises(ValueError, match="Text is required"):
            manager.manage_comments(
                temp_workbook, "add", "TestSheet", "A1"
            )
    
    def test_case_sensitivity_in_search(self, temp_workbook):
        """Test case sensitivity handling in search operations."""
        manager = AdvancedManager()
        
        # Set up test data with mixed case
        manager.create_named_range(temp_workbook, "CaseRange", "A1:A1", "TestSheet")
        manager.manage_comments(temp_workbook, "add", "TestSheet", "A1", "Case Sensitive Comment")
        
        # Case insensitive search (default)
        result = manager.search_advanced_features(
            temp_workbook, "comments", "CASE"
        )
        assert result["success"] is True
        assert result["total_matches"] >= 1
        
        # Case sensitive search
        result = manager.search_advanced_features(
            temp_workbook, "comments", "CASE", case_sensitive=True
        )
        assert result["success"] is True
        assert result["total_matches"] == 0  # Should not match lowercase "case"
        
        result = manager.search_advanced_features(
            temp_workbook, "comments", "Case", case_sensitive=True
        )
        assert result["success"] is True
        assert result["total_matches"] >= 1  # Should match "Case"
    
    def test_sheet_scope_operations(self, temp_workbook):
        """Test operations scoped to specific sheets."""
        manager = AdvancedManager()
        
        # Add features to specific sheet
        manager.add_hyperlink(temp_workbook, "TestSheet", "A1", "https://example.com")
        manager.manage_comments(temp_workbook, "add", "TestSheet", "B1", "Sheet comment")
        
        # Get summary for specific sheet
        result = manager.get_advanced_summary(temp_workbook, "TestSheet")
        assert result["success"] is True
        assert result["sheet_scope"] == "TestSheet"
        
        # Search in specific sheet
        result = manager.search_advanced_features(
            temp_workbook, "hyperlinks", "example", "TestSheet"
        )
        assert result["success"] is True