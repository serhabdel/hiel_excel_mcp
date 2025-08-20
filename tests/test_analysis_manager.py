"""
Tests for Analysis Manager Tool.

Tests chart creation, pivot table generation, Excel table creation,
and data analysis functionality.
"""

import pytest
import json
import os
import tempfile
from pathlib import Path
from openpyxl import Workbook

from ..tools.analysis_manager import analysis_manager, AnalysisManager
from ..core.base_tool import OperationResponse


class TestAnalysisManager:
    """Test suite for AnalysisManager tool."""
    
    @pytest.fixture
    def temp_excel_file(self):
        """Create a temporary Excel file with test data."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "TestData"
            
            # Add headers
            headers = ["Product", "Sales", "Profit", "Region"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Add test data
            test_data = [
                ["Product A", 1000, 200, "North"],
                ["Product B", 1500, 300, "South"],
                ["Product C", 800, 150, "East"],
                ["Product D", 1200, 250, "West"],
                ["Product E", 900, 180, "North"],
                ["Product F", 1100, 220, "South"]
            ]
            
            for row_idx, row_data in enumerate(test_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    @pytest.fixture
    def analysis_manager_instance(self):
        """Create AnalysisManager instance for testing."""
        return AnalysisManager()
    
    def test_tool_info(self, analysis_manager_instance):
        """Test tool information retrieval."""
        tool_info = analysis_manager_instance.get_tool_info()
        
        assert tool_info["name"] == "analysis_manager"
        assert "analysis and visualization management" in tool_info["description"].lower()
        assert "operations" in tool_info
        
        # Check that all expected operations are present
        expected_operations = ["create_chart", "create_pivot_table", "create_table", "analyze_data"]
        for operation in expected_operations:
            assert operation in tool_info["operations"]
    
    def test_create_chart_operation(self, analysis_manager_instance, temp_excel_file):
        """Test chart creation operation."""
        response = analysis_manager_instance.execute_operation(
            "create_chart",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:C7",
            chart_type="bar",
            target_cell="E2",
            title="Sales Analysis",
            x_axis="Products",
            y_axis="Values"
        )
        
        assert isinstance(response, OperationResponse)
        assert response.success is True
        assert response.operation == "create_chart"
        assert "chart created successfully" in response.message.lower()
        assert response.data is not None
        assert response.data["chart_type"] == "bar"
        assert response.data["target_cell"] == "E2"
    
    def test_create_chart_with_style(self, analysis_manager_instance, temp_excel_file):
        """Test chart creation with custom styling."""
        style = {
            "show_legend": True,
            "show_data_labels": True,
            "legend_position": "b"
        }
        
        response = analysis_manager_instance.execute_operation(
            "create_chart",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="B1:C7",
            chart_type="line",
            target_cell="F2",
            title="Trend Analysis",
            style=style
        )
        
        assert response.success is True
        assert response.data["chart_type"] == "line"
    
    def test_create_pivot_table_operation(self, analysis_manager_instance, temp_excel_file):
        """Test pivot table creation operation."""
        response = analysis_manager_instance.execute_operation(
            "create_pivot_table",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            rows=["Region"],
            values=["Sales", "Profit"],
            agg_func="sum"
        )
        
        assert isinstance(response, OperationResponse)
        assert response.success is True
        assert response.operation == "create_pivot_table"
        assert "summary table created successfully" in response.message.lower()
        assert response.data is not None
        assert response.data["rows"] == ["Region"]
        assert response.data["values"] == ["Sales", "Profit"]
        assert response.data["aggregation"] == "sum"
    
    def test_create_pivot_table_with_columns(self, analysis_manager_instance, temp_excel_file):
        """Test pivot table creation with column grouping."""
        response = analysis_manager_instance.execute_operation(
            "create_pivot_table",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            rows=["Product"],
            values=["Sales"],
            columns=["Region"],
            agg_func="average"
        )
        
        assert response.success is True
        assert response.data["columns"] == ["Region"]
        assert response.data["aggregation"] == "average"
    
    def test_create_table_operation(self, analysis_manager_instance, temp_excel_file):
        """Test Excel table creation operation."""
        response = analysis_manager_instance.execute_operation(
            "create_table",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            table_name="SalesTable",
            table_style="TableStyleMedium2"
        )
        
        assert isinstance(response, OperationResponse)
        assert response.success is True
        assert response.operation == "create_table"
        assert "successfully created table" in response.message.lower()
        assert response.data is not None
        assert response.data["table_name"] == "SalesTable"
        assert response.data["data_range"] == "A1:D7"
    
    def test_create_table_auto_name(self, analysis_manager_instance, temp_excel_file):
        """Test Excel table creation with auto-generated name."""
        response = analysis_manager_instance.execute_operation(
            "create_table",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7"
        )
        
        assert response.success is True
        assert "table_" in response.data["table_name"].lower()
    
    def test_analyze_data_descriptive(self, analysis_manager_instance, temp_excel_file):
        """Test descriptive data analysis."""
        response = analysis_manager_instance.execute_operation(
            "analyze_data",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            analysis_type="descriptive"
        )
        
        assert isinstance(response, OperationResponse)
        assert response.success is True
        assert response.operation == "analyze_data"
        assert "data analysis completed" in response.message.lower()
        assert response.data is not None
        
        analysis_results = response.data["analysis_results"]
        assert "column_analysis" in analysis_results
        assert "Sales" in analysis_results["column_analysis"]
        assert "Profit" in analysis_results["column_analysis"]
        
        # Check that statistical measures are present
        sales_stats = analysis_results["column_analysis"]["Sales"]
        assert "mean" in sales_stats
        assert "median" in sales_stats
        assert "min" in sales_stats
        assert "max" in sales_stats
        assert "count" in sales_stats
    
    def test_analyze_data_correlation(self, analysis_manager_instance, temp_excel_file):
        """Test correlation analysis."""
        response = analysis_manager_instance.execute_operation(
            "analyze_data",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            analysis_type="correlation"
        )
        
        assert response.success is True
        analysis_results = response.data["analysis_results"]
        assert "correlations" in analysis_results
        assert "numeric_columns" in analysis_results
        
        # Should have Sales and Profit as numeric columns
        assert "Sales" in analysis_results["numeric_columns"]
        assert "Profit" in analysis_results["numeric_columns"]
    
    def test_analyze_data_trend(self, analysis_manager_instance, temp_excel_file):
        """Test trend analysis."""
        response = analysis_manager_instance.execute_operation(
            "analyze_data",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            analysis_type="trend"
        )
        
        assert response.success is True
        analysis_results = response.data["analysis_results"]
        assert "trends" in analysis_results
        
        # Check trend information for numeric columns
        if "Sales" in analysis_results["trends"]:
            sales_trend = analysis_results["trends"]["Sales"]
            assert "direction" in sales_trend
            assert "total_change" in sales_trend
            assert "start_value" in sales_trend
            assert "end_value" in sales_trend
    
    def test_analyze_data_with_charts(self, analysis_manager_instance, temp_excel_file):
        """Test data analysis with chart creation."""
        response = analysis_manager_instance.execute_operation(
            "analyze_data",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:D7",
            analysis_type="descriptive",
            include_charts=True
        )
        
        assert response.success is True
        assert "charts_created" in response.data
        # Charts creation might fail due to data format, but should not cause operation failure
    
    def test_invalid_file_path(self, analysis_manager_instance):
        """Test operations with invalid file path."""
        response = analysis_manager_instance.execute_operation(
            "create_chart",
            filepath="/nonexistent/file.xlsx",
            sheet_name="Sheet1",
            data_range="A1:C10",
            chart_type="bar",
            target_cell="E2"
        )
        
        assert response.success is False
        assert "error" in response.status.value.lower()
    
    def test_invalid_chart_type(self, analysis_manager_instance, temp_excel_file):
        """Test chart creation with invalid chart type."""
        response = analysis_manager_instance.execute_operation(
            "create_chart",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:C7",
            chart_type="invalid_type",
            target_cell="E2"
        )
        
        assert response.success is False
    
    def test_invalid_sheet_name(self, analysis_manager_instance, temp_excel_file):
        """Test operations with invalid sheet name."""
        response = analysis_manager_instance.execute_operation(
            "create_pivot_table",
            filepath=temp_excel_file,
            sheet_name="NonexistentSheet",
            data_range="A1:D7",
            rows=["Region"],
            values=["Sales"]
        )
        
        assert response.success is False
    
    def test_missing_required_parameters(self, analysis_manager_instance, temp_excel_file):
        """Test operations with missing required parameters."""
        response = analysis_manager_instance.execute_operation(
            "create_chart",
            filepath=temp_excel_file,
            sheet_name="TestData"
            # Missing required parameters: data_range, chart_type, target_cell
        )
        
        assert response.success is False
        assert "missing required parameters" in response.message.lower()
    
    def test_empty_data_range(self, analysis_manager_instance):
        """Test analysis with empty data range."""
        # Create a file with no data
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            ws = wb.active
            ws.title = "EmptySheet"
            wb.save(tmp.name)
            wb.close()
            
            try:
                response = analysis_manager_instance.execute_operation(
                    "analyze_data",
                    filepath=tmp.name,
                    sheet_name="EmptySheet",
                    data_range="A1:C10",
                    analysis_type="descriptive"
                )
                
                assert response.success is True
                assert "no data found" in response.message.lower() or "warnings" in response.data
                
            finally:
                if os.path.exists(tmp.name):
                    os.unlink(tmp.name)
    
    def test_tool_function_wrapper(self, temp_excel_file):
        """Test the MCP tool function wrapper."""
        result_json = analysis_manager.analysis_manager_tool(
            operation="create_chart",
            filepath=temp_excel_file,
            sheet_name="TestData",
            data_range="A1:C7",
            chart_type="pie",
            target_cell="G2",
            title="Distribution Chart"
        )
        
        result = json.loads(result_json)
        assert result["success"] is True
        assert result["operation"] == "create_chart"
        assert result["data"]["chart_type"] == "pie"
    
    def test_tool_function_error_handling(self):
        """Test error handling in tool function wrapper."""
        result_json = analysis_manager.analysis_manager_tool(
            operation="invalid_operation"
        )
        
        result = json.loads(result_json)
        assert result["success"] is False
        assert "not supported" in result["message"].lower()


class TestAnalysisManagerIntegration:
    """Integration tests for AnalysisManager with real Excel operations."""
    
    @pytest.fixture
    def complex_excel_file(self):
        """Create a more complex Excel file for integration testing."""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            wb = Workbook()
            
            # Create multiple sheets
            ws1 = wb.active
            ws1.title = "SalesData"
            
            # Add comprehensive sales data
            headers = ["Date", "Product", "Category", "Sales", "Profit", "Region", "Salesperson"]
            for col, header in enumerate(headers, 1):
                ws1.cell(row=1, column=col, value=header)
            
            # Add more comprehensive test data
            import datetime
            base_date = datetime.date(2024, 1, 1)
            
            test_data = []
            products = ["Laptop", "Mouse", "Keyboard", "Monitor", "Tablet"]
            categories = ["Electronics", "Accessories", "Electronics", "Electronics", "Electronics"]
            regions = ["North", "South", "East", "West"]
            salespeople = ["Alice", "Bob", "Charlie", "Diana"]
            
            for i in range(50):  # Create 50 rows of data
                date = base_date + datetime.timedelta(days=i)
                product = products[i % len(products)]
                category = categories[i % len(categories)]
                sales = 500 + (i * 23) % 1000  # Varying sales figures
                profit = sales * 0.2 + (i * 7) % 100  # Profit with some variation
                region = regions[i % len(regions)]
                salesperson = salespeople[i % len(salespeople)]
                
                test_data.append([date, product, category, sales, profit, region, salesperson])
            
            for row_idx, row_data in enumerate(test_data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws1.cell(row=row_idx, column=col_idx, value=value)
            
            wb.save(tmp.name)
            wb.close()
            
            yield tmp.name
            
            # Cleanup
            if os.path.exists(tmp.name):
                os.unlink(tmp.name)
    
    def test_comprehensive_analysis_workflow(self, complex_excel_file):
        """Test a complete analysis workflow with charts, pivot tables, and analysis."""
        manager = AnalysisManager()
        
        # Step 1: Create an Excel table
        table_response = manager.execute_operation(
            "create_table",
            filepath=complex_excel_file,
            sheet_name="SalesData",
            data_range="A1:G51",
            table_name="ComprehensiveSalesTable"
        )
        assert table_response.success is True
        
        # Step 2: Create a pivot table
        pivot_response = manager.execute_operation(
            "create_pivot_table",
            filepath=complex_excel_file,
            sheet_name="SalesData",
            data_range="A1:G51",
            rows=["Region", "Product"],
            values=["Sales", "Profit"],
            agg_func="sum"
        )
        assert pivot_response.success is True
        
        # Step 3: Create charts
        chart_response = manager.execute_operation(
            "create_chart",
            filepath=complex_excel_file,
            sheet_name="SalesData",
            data_range="D1:E51",  # Sales and Profit columns
            chart_type="scatter",
            target_cell="I2",
            title="Sales vs Profit Analysis",
            x_axis="Sales",
            y_axis="Profit"
        )
        assert chart_response.success is True
        
        # Step 4: Perform comprehensive data analysis
        analysis_response = manager.execute_operation(
            "analyze_data",
            filepath=complex_excel_file,
            sheet_name="SalesData",
            data_range="A1:G51",
            analysis_type="descriptive",
            include_charts=False
        )
        assert analysis_response.success is True
        
        # Verify analysis results
        analysis_results = analysis_response.data["analysis_results"]
        assert "column_analysis" in analysis_results
        assert "Sales" in analysis_results["column_analysis"]
        assert "Profit" in analysis_results["column_analysis"]
        
        # Check that we have meaningful statistics
        sales_stats = analysis_results["column_analysis"]["Sales"]
        assert sales_stats["count"] > 0
        assert sales_stats["mean"] > 0
        assert sales_stats["min"] <= sales_stats["max"]
    
    def test_correlation_analysis_integration(self, complex_excel_file):
        """Test correlation analysis with real data."""
        manager = AnalysisManager()
        
        response = manager.execute_operation(
            "analyze_data",
            filepath=complex_excel_file,
            sheet_name="SalesData",
            data_range="A1:G51",
            analysis_type="correlation"
        )
        
        assert response.success is True
        analysis_results = response.data["analysis_results"]
        
        # Should find correlations between Sales and Profit
        assert "correlations" in analysis_results
        correlations = analysis_results["correlations"]
        
        # Look for Sales vs Profit correlation
        sales_profit_key = None
        for key in correlations.keys():
            if "sales" in key.lower() and "profit" in key.lower():
                sales_profit_key = key
                break
        
        if sales_profit_key:
            correlation_data = correlations[sales_profit_key]
            assert "correlation" in correlation_data
            assert "strength" in correlation_data
            # Sales and Profit should be positively correlated
            assert correlation_data["correlation"] > 0


if __name__ == "__main__":
    pytest.main([__file__])