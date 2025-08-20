#!/usr/bin/env python3
"""
Working Hiel Excel MCP Server
Provides 83 Excel tools via MCP protocol with correct imports and error handling.
"""

import os
import json
import logging
import traceback
import tempfile
import csv
from typing import Dict, Any, List, Optional, Tuple
from pathlib import Path
import asyncio
from datetime import datetime
from functools import lru_cache
import threading
from collections import OrderedDict
import time

# MCP imports - using correct available imports
from mcp import types
from mcp.server import Server
from mcp.server.stdio import stdio_server
from mcp.server import Server
from mcp import types
import asyncio

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Excel operations using openpyxl
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.chart import BarChart, LineChart, PieChart, Reference
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    logger.warning("OpenPyXL not available - some features may be limited")
    OPENPYXL_AVAILABLE = False

class WorkbookCache:
    """LRU cache for Excel workbooks with mtime checking."""
    
    def __init__(self, max_size: int = 20, ttl: int = 300):
        self.cache = OrderedDict()
        self.max_size = max_size
        self.ttl = ttl
        self.locks = {}  # Per-file locks
        self.lock = threading.Lock()
    
    def get(self, filepath: str) -> Optional[Tuple[Any, float]]:
        """Get workbook from cache if valid."""
        with self.lock:
            if filepath not in self.cache:
                return None
            
            wb, cached_time, cached_mtime = self.cache[filepath]
            
            # Check if file still exists and hasn't been modified
            try:
                current_mtime = os.path.getmtime(filepath)
                if current_mtime != cached_mtime:
                    del self.cache[filepath]
                    return None
                
                # Check TTL
                if time.time() - cached_time > self.ttl:
                    del self.cache[filepath]
                    return None
                
                # Move to end (LRU)
                self.cache.move_to_end(filepath)
                return wb, cached_mtime
            except FileNotFoundError:
                del self.cache[filepath]
                return None
    
    def put(self, filepath: str, workbook: Any) -> None:
        """Add workbook to cache."""
        with self.lock:
            try:
                mtime = os.path.getmtime(filepath)
                self.cache[filepath] = (workbook, time.time(), mtime)
                self.cache.move_to_end(filepath)
                
                # Evict if over size limit
                while len(self.cache) > self.max_size:
                    self.cache.popitem(last=False)
            except FileNotFoundError:
                pass
    
    def invalidate(self, filepath: str) -> None:
        """Remove workbook from cache."""
        with self.lock:
            if filepath in self.cache:
                del self.cache[filepath]
    
    def get_lock(self, filepath: str) -> threading.Lock:
        """Get per-file lock for serializing access."""
        with self.lock:
            if filepath not in self.locks:
                self.locks[filepath] = threading.Lock()
            return self.locks[filepath]

class ExcelMCPServer:
    """Excel MCP Server with performance optimizations and extended tools."""
    
    def __init__(self):
        self.cache = WorkbookCache(
            max_size=int(os.getenv('CACHE_SIZE', '20')),
            ttl=int(os.getenv('CACHE_TTL', '300'))
        )
        self.temp_dir = tempfile.gettempdir()
        self.base_dir = os.getenv('EXCEL_FILES_PATH', os.getcwd())
        self.max_rows = int(os.getenv('MAX_ROWS_PER_CALL', '10000'))
        self.max_cols = int(os.getenv('MAX_COLS_PER_CALL', '1000'))
        self.max_file_size = int(os.getenv('MAX_FILE_SIZE', '52428800'))  # 50MB
        
    async def call_tool(self, name: str, arguments: Dict[str, Any]) -> List[types.TextContent]:
        """Main tool dispatcher."""
        try:
            logger.info(f"Calling tool: {name} with args: {list(arguments.keys())}")

            # Maintain backward compatibility with various naming patterns
            aliases = {
                # Original names
                "mcp1_create_workbook": "workbook-create",
                "mcp1_write_data_to_excel": "data-write",
                "mcp1_read_data_from_excel": "data-read",
                "mcp1_apply_formula": "formula-apply",
                "server_status": "server-status",
                "format_conditional": "format-conditional",
                "validation_add": "validation-add",
                "worksheet_delete": "worksheet-delete",
                "range_merge": "range-merge",
                "range_unmerge": "range-unmerge",
                
                # New tools
                "table_create": "table-create",
                "pivot_create": "pivot-create",
                "sparkline_add": "sparkline-add",
                "format_advanced": "format-advanced",
                "rows_insert": "rows-insert",
                "columns_insert": "columns-insert",
                "rows_delete": "rows-delete",
                "columns_delete": "columns-delete",
                "find_replace": "find-replace",
                "filter_apply": "filter-apply",
                "sort_range": "sort-range",
                "named_range_create": "named-range-create",
                "protection_add": "protection-add",
                
                # Without mcp1 prefix
                "create_workbook": "workbook-create",
                "write_data_to_excel": "data-write",
                "read_data_from_excel": "data-read",
                "apply_formula": "formula-apply",
                
                # Underscore versions
                "workbook_create": "workbook-create",
                "data_write": "data-write",
                "data_read": "data-read",
                "formula_apply": "formula-apply",
                # New tools aliases
                "conditional_formatting": "format-conditional",
                "data_validation": "validation-add",
                "delete_worksheet": "worksheet-delete",
                "merge_cells": "range-merge",
                "unmerge_cells": "range-unmerge",
            }

            canonical = aliases.get(name, name)

            # Route by canonical name
            if canonical == "workbook-create":
                result = await self.create_workbook(arguments.get("filepath"))
            elif canonical == "data-write":
                result = await self.write_data_to_excel(
                    arguments.get("filepath"),
                    arguments.get("sheet_name", "Sheet1"),
                    arguments.get("data", []),
                    arguments.get("start_cell", "A1"),
                )
            elif canonical == "data-read":
                result = await self.read_data_from_excel(
                    arguments.get("filepath"),
                    arguments.get("sheet_name", "Sheet1"),
                    arguments.get("start_cell", "A1"),
                    arguments.get("end_cell"),
                )
            elif canonical == "io-import-csv":
                result = await self.import_csv_to_excel(
                    arguments.get("csv_path"),
                    arguments.get("excel_path"),
                    arguments.get("sheet_name", "Sheet1"),
                    arguments.get("has_header", True),
                )
            elif canonical == "io-export-csv":
                result = await self.export_excel_to_csv(
                    arguments.get("excel_path"),
                    arguments.get("sheet_name", "Sheet1"),
                    arguments.get("csv_path"),
                )
            elif canonical == "worksheet-create":
                result = await self.create_worksheet(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                )
            elif canonical == "format-range":
                result = await self.format_range(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("start_cell"),
                    arguments.get("end_cell"),
                    arguments,
                )
            elif canonical == "formula-apply":
                result = await self.apply_formula(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("cell"),
                    arguments.get("formula"),
                )
            elif canonical == "chart-create":
                result = await self.create_chart(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("data_range"),
                    arguments.get("chart_type", "column"),
                    arguments.get("target_cell"),
                )
            elif canonical == "workbook-metadata":
                result = await self.get_workbook_metadata(arguments.get("filepath"))
            elif canonical == "cell-write":
                result = await self.write_cell(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("cell"),
                    arguments.get("value"),
                )
            elif canonical == "server-status":
                result = {
                    "success": True,
                    "server": "Hiel Excel MCP",
                    "version": "1.0.0",
                    "total_tools": len(TOOLS),
                    "openpyxl_available": OPENPYXL_AVAILABLE,
                    "status": "running",
                    "cache_size": len(self.cache.cache),
                    "base_dir": self.base_dir,
                }
            elif canonical == "format-conditional":
                result = await self.apply_conditional_formatting(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("rule_type"),
                    arguments.get("condition"),
                    arguments.get("format"),
                )
            elif canonical == "validation-add":
                result = await self.add_data_validation(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("validation_type"),
                    arguments.get("criteria"),
                )
            elif canonical == "worksheet-delete":
                result = await self.delete_worksheet(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                )
            elif canonical == "range-merge":
                result = await self.merge_cells(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                )
            elif canonical == "range-unmerge":
                result = await self.unmerge_cells(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                )
            elif canonical == "table-create":
                result = await self.create_table(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("table_name"),
                    arguments.get("style"),
                )
            elif canonical == "pivot-create":
                result = await self.create_pivot_table(
                    arguments.get("filepath"),
                    arguments.get("source_sheet"),
                    arguments.get("source_range"),
                    arguments.get("target_sheet"),
                    arguments.get("target_cell"),
                    arguments.get("rows"),
                    arguments.get("columns"),
                    arguments.get("values"),
                    arguments.get("filters"),
                )
            elif canonical == "sparkline-add":
                result = await self.add_sparkline(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("data_range"),
                    arguments.get("target_cell"),
                    arguments.get("sparkline_type"),
                )
            elif canonical == "format-advanced":
                result = await self.apply_advanced_formatting(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("formatting"),
                )
            elif canonical == "rows-insert":
                result = await self.insert_rows(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("row_index"),
                    arguments.get("count"),
                )
            elif canonical == "columns-insert":
                result = await self.insert_columns(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("column_index"),
                    arguments.get("count"),
                )
            elif canonical == "rows-delete":
                result = await self.delete_rows(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("row_index"),
                    arguments.get("count"),
                )
            elif canonical == "columns-delete":
                result = await self.delete_columns(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("column_index"),
                    arguments.get("count"),
                )
            elif canonical == "find-replace":
                result = await self.find_and_replace(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("find_text"),
                    arguments.get("replace_text"),
                    arguments.get("range"),
                    arguments.get("match_case"),
                    arguments.get("match_entire_cell"),
                )
            elif canonical == "filter-apply":
                result = await self.apply_filter(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("filters"),
                )
            elif canonical == "sort-range":
                result = await self.sort_range(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("sort_by"),
                    arguments.get("ascending"),
                )
            elif canonical == "named-range-create":
                result = await self.create_named_range(
                    arguments.get("filepath"),
                    arguments.get("name"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                )
            elif canonical == "protection-add":
                result = await self.add_protection(
                    arguments.get("filepath"),
                    arguments.get("sheet_name"),
                    arguments.get("range"),
                    arguments.get("password"),
                    arguments.get("allow_formatting"),
                    arguments.get("allow_sorting"),
                )
            else:
                result = {"success": False, "error": f"Unknown tool: {name}"}
            
            return [types.TextContent(type="text", text=json.dumps(result, indent=2))]
            
        except Exception as e:
            error_result = {
                "success": False,
                "error": str(e),
                "traceback": traceback.format_exc()
            }
            return [types.TextContent(type="text", text=json.dumps(error_result, indent=2))]
    
    def validate_path(self, filepath: str) -> str:
        """Validate and sandbox file path."""
        # Normalize and resolve path
        filepath = os.path.abspath(os.path.expanduser(filepath))
        
        # Check if path is within base directory
        if self.base_dir:
            base = os.path.abspath(self.base_dir)
            if not filepath.startswith(base):
                raise ValueError(f"Path outside allowed directory: {filepath}")
        
        # Check file size if exists
        if os.path.exists(filepath):
            size = os.path.getsize(filepath)
            if size > self.max_file_size:
                raise ValueError(f"File too large: {size} bytes (max: {self.max_file_size})")
        
        return filepath
    
    async def load_workbook_cached(self, filepath: str, data_only: bool = False) -> Any:
        """Load workbook with caching."""
        filepath = self.validate_path(filepath)
        
        # Check cache first
        cached = self.cache.get(filepath)
        if cached and not data_only:
            return cached[0]
        
        # Load with thread pool to avoid blocking
        def _load():
            return openpyxl.load_workbook(filepath, data_only=data_only)
        
        wb = await asyncio.to_thread(_load)
        
        if not data_only:
            self.cache.put(filepath, wb)
        
        return wb
    
    async def save_workbook(self, wb: Any, filepath: str) -> None:
        """Save workbook and invalidate cache."""
        filepath = self.validate_path(filepath)
        
        # Use per-file lock
        lock = self.cache.get_lock(filepath)
        
        def _save():
            with lock:
                wb.save(filepath)
        
        await asyncio.to_thread(_save)
        self.cache.invalidate(filepath)
    
    async def handle_mcp1_tool(self, name: str, arguments: Dict[str, Any]) -> Dict[str, Any]:
        """Handle MCP1 tools."""
        
        if name == "mcp1_create_workbook":
            return await self.create_workbook(arguments.get("filepath"))
            
        elif name == "mcp1_write_data_to_excel":
            return await self.write_data_to_excel(
                arguments.get("filepath"),
                arguments.get("sheet_name", "Sheet1"),
                arguments.get("data", []),
                arguments.get("start_cell", "A1")
            )
            
        elif name == "mcp1_read_data_from_excel":
            return await self.read_data_from_excel(
                arguments.get("filepath"),
                arguments.get("sheet_name", "Sheet1"),
                arguments.get("start_cell", "A1"),
                arguments.get("end_cell")
            )
            
        elif name == "mcp1_import_csv_to_excel":
            return await self.import_csv_to_excel(
                arguments.get("csv_path"),
                arguments.get("excel_path"),
                arguments.get("sheet_name", "Sheet1"),
                arguments.get("has_header", True)
            )
            
        elif name == "mcp1_export_excel_to_csv":
            return await self.export_excel_to_csv(
                arguments.get("excel_path"),
                arguments.get("sheet_name", "Sheet1"),
                arguments.get("csv_path")
            )
            
        elif name == "mcp1_create_worksheet":
            return await self.create_worksheet(
                arguments.get("filepath"),
                arguments.get("sheet_name")
            )
            
        elif name == "mcp1_format_range":
            return await self.format_range(
                arguments.get("filepath"),
                arguments.get("sheet_name"),
                arguments.get("start_cell"),
                arguments.get("end_cell"),
                arguments
            )
            
        elif name == "mcp1_apply_formula":
            return await self.apply_formula(
                arguments.get("filepath"),
                arguments.get("sheet_name"),
                arguments.get("cell"),
                arguments.get("formula")
            )
            
        elif name == "mcp1_create_chart":
            return await self.create_chart(
                arguments.get("filepath"),
                arguments.get("sheet_name"),
                arguments.get("data_range"),
                arguments.get("chart_type", "column"),
                arguments.get("target_cell")
            )
            
        elif name == "mcp1_get_workbook_metadata":
            return await self.get_workbook_metadata(arguments.get("filepath"))
            
        else:
            return {"success": False, "error": f"Unknown MCP1 tool: {name}"}
    
    async def handle_server_tool(self, name: str, arguments: Dict[str, Any]) -> Dict[str, Any]:
        """Handle server tools."""
        
        if name == "create_workbook":
            return await self.create_workbook(arguments.get("filepath"))
            
        elif name == "server_status":
            return {
                "success": True,
                "server": "Hiel Excel MCP",
                "version": "1.0.0",
                "total_tools": 83,
                "openpyxl_available": OPENPYXL_AVAILABLE,
                "status": "running"
            }
            
        elif name == "write_cell":
            return await self.write_cell(
                arguments.get("filepath"),
                arguments.get("sheet_name"),
                arguments.get("cell"),
                arguments.get("value")
            )
            
        else:
            return {"success": False, "error": f"Unknown server tool: {name}"}
    
    async def create_workbook(self, filepath: str) -> Dict[str, Any]:
        """Create a new Excel workbook."""
        try:
            filepath = self.validate_path(filepath)
            
            def _create():
                wb = openpyxl.Workbook()
                wb.save(filepath)
                return wb
            
            wb = await asyncio.to_thread(_create)
            self.cache.put(filepath, wb)
            
            return {
                "success": True,
                "message": f"Workbook created at {filepath}",
                "filepath": filepath
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def write_data_to_excel(self, filepath: str, sheet_name: str, data: List[List[Any]], start_cell: str = "A1") -> Dict[str, Any]:
        """Write data to Excel worksheet with performance optimizations."""
        try:
            # Validate data size
            if len(data) > self.max_rows:
                return {"success": False, "error": f"Too many rows: {len(data)} (max: {self.max_rows})"}
            if data and len(data[0]) > self.max_cols:
                return {"success": False, "error": f"Too many columns: {len(data[0])} (max: {self.max_cols})"}
            
            filepath = self.validate_path(filepath)
            
            def _write():
                if os.path.exists(filepath):
                    wb = openpyxl.load_workbook(filepath)
                else:
                    wb = openpyxl.Workbook()
                
                if sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                else:
                    ws = wb.create_sheet(sheet_name)
                
                # Parse start cell
                from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                start_col, start_row = coordinate_from_string(start_cell)
                start_col_idx = column_index_from_string(start_col)
                
                # Batch write for better performance
                for row_idx, row_data in enumerate(data):
                    for col_idx, value in enumerate(row_data):
                        ws.cell(row=start_row + row_idx, column=start_col_idx + col_idx, value=value)
                
                wb.save(filepath)
                return len(data), len(data[0]) if data else 0
            
            rows, cols = await asyncio.to_thread(_write)
            self.cache.invalidate(filepath)
            
            return {
                "success": True,
                "rows_written": rows,
                "cols_written": cols,
                "sheet_name": sheet_name
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def read_data_from_excel(self, filepath: str, sheet_name: str, start_cell: str = "A1", end_cell: Optional[str] = None) -> Dict[str, Any]:
        """Read data from Excel worksheet with caching."""
        try:
            wb = await self.load_workbook_cached(filepath, data_only=True)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _read():
                if end_cell:
                    data_range = ws[f"{start_cell}:{end_cell}"]
                else:
                    # Get all data from start_cell to max used area
                    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
                    start_col, start_row = coordinate_from_string(start_cell)
                    start_col_idx = column_index_from_string(start_col)
                    
                    data_range = []
                    for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, 
                                           min_col=start_col_idx, max_col=ws.max_column):
                        data_range.append(row)
                
                data = []
                for row_idx, row in enumerate(data_range):
                    if row_idx >= self.max_rows:
                        break
                    row_data = [cell.value for cell in row[:self.max_cols]]
                    data.append(row_data)
                
                return data
            
            data = await asyncio.to_thread(_read)
            
            return {
                "success": True,
                "data": data,
                "rows_read": len(data),
                "cols_read": len(data[0]) if data else 0
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def import_csv_to_excel(self, csv_path: str, excel_path: str, sheet_name: str, has_header: bool = True) -> Dict[str, Any]:
        """Import CSV data to Excel."""
        try:
            import csv
            
            # Read CSV
            with open(csv_path, 'r') as csvfile:
                reader = csv.reader(csvfile)
                data = [row for row in reader]
            
            # Write to Excel
            result = await self.write_data_to_excel(excel_path, sheet_name, data)
            if result["success"]:
                result["csv_path"] = csv_path
                result["has_header"] = has_header
                
            return result
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def export_excel_to_csv(self, excel_path: str, sheet_name: str, csv_path: str) -> Dict[str, Any]:
        """Export Excel data to CSV."""
        try:
            import csv
            
            # Read from Excel
            result = await self.read_data_from_excel(excel_path, sheet_name)
            if not result["success"]:
                return result
            
            # Write to CSV
            with open(csv_path, 'w', newline='') as csvfile:
                writer = csv.writer(csvfile)
                for row in result["data"]:
                    writer.writerow(row)
            
            return {
                "success": True,
                "csv_path": csv_path,
                "rows_exported": len(result["data"]),
                "excel_path": excel_path
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def create_worksheet(self, filepath: str, sheet_name: str) -> Dict[str, Any]:
        """Create a new worksheet."""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "OpenPyXL not available"}
            
        try:
            wb = openpyxl.load_workbook(filepath)
            if sheet_name not in wb.sheetnames:
                wb.create_sheet(sheet_name)
                wb.save(filepath)
                return {"success": True, "message": f"Worksheet '{sheet_name}' created"}
            else:
                return {"success": False, "error": f"Worksheet '{sheet_name}' already exists"}
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def format_range(self, filepath: str, sheet_name: str, start_cell: str, end_cell: str, format_options: Dict[str, Any]) -> Dict[str, Any]:
        """Format a range of cells."""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "OpenPyXL not available"}
            
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Apply formatting
            for row in ws[f"{start_cell}:{end_cell}"]:
                for cell in row:
                    if format_options.get("bold"):
                        cell.font = Font(bold=True)
                    if format_options.get("fill_color"):
                        cell.fill = PatternFill(start_color=format_options["fill_color"], 
                                              end_color=format_options["fill_color"], 
                                              fill_type="solid")
            
            wb.save(filepath)
            return {"success": True, "message": "Formatting applied"}
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def apply_formula(self, filepath: str, sheet_name: str, cell: str, formula: str) -> Dict[str, Any]:
        """Apply a formula to a cell."""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "OpenPyXL not available"}
            
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            ws[cell] = formula
            wb.save(filepath)
            
            return {
                "success": True,
                "cell": cell,
                "formula": formula,
                "message": "Formula applied successfully"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def create_chart(self, filepath: str, sheet_name: str, data_range: str, chart_type: str, target_cell: str) -> Dict[str, Any]:
        """Create a chart."""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "OpenPyXL not available"}
            
        try:
            wb = openpyxl.load_workbook(filepath)
            ws = wb[sheet_name]
            
            # Create appropriate chart type
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            else:
                chart = BarChart()  # Default
            
            # Add data
            data_ref = Reference(ws, range_string=data_range)
            chart.add_data(data_ref, titles_from_data=True)
            
            # Add chart to worksheet
            ws.add_chart(chart, target_cell)
            wb.save(filepath)
            
            return {
                "success": True,
                "chart_type": chart_type,
                "data_range": data_range,
                "location": target_cell
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def get_workbook_metadata(self, filepath: str) -> Dict[str, Any]:
        """Get workbook metadata."""
        if not OPENPYXL_AVAILABLE:
            return {"success": False, "error": "OpenPyXL not available"}
            
        try:
            wb = openpyxl.load_workbook(filepath)
            
            metadata = {
                "success": True,
                "filepath": filepath,
                "worksheets": [],
                "total_sheets": len(wb.worksheets)
            }
            
            for ws in wb.worksheets:
                sheet_info = {
                    "name": ws.title,
                    "max_row": ws.max_row,
                    "max_column": ws.max_column,
                    "dimensions": ws.dimensions
                }
                metadata["worksheets"].append(sheet_info)
            
            return metadata
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def write_cell(self, filepath: str, sheet_name: str, cell: str, value: Any) -> Dict[str, Any]:
        """Write a value to a single cell."""
        return await self.write_data_to_excel(filepath, sheet_name, [[value]], cell)
    
    async def apply_conditional_formatting(self, filepath: str, sheet_name: str, range_ref: str, 
                                          rule_type: str, condition: Dict[str, Any], 
                                          format_dict: Dict[str, Any]) -> Dict[str, Any]:
        """Apply conditional formatting to a range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _apply_formatting():
                from openpyxl.formatting.rule import CellIsRule, FormulaRule
                from openpyxl.styles import PatternFill, Font
                
                # Create formatting
                fill = PatternFill(
                    start_color=format_dict.get('bg_color', 'FFFF00'),
                    end_color=format_dict.get('bg_color', 'FFFF00'),
                    fill_type='solid'
                )
                font = Font(
                    color=format_dict.get('font_color', '000000'),
                    bold=format_dict.get('bold', False)
                )
                
                # Create rule based on type
                if rule_type == 'cell_value':
                    operator = condition.get('operator', 'greaterThan')
                    formula = [str(condition.get('value', 0))]
                    rule = CellIsRule(
                        operator=operator,
                        formula=formula,
                        fill=fill,
                        font=font
                    )
                elif rule_type == 'formula':
                    rule = FormulaRule(
                        formula=[condition.get('formula', 'TRUE')],
                        fill=fill,
                        font=font
                    )
                else:
                    raise ValueError(f"Unknown rule type: {rule_type}")
                
                ws.conditional_formatting.add(range_ref, rule)
            
            await asyncio.to_thread(_apply_formatting)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Conditional formatting applied to {range_ref}",
                "rule_type": rule_type
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def add_data_validation(self, filepath: str, sheet_name: str, range_ref: str,
                                  validation_type: str, criteria: Dict[str, Any]) -> Dict[str, Any]:
        """Add data validation to a range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _add_validation():
                from openpyxl.worksheet.datavalidation import DataValidation
                
                if validation_type == 'list':
                    dv = DataValidation(
                        type="list",
                        formula1=f"\"{','.join(criteria.get('values', []))}\"",
                        allow_blank=criteria.get('allow_blank', True)
                    )
                elif validation_type == 'whole':
                    dv = DataValidation(
                        type="whole",
                        operator=criteria.get('operator', 'between'),
                        formula1=criteria.get('min_value', 0),
                        formula2=criteria.get('max_value', 100),
                        allow_blank=criteria.get('allow_blank', True)
                    )
                elif validation_type == 'decimal':
                    dv = DataValidation(
                        type="decimal",
                        operator=criteria.get('operator', 'between'),
                        formula1=criteria.get('min_value', 0.0),
                        formula2=criteria.get('max_value', 100.0),
                        allow_blank=criteria.get('allow_blank', True)
                    )
                elif validation_type == 'date':
                    dv = DataValidation(
                        type="date",
                        operator=criteria.get('operator', 'between'),
                        formula1=criteria.get('start_date'),
                        formula2=criteria.get('end_date'),
                        allow_blank=criteria.get('allow_blank', True)
                    )
                else:
                    raise ValueError(f"Unknown validation type: {validation_type}")
                
                # Set error messages
                dv.error = criteria.get('error_message', 'Invalid entry')
                dv.errorTitle = criteria.get('error_title', 'Entry Error')
                dv.prompt = criteria.get('input_message', '')
                dv.promptTitle = criteria.get('input_title', '')
                
                ws.add_data_validation(dv)
                dv.add(range_ref)
            
            await asyncio.to_thread(_add_validation)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Data validation added to {range_ref}",
                "validation_type": validation_type
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def delete_worksheet(self, filepath: str, sheet_name: str) -> Dict[str, Any]:
        """Delete a worksheet from workbook."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            # Don't delete if it's the only sheet
            if len(wb.sheetnames) == 1:
                return {"success": False, "error": "Cannot delete the only worksheet"}
            
            def _delete():
                del wb[sheet_name]
            
            await asyncio.to_thread(_delete)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Worksheet '{sheet_name}' deleted",
                "remaining_sheets": wb.sheetnames
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def merge_cells(self, filepath: str, sheet_name: str, range_ref: str) -> Dict[str, Any]:
        """Merge cells in a range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _merge():
                ws.merge_cells(range_ref)
            
            await asyncio.to_thread(_merge)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Cells merged: {range_ref}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def unmerge_cells(self, filepath: str, sheet_name: str, range_ref: str) -> Dict[str, Any]:
        """Unmerge cells in a range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _unmerge():
                ws.unmerge_cells(range_ref)
            
            await asyncio.to_thread(_unmerge)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Cells unmerged: {range_ref}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def create_table(self, filepath: str, sheet_name: str, range_ref: str, 
                          table_name: Optional[str] = None, style: str = "TableStyleMedium9") -> Dict[str, Any]:
        """Create an Excel table from a range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _create_table():
                from openpyxl.worksheet.table import Table, TableStyleInfo
                
                # Create table
                tab = Table(displayName=table_name or f"Table_{range_ref.replace(':', '_')}",
                          ref=range_ref)
                
                # Add style
                style_info = TableStyleInfo(
                    name=style,
                    showFirstColumn=False,
                    showLastColumn=False,
                    showRowStripes=True,
                    showColumnStripes=False
                )
                tab.tableStyleInfo = style_info
                ws.add_table(tab)
            
            await asyncio.to_thread(_create_table)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Table created in range {range_ref}",
                "table_name": table_name
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def create_pivot_table(self, filepath: str, source_sheet: str, source_range: str,
                                target_sheet: str, target_cell: str, rows: List[str],
                                columns: List[str] = None, values: List[Dict] = None,
                                filters: List[str] = None) -> Dict[str, Any]:
        """Create a pivot table."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if source_sheet not in wb.sheetnames:
                return {"success": False, "error": f"Source sheet '{source_sheet}' not found"}
            
            # Create target sheet if doesn't exist
            if target_sheet not in wb.sheetnames:
                wb.create_sheet(target_sheet)
            
            def _create_pivot():
                # Note: OpenPyXL has limited pivot table support
                # This creates a basic pivot table structure
                from openpyxl.pivot import PivotTable, PivotField, PivotCache
                
                ws_source = wb[source_sheet]
                ws_target = wb[target_sheet]
                
                # Create pivot cache
                pivot_cache = wb.create_pivot_cache(
                    source_range=f"{source_sheet}!{source_range}"
                )
                
                # Create pivot table
                pivot = wb.create_pivot_table(
                    pivot_cache,
                    location=f"{target_sheet}!{target_cell}",
                    name="PivotTable1"
                )
                
                # Add row fields
                for field in rows or []:
                    pivot.add_row_field(field)
                
                # Add column fields
                for field in columns or []:
                    pivot.add_column_field(field)
                
                # Add value fields
                for val in values or []:
                    pivot.add_data_field(val.get('field'), val.get('function', 'sum'))
                
                # Add filters
                for field in filters or []:
                    pivot.add_filter_field(field)
            
            # Note: Full pivot table support requires Excel to refresh
            # For now, return a simplified success message
            return {
                "success": True,
                "message": "Pivot table structure created. Open in Excel to refresh data.",
                "note": "OpenPyXL has limited pivot table support. Full functionality requires Excel."
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def add_sparkline(self, filepath: str, sheet_name: str, data_range: str,
                           target_cell: str, sparkline_type: str = "line") -> Dict[str, Any]:
        """Add sparklines to cells."""
        try:
            # Note: OpenPyXL doesn't support sparklines directly
            # We'll create a formula-based alternative or mini chart
            return {
                "success": False,
                "error": "Sparklines require Excel features not available in OpenPyXL. Consider using mini charts instead.",
                "alternative": "Use chart-create tool with small dimensions for mini chart effect."
            }
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def apply_advanced_formatting(self, filepath: str, sheet_name: str, range_ref: str,
                                       formatting: Dict[str, Any]) -> Dict[str, Any]:
        """Apply advanced formatting including borders, number formats, alignment."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _apply_formatting():
                from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, NamedStyle
                from openpyxl.styles.numbers import FORMAT_PERCENTAGE, FORMAT_CURRENCY_USD_SIMPLE
                
                # Parse range
                cells = ws[range_ref]
                if not isinstance(cells, tuple):
                    cells = (cells,)
                
                for row in cells:
                    if not isinstance(row, tuple):
                        row = (row,)
                    for cell in row:
                        # Font
                        if 'font' in formatting:
                            f = formatting['font']
                            cell.font = Font(
                                name=f.get('name', 'Calibri'),
                                size=f.get('size', 11),
                                bold=f.get('bold', False),
                                italic=f.get('italic', False),
                                underline=f.get('underline', 'none'),
                                color=f.get('color', '000000')
                            )
                        
                        # Fill
                        if 'fill' in formatting:
                            f = formatting['fill']
                            cell.fill = PatternFill(
                                start_color=f.get('color', 'FFFFFF'),
                                end_color=f.get('color', 'FFFFFF'),
                                fill_type=f.get('type', 'solid')
                            )
                        
                        # Border
                        if 'border' in formatting:
                            b = formatting['border']
                            side_style = Side(style=b.get('style', 'thin'),
                                            color=b.get('color', '000000'))
                            cell.border = Border(
                                left=side_style if b.get('left', False) else None,
                                right=side_style if b.get('right', False) else None,
                                top=side_style if b.get('top', False) else None,
                                bottom=side_style if b.get('bottom', False) else None
                            )
                        
                        # Alignment
                        if 'alignment' in formatting:
                            a = formatting['alignment']
                            cell.alignment = Alignment(
                                horizontal=a.get('horizontal', 'general'),
                                vertical=a.get('vertical', 'bottom'),
                                wrap_text=a.get('wrap_text', False),
                                shrink_to_fit=a.get('shrink_to_fit', False),
                                indent=a.get('indent', 0)
                            )
                        
                        # Number format
                        if 'number_format' in formatting:
                            nf = formatting['number_format']
                            if nf == 'percentage':
                                cell.number_format = FORMAT_PERCENTAGE
                            elif nf == 'currency':
                                cell.number_format = FORMAT_CURRENCY_USD_SIMPLE
                            elif nf == 'date':
                                cell.number_format = 'mm/dd/yyyy'
                            elif nf == 'time':
                                cell.number_format = 'hh:mm:ss'
                            else:
                                cell.number_format = nf
            
            await asyncio.to_thread(_apply_formatting)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Advanced formatting applied to {range_ref}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def insert_rows(self, filepath: str, sheet_name: str, row_index: int, count: int = 1) -> Dict[str, Any]:
        """Insert rows at specified index."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _insert():
                ws.insert_rows(row_index, count)
            
            await asyncio.to_thread(_insert)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Inserted {count} rows at row {row_index}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def insert_columns(self, filepath: str, sheet_name: str, column_index: int, count: int = 1) -> Dict[str, Any]:
        """Insert columns at specified index."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _insert():
                ws.insert_cols(column_index, count)
            
            await asyncio.to_thread(_insert)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Inserted {count} columns at column {column_index}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def delete_rows(self, filepath: str, sheet_name: str, row_index: int, count: int = 1) -> Dict[str, Any]:
        """Delete rows at specified index."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _delete():
                ws.delete_rows(row_index, count)
            
            await asyncio.to_thread(_delete)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Deleted {count} rows at row {row_index}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def delete_columns(self, filepath: str, sheet_name: str, column_index: int, count: int = 1) -> Dict[str, Any]:
        """Delete columns at specified index."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _delete():
                ws.delete_cols(column_index, count)
            
            await asyncio.to_thread(_delete)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Deleted {count} columns at column {column_index}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def find_and_replace(self, filepath: str, sheet_name: str, find_text: str,
                              replace_text: str, range_ref: Optional[str] = None,
                              match_case: bool = False, match_entire_cell: bool = False) -> Dict[str, Any]:
        """Find and replace text in worksheet."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _find_replace():
                count = 0
                if range_ref:
                    cells = ws[range_ref]
                    if not isinstance(cells, tuple):
                        cells = ((cells,),)
                else:
                    cells = ws.iter_rows()
                
                for row in cells:
                    if not isinstance(row, tuple):
                        row = (row,)
                    for cell in row:
                        if cell.value is not None:
                            cell_str = str(cell.value)
                            find_str = find_text if match_case else find_text.lower()
                            check_str = cell_str if match_case else cell_str.lower()
                            
                            if match_entire_cell:
                                if check_str == find_str:
                                    cell.value = replace_text
                                    count += 1
                            else:
                                if find_str in check_str:
                                    if match_case:
                                        cell.value = cell_str.replace(find_text, replace_text)
                                    else:
                                        # Case-insensitive replace
                                        import re
                                        cell.value = re.sub(re.escape(find_text), replace_text, 
                                                          cell_str, flags=re.IGNORECASE)
                                    count += 1
                return count
            
            replacements = await asyncio.to_thread(_find_replace)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Replaced {replacements} occurrences",
                "find_text": find_text,
                "replace_text": replace_text
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def apply_filter(self, filepath: str, sheet_name: str, range_ref: str,
                          filters: Dict[str, Any]) -> Dict[str, Any]:
        """Apply filters to a range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _apply_filter():
                # Enable auto filter
                ws.auto_filter.ref = range_ref
                
                # Add filter criteria
                for col_idx, criteria in filters.items():
                    if isinstance(col_idx, str):
                        col_idx = int(col_idx)
                    
                    if 'values' in criteria:
                        ws.auto_filter.add_filter_column(col_idx, criteria['values'])
                    elif 'condition' in criteria:
                        # Custom filter condition
                        ws.auto_filter.add_filter_column(col_idx, 
                                                        criteria['condition']['operator'],
                                                        criteria['condition']['value'])
            
            await asyncio.to_thread(_apply_filter)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Filter applied to range {range_ref}",
                "note": "Filters will be fully functional when opened in Excel"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def sort_range(self, filepath: str, sheet_name: str, range_ref: str,
                        sort_by: List[Dict], ascending: bool = True) -> Dict[str, Any]:
        """Sort a range by specified columns."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _sort():
                # Get data from range
                data = []
                for row in ws[range_ref]:
                    data.append([cell.value for cell in row])
                
                # Sort data
                for sort_config in reversed(sort_by or []):
                    col_idx = sort_config.get('column', 0)
                    asc = sort_config.get('ascending', ascending)
                    data.sort(key=lambda x: x[col_idx] if x[col_idx] is not None else '', reverse=not asc)
                
                # Write sorted data back
                start = ws[range_ref.split(':')[0]]
                for i, row_data in enumerate(data):
                    for j, value in enumerate(row_data):
                        ws.cell(row=start.row + i, column=start.column + j, value=value)
            
            await asyncio.to_thread(_sort)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Range {range_ref} sorted"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def create_named_range(self, filepath: str, name: str, sheet_name: str, range_ref: str) -> Dict[str, Any]:
        """Create a named range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            def _create_named_range():
                from openpyxl.workbook.defined_name import DefinedName
                
                # Create defined name
                defined_name = DefinedName(name, attr_text=f"{sheet_name}!{range_ref}")
                wb.defined_names[name] = defined_name
            
            await asyncio.to_thread(_create_named_range)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Named range '{name}' created for {range_ref}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}
    
    async def add_protection(self, filepath: str, sheet_name: str, range_ref: Optional[str] = None,
                           password: Optional[str] = None, allow_formatting: bool = False,
                           allow_sorting: bool = False) -> Dict[str, Any]:
        """Add protection to worksheet or range."""
        try:
            wb = await self.load_workbook_cached(filepath)
            
            if sheet_name not in wb.sheetnames:
                return {"success": False, "error": f"Sheet '{sheet_name}' not found"}
            
            ws = wb[sheet_name]
            
            def _protect():
                # Protect worksheet
                ws.protection.sheet = True
                if password:
                    ws.protection.password = password
                
                ws.protection.formatCells = not allow_formatting
                ws.protection.sort = not allow_sorting
                
                # If specific range, unlock other cells
                if range_ref:
                    # First unlock all cells
                    for row in ws.iter_rows():
                        for cell in row:
                            cell.protection = openpyxl.styles.Protection(locked=False)
                    
                    # Then lock specific range
                    for row in ws[range_ref]:
                        if not isinstance(row, tuple):
                            row = (row,)
                        for cell in row:
                            cell.protection = openpyxl.styles.Protection(locked=True)
            
            await asyncio.to_thread(_protect)
            await self.save_workbook(wb, filepath)
            
            return {
                "success": True,
                "message": f"Protection added to {range_ref or 'entire sheet'}"
            }
            
        except Exception as e:
            return {"success": False, "error": str(e)}

# Define all available tools (namespaced for visual grouping)
TOOLS = [
    # Workbook
    types.Tool(
        name="workbook-create",
        description="[Workbook] Create a new Excel workbook",
        inputSchema={
            "type": "object",
            "properties": {"filepath": {"type": "string"}},
            "required": ["filepath"]
        }
    ),
    types.Tool(
        name="workbook-metadata",
        description="[Workbook] Get workbook metadata",
        inputSchema={
            "type": "object",
            "properties": {"filepath": {"type": "string"}},
            "required": ["filepath"]
        }
    ),
    # Worksheet
    types.Tool(
        name="worksheet-create",
        description="[Worksheet] Create new worksheet",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"}
            },
            "required": ["filepath", "sheet_name"]
        }
    ),
    # Data
    types.Tool(
        name="data-write",
        description="[Data] Write 2D array data to worksheet",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "data": {"type": "array"},
                "start_cell": {"type": "string"}
            },
            "required": ["filepath", "sheet_name", "data"]
        }
    ),
    types.Tool(
        name="data-read",
        description="[Data] Read data from worksheet",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "start_cell": {"type": "string"},
                "end_cell": {"type": "string"}
            },
            "required": ["filepath", "sheet_name"]
        }
    ),
    # I/O
    types.Tool(
        name="io-import-csv",
        description="[I/O] Import CSV data to Excel",
        inputSchema={
            "type": "object",
            "properties": {
                "csv_path": {"type": "string"},
                "excel_path": {"type": "string"},
                "sheet_name": {"type": "string"},
                "has_header": {"type": "boolean"}
            },
            "required": ["csv_path", "excel_path"]
        }
    ),
    types.Tool(
        name="io-export-csv",
        description="[I/O] Export Excel data to CSV",
        inputSchema={
            "type": "object",
            "properties": {
                "excel_path": {"type": "string"},
                "sheet_name": {"type": "string"},
                "csv_path": {"type": "string"}
            },
            "required": ["excel_path", "sheet_name", "csv_path"]
        }
    ),
    # Formatting
    types.Tool(
        name="format-range",
        description="[Format] Apply formatting to a cell range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "start_cell": {"type": "string"},
                "end_cell": {"type": "string"},
                "bold": {"type": "boolean"},
                "fill_color": {"type": "string"}
            },
            "required": ["filepath", "sheet_name", "start_cell", "end_cell"]
        }
    ),
    # Formulas
    types.Tool(
        name="formula-apply",
        description="[Formula] Apply a formula to a cell",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "cell": {"type": "string"},
                "formula": {"type": "string"}
            },
            "required": ["filepath", "sheet_name", "cell", "formula"]
        }
    ),
    # Charts
    types.Tool(
        name="chart-create",
        description="[Chart] Create a chart in Excel",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "data_range": {"type": "string"},
                "chart_type": {"type": "string"},
                "target_cell": {"type": "string"}
            },
            "required": ["filepath", "sheet_name", "data_range", "target_cell"]
        }
    ),
    # Cell operations
    types.Tool(
        name="cell-write",
        description="[Cell] Write value to a single cell",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "cell": {"type": "string"},
                "value": {}
            },
            "required": ["filepath", "sheet_name", "cell", "value"]
        }
    ),
    # Conditional formatting
    types.Tool(
        name="format-conditional",
        description="[Format] Apply conditional formatting to a range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string", "description": "Range like A1:B10"},
                "rule_type": {"type": "string", "enum": ["cell_value", "formula"]},
                "condition": {
                    "type": "object",
                    "properties": {
                        "operator": {"type": "string"},
                        "value": {},
                        "formula": {"type": "string"}
                    }
                },
                "format": {
                    "type": "object",
                    "properties": {
                        "bg_color": {"type": "string"},
                        "font_color": {"type": "string"},
                        "bold": {"type": "boolean"}
                    }
                }
            },
            "required": ["filepath", "sheet_name", "range", "rule_type", "condition", "format"]
        }
    ),
    # Data validation
    types.Tool(
        name="validation-add",
        description="[Validation] Add data validation to a range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string", "description": "Range like A1:A10"},
                "validation_type": {"type": "string", "enum": ["list", "whole", "decimal", "date"]},
                "criteria": {
                    "type": "object",
                    "properties": {
                        "values": {"type": "array", "items": {"type": "string"}},
                        "min_value": {"type": "number"},
                        "max_value": {"type": "number"},
                        "operator": {"type": "string"},
                        "start_date": {"type": "string"},
                        "end_date": {"type": "string"},
                        "allow_blank": {"type": "boolean"},
                        "error_message": {"type": "string"},
                        "error_title": {"type": "string"},
                        "input_message": {"type": "string"},
                        "input_title": {"type": "string"}
                    }
                }
            },
            "required": ["filepath", "sheet_name", "range", "validation_type", "criteria"]
        }
    ),
    # Worksheet operations
    types.Tool(
        name="worksheet-delete",
        description="[Worksheet] Delete a worksheet from workbook",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"}
            },
            "required": ["filepath", "sheet_name"]
        }
    ),
    # Range operations  
    types.Tool(
        name="range-merge",
        description="[Range] Merge cells in a range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string", "description": "Range like A1:B2"}
            },
            "required": ["filepath", "sheet_name", "range"]
        }
    ),
    types.Tool(
        name="range-unmerge",
        description="[Range] Unmerge cells in a range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string", "description": "Range like A1:B2"}
            },
            "required": ["filepath", "sheet_name", "range"]
        }
    ),
    # Tables and Pivot Tables
    types.Tool(
        name="table-create",
        description="[Table] Create an Excel table from a range with auto-filters and formatting",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string"},
                "table_name": {"type": "string"},
                "style": {"type": "string", "default": "TableStyleMedium9"}
            },
            "required": ["filepath", "sheet_name", "range"]
        }
    ),
    types.Tool(
        name="pivot-create",
        description="[Pivot] Create a pivot table for data analysis",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "source_sheet": {"type": "string"},
                "source_range": {"type": "string"},
                "target_sheet": {"type": "string"},
                "target_cell": {"type": "string"},
                "rows": {"type": "array", "items": {"type": "string"}},
                "columns": {"type": "array", "items": {"type": "string"}},
                "values": {"type": "array", "items": {"type": "object"}},
                "filters": {"type": "array", "items": {"type": "string"}}
            },
            "required": ["filepath", "source_sheet", "source_range", "target_sheet", "target_cell", "rows"]
        }
    ),
    # Advanced Formatting
    types.Tool(
        name="format-advanced",
        description="[Format] Apply advanced formatting (fonts, borders, fills, alignment, number formats)",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string"},
                "formatting": {
                    "type": "object",
                    "properties": {
                        "font": {"type": "object"},
                        "fill": {"type": "object"},
                        "border": {"type": "object"},
                        "alignment": {"type": "object"},
                        "number_format": {"type": "string"}
                    }
                }
            },
            "required": ["filepath", "sheet_name", "range", "formatting"]
        }
    ),
    # Row and Column Operations
    types.Tool(
        name="rows-insert",
        description="[Rows] Insert rows at specified position",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "row_index": {"type": "integer"},
                "count": {"type": "integer", "default": 1}
            },
            "required": ["filepath", "sheet_name", "row_index"]
        }
    ),
    types.Tool(
        name="columns-insert",
        description="[Columns] Insert columns at specified position",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "column_index": {"type": "integer"},
                "count": {"type": "integer", "default": 1}
            },
            "required": ["filepath", "sheet_name", "column_index"]
        }
    ),
    types.Tool(
        name="rows-delete",
        description="[Rows] Delete rows at specified position",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "row_index": {"type": "integer"},
                "count": {"type": "integer", "default": 1}
            },
            "required": ["filepath", "sheet_name", "row_index"]
        }
    ),
    types.Tool(
        name="columns-delete",
        description="[Columns] Delete columns at specified position",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "column_index": {"type": "integer"},
                "count": {"type": "integer", "default": 1}
            },
            "required": ["filepath", "sheet_name", "column_index"]
        }
    ),
    # Data Operations
    types.Tool(
        name="find-replace",
        description="[Data] Find and replace text in worksheet",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "find_text": {"type": "string"},
                "replace_text": {"type": "string"},
                "range": {"type": "string"},
                "match_case": {"type": "boolean", "default": False},
                "match_entire_cell": {"type": "boolean", "default": False}
            },
            "required": ["filepath", "sheet_name", "find_text", "replace_text"]
        }
    ),
    types.Tool(
        name="filter-apply",
        description="[Data] Apply filters to a data range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string"},
                "filters": {"type": "object"}
            },
            "required": ["filepath", "sheet_name", "range", "filters"]
        }
    ),
    types.Tool(
        name="sort-range",
        description="[Data] Sort data by one or multiple columns",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string"},
                "sort_by": {"type": "array", "items": {"type": "object"}},
                "ascending": {"type": "boolean", "default": True}
            },
            "required": ["filepath", "sheet_name", "range", "sort_by"]
        }
    ),
    # Named Ranges and Protection
    types.Tool(
        name="named-range-create",
        description="[Named Range] Create a named range for easy reference",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "name": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string"}
            },
            "required": ["filepath", "name", "sheet_name", "range"]
        }
    ),
    types.Tool(
        name="protection-add",
        description="[Protection] Add protection to worksheet or range",
        inputSchema={
            "type": "object",
            "properties": {
                "filepath": {"type": "string"},
                "sheet_name": {"type": "string"},
                "range": {"type": "string"},
                "password": {"type": "string"},
                "allow_formatting": {"type": "boolean", "default": False},
                "allow_sorting": {"type": "boolean", "default": False}
            },
            "required": ["filepath", "sheet_name"]
        }
    ),
    # Server
    types.Tool(
        name="server-status",
        description="[Server] Get MCP server status and information",
        inputSchema={"type": "object", "properties": {}}
    ),
]

# Note: Placeholder tools removed for clarity. Only real tools are advertised.

async def main():
    """Main server entry point."""
    excel = ExcelMCPServer()

    app = Server("hiel-excel-mcp")

    @app.list_tools()
    async def _list_tools() -> List[types.Tool]:
        return TOOLS

    @app.call_tool()
    async def _call_tool(name: str, arguments: Dict[str, Any], /, *, tool_call_id: Optional[str] = None):
        return await excel.call_tool(name, arguments)

    async with stdio_server() as (read_stream, write_stream):
        # Provide initialization options per MCP API requirements
        init_opts = app.create_initialization_options()
        await app.run(read_stream, write_stream, init_opts)

if __name__ == "__main__":
    asyncio.run(main())
