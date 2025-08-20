"""
Workbook context management and caching system for performance optimization.

This module provides intelligent workbook caching and context management
to optimize performance when performing multiple operations on the same
Excel files.
"""

import threading
import time
from contextlib import contextmanager
from pathlib import Path
from typing import Dict, Optional, Any, Generator
from weakref import WeakValueDictionary
import logging

from openpyxl import Workbook, load_workbook
from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

logger = logging.getLogger(__name__)


class WorkbookContext:
    """
    Context manager for workbook operations with intelligent caching.
    
    Provides automatic workbook lifecycle management, caching for performance,
    and thread-safe operations.
    """
    
    def __init__(self, filepath: str, read_only: bool = False, data_only: bool = False):
        """
        Initialize workbook context.
        
        Args:
            filepath: Path to Excel file
            read_only: Whether to open in read-only mode
            data_only: Whether to load only data (no formulas)
        """
        self.filepath = Path(filepath).resolve()
        self.read_only = read_only
        self.data_only = data_only
        self.workbook: Optional[OpenpyxlWorkbook] = None
        self.last_accessed = time.time()
        self.access_count = 0
        self.is_dirty = False
        self._lock = threading.RLock()
    
    def __enter__(self) -> OpenpyxlWorkbook:
        """Enter context manager and return workbook."""
        return self.get_workbook()
    
    def __exit__(self, exc_type, exc_val, exc_tb):
        """Exit context manager and handle cleanup."""
        if exc_type is None and self.is_dirty and not self.read_only:
            self.save()
    
    def get_workbook(self) -> OpenpyxlWorkbook:
        """
        Get workbook instance, loading if necessary.
        
        Returns:
            Loaded workbook instance
        """
        with self._lock:
            if self.workbook is None:
                self._load_workbook()
            
            self.last_accessed = time.time()
            self.access_count += 1
            return self.workbook
    
    def _load_workbook(self):
        """Load workbook from file."""
        try:
            if self.filepath.exists() and self.filepath.stat().st_size > 0:
                try:
                    logger.debug(f"Loading workbook: {self.filepath}")
                    self.workbook = load_workbook(
                        filename=str(self.filepath),
                        read_only=self.read_only,
                        data_only=self.data_only
                    )
                except Exception as load_error:
                    # If file exists but can't be loaded (corrupted, empty, etc.)
                    # create new workbook instead
                    logger.warning(f"Could not load existing file {self.filepath}: {load_error}")
                    logger.debug(f"Creating new workbook: {self.filepath}")
                    self.workbook = Workbook()
                    self.is_dirty = True
            else:
                logger.debug(f"Creating new workbook: {self.filepath}")
                self.workbook = Workbook()
                self.is_dirty = True
        except Exception as e:
            logger.error(f"Failed to load workbook {self.filepath}: {e}")
            raise
    
    def save(self, filepath: Optional[str] = None):
        """
        Save workbook to file.
        
        Args:
            filepath: Optional different path to save to
        """
        if self.read_only:
            raise ValueError("Cannot save read-only workbook")
        
        with self._lock:
            if self.workbook is None:
                return
            
            save_path = Path(filepath) if filepath else self.filepath
            logger.debug(f"Saving workbook: {save_path}")
            
            try:
                # Ensure directory exists
                save_path.parent.mkdir(parents=True, exist_ok=True)
                self.workbook.save(str(save_path))
                self.is_dirty = False
            except Exception as e:
                logger.error(f"Failed to save workbook {save_path}: {e}")
                raise
    
    def mark_dirty(self):
        """Mark workbook as modified."""
        self.is_dirty = True
    
    def close(self):
        """Close workbook and release resources."""
        with self._lock:
            if self.workbook is not None:
                try:
                    self.workbook.close()
                except Exception as e:
                    logger.warning(f"Error closing workbook {self.filepath}: {e}")
                finally:
                    self.workbook = None
    
    def is_expired(self, max_age_seconds: int = 300) -> bool:
        """
        Check if context has expired based on last access time.
        
        Args:
            max_age_seconds: Maximum age in seconds before expiration
            
        Returns:
            True if context has expired
        """
        return time.time() - self.last_accessed > max_age_seconds
    
    @property
    def is_loaded(self) -> bool:
        """Check if workbook is currently loaded."""
        return self.workbook is not None


class WorkbookCache:
    """
    Thread-safe cache for workbook contexts with automatic cleanup.
    
    Manages multiple workbook contexts with intelligent caching,
    automatic expiration, and memory management.
    """
    
    def __init__(self, max_size: int = 10, max_age_seconds: int = 300):
        """
        Initialize workbook cache.
        
        Args:
            max_size: Maximum number of cached workbooks
            max_age_seconds: Maximum age before expiration
        """
        self.max_size = max_size
        self.max_age_seconds = max_age_seconds
        self._cache: Dict[str, WorkbookContext] = {}
        self._lock = threading.RLock()
        self._stats = {
            'hits': 0,
            'misses': 0,
            'evictions': 0,
            'total_accesses': 0
        }
    
    def get_context(self, filepath: str, read_only: bool = False, 
                   data_only: bool = False) -> WorkbookContext:
        """
        Get or create workbook context.
        
        Args:
            filepath: Path to Excel file
            read_only: Whether to open in read-only mode
            data_only: Whether to load only data
            
        Returns:
            WorkbookContext instance
        """
        cache_key = self._make_cache_key(filepath, read_only, data_only)
        
        with self._lock:
            self._stats['total_accesses'] += 1
            
            # Check if context exists and is valid
            if cache_key in self._cache:
                context = self._cache[cache_key]
                if not context.is_expired(self.max_age_seconds):
                    self._stats['hits'] += 1
                    logger.debug(f"Cache hit for {filepath}")
                    return context
                else:
                    # Remove expired context
                    self._remove_context(cache_key)
            
            # Create new context
            self._stats['misses'] += 1
            logger.debug(f"Cache miss for {filepath}")
            
            context = WorkbookContext(filepath, read_only, data_only)
            self._add_context(cache_key, context)
            return context
    
    def _make_cache_key(self, filepath: str, read_only: bool, data_only: bool) -> str:
        """Create cache key from parameters."""
        resolved_path = str(Path(filepath).resolve())
        return f"{resolved_path}|{read_only}|{data_only}"
    
    def _add_context(self, cache_key: str, context: WorkbookContext):
        """Add context to cache, evicting if necessary."""
        # Evict expired contexts first
        self._cleanup_expired()
        
        # Evict least recently used if at capacity
        if len(self._cache) >= self.max_size:
            self._evict_lru()
        
        self._cache[cache_key] = context
        logger.debug(f"Added context to cache: {cache_key}")
    
    def _remove_context(self, cache_key: str):
        """Remove context from cache."""
        if cache_key in self._cache:
            context = self._cache.pop(cache_key)
            context.close()
            logger.debug(f"Removed context from cache: {cache_key}")
    
    def _cleanup_expired(self):
        """Remove expired contexts from cache."""
        expired_keys = [
            key for key, context in self._cache.items()
            if context.is_expired(self.max_age_seconds)
        ]
        
        for key in expired_keys:
            self._remove_context(key)
            self._stats['evictions'] += 1
    
    def _evict_lru(self):
        """Evict least recently used context."""
        if not self._cache:
            return
        
        lru_key = min(
            self._cache.keys(),
            key=lambda k: self._cache[k].last_accessed
        )
        
        self._remove_context(lru_key)
        self._stats['evictions'] += 1
        logger.debug(f"Evicted LRU context: {lru_key}")
    
    def invalidate(self, filepath: str):
        """
        Invalidate all cached contexts for a file.
        
        Args:
            filepath: Path to Excel file
        """
        resolved_path = str(Path(filepath).resolve())
        
        with self._lock:
            keys_to_remove = [
                key for key in self._cache.keys()
                if key.startswith(resolved_path + "|")
            ]
            
            for key in keys_to_remove:
                self._remove_context(key)
                logger.debug(f"Invalidated cache for {filepath}")
    
    def clear(self):
        """Clear all cached contexts."""
        with self._lock:
            for context in self._cache.values():
                context.close()
            self._cache.clear()
            logger.debug("Cleared workbook cache")
    
    def get_stats(self) -> Dict[str, Any]:
        """
        Get cache statistics.
        
        Returns:
            Dictionary with cache statistics
        """
        with self._lock:
            hit_rate = (
                self._stats['hits'] / self._stats['total_accesses']
                if self._stats['total_accesses'] > 0 else 0
            )
            
            return {
                'size': len(self._cache),
                'max_size': self.max_size,
                'hits': self._stats['hits'],
                'misses': self._stats['misses'],
                'evictions': self._stats['evictions'],
                'total_accesses': self._stats['total_accesses'],
                'hit_rate': hit_rate,
                'contexts': [
                    {
                        'filepath': str(context.filepath),
                        'last_accessed': context.last_accessed,
                        'access_count': context.access_count,
                        'is_dirty': context.is_dirty,
                        'is_loaded': context.is_loaded
                    }
                    for context in self._cache.values()
                ]
            }


# Global cache instance
_global_cache = WorkbookCache()


@contextmanager
def workbook_context(filepath: str, read_only: bool = False, 
                    data_only: bool = False) -> Generator[OpenpyxlWorkbook, None, None]:
    """
    Context manager for workbook operations with caching.
    
    Args:
        filepath: Path to Excel file
        read_only: Whether to open in read-only mode
        data_only: Whether to load only data
        
    Yields:
        Workbook instance
        
    Example:
        with workbook_context('file.xlsx') as wb:
            ws = wb.active
            ws['A1'] = 'Hello'
    """
    context = _global_cache.get_context(filepath, read_only, data_only)
    
    try:
        with context as workbook:
            yield workbook
    except Exception:
        # Don't save on exception
        context.is_dirty = False
        raise


def get_cache_stats() -> Dict[str, Any]:
    """Get global cache statistics."""
    return _global_cache.get_stats()


def invalidate_cache(filepath: str):
    """Invalidate cache for specific file."""
    _global_cache.invalidate(filepath)


def clear_cache():
    """Clear entire cache."""
    _global_cache.clear()


def configure_cache(max_size: int = 10, max_age_seconds: int = 300):
    """
    Configure global cache settings.
    
    Args:
        max_size: Maximum number of cached workbooks
        max_age_seconds: Maximum age before expiration
    """
    global _global_cache
    _global_cache.clear()
    _global_cache = WorkbookCache(max_size, max_age_seconds)